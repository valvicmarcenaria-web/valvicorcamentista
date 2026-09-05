#!/usr/bin/env python3
"""
Funções para montar planilhas da Valvic com a cara da casa e menus suspensos que
funcionam em todo lugar. Importe daqui em vez de reescrever a formatação toda vez:

    import sys; sys.path.insert(0, 'ferramentas')
    from gerar_planilha import nova, cabecalho, titulo_secao, menu, largura, salvar

A LIÇÃO MAIS CARA DESTE ARQUIVO — menu suspenso:
O openpyxl aceita três formas de definir a lista de um menu suspenso, e só uma sobrevive
em todos os programas:

  1. Intervalo de outra aba  ("=Listas!$D$2:$D$17")  → o Google Sheets DESCARTA na importação
  2. Nome definido           ("=EQUIPE")             → importado de forma inconsistente
  3. Lista literal           ('"PIX,Boleto,Cartão"') → funciona em TODOS

Use sempre a forma 3, via a função `menu()`. Limites: 255 caracteres no total e nenhum
item pode conter vírgula ou aspas — a função confere isso e falha alto se passar.

Detalhe traiçoeiro do formato OOXML: `showDropDown=False` é o que FAZ APARECER a setinha.
Está invertido mesmo. A função já cuida disso.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = '0E2038'
NAVY2 = '16314F'
GOLD = 'C2A05A'
GOLDBG = 'F6EDD6'
INK = '1B2733'
CINZA = 'F2F4F7'

_fina = Side(style='thin', color='D6DBE2')
BORDA = Border(left=_fina, right=_fina, top=_fina, bottom=_fina)


def nova(nome_primeira_aba='Dados'):
    """Cria o arquivo já com a primeira aba nomeada."""
    wb = Workbook()
    wb.active.title = nome_primeira_aba
    return wb


def cabecalho(ws, colunas, linha=1, altura=26):
    """Escreve a linha de cabeçalho no navy da marca, com texto dourado."""
    for i, texto in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=i, value=texto)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.font = Font(name='Calibri', size=10, bold=True, color=GOLD)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDA
    ws.row_dimensions[linha].height = altura
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def titulo_secao(ws, texto, linha, ate_coluna, altura=22):
    """Faixa dourada de seção, mesclada de A até a coluna indicada."""
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ate_coluna)
    c = ws.cell(row=linha, column=1, value=texto)
    c.fill = PatternFill('solid', fgColor=GOLDBG)
    c.font = Font(name='Calibri', size=11, bold=True, color=NAVY2)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha].height = altura


def _literal(itens):
    """Monta a lista embutida no formato que todo programa entende."""
    for it in itens:
        assert ',' not in it, f'item com vírgula quebra a lista literal: {it!r}'
        assert '"' not in it, f'item com aspas quebra a lista literal: {it!r}'
    lit = '"' + ','.join(itens) + '"'
    assert len(lit) <= 255, (
        f'lista literal com {len(lit)} caracteres (máximo 255). '
        'Encurte os itens ou divida a lista em duas colunas.')
    return lit


def menu(ws, itens, intervalo, permitir_vazio=True):
    """Aplica um menu suspenso a um intervalo. Ex.: menu(ws, ['PIX','Boleto'], 'D2:D200')"""
    dv = DataValidation(type='list', formula1=_literal(itens),
                        allow_blank=permitir_vazio, showDropDown=False)
    dv.error = 'Escolha um valor da lista.'
    dv.errorTitle = 'Valor fora da lista'
    ws.add_data_validation(dv)
    dv.add(intervalo)
    return dv


def largura(ws, larguras):
    """largura(ws, [28, 14, 14, 40]) — em caracteres, na ordem das colunas."""
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def zebra(ws, primeira_linha, ultima_linha, ate_coluna):
    """Fundo alternado, para a leitura não escorregar de linha."""
    for r in range(primeira_linha, ultima_linha + 1):
        if (r - primeira_linha) % 2:
            for c in range(1, ate_coluna + 1):
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=CINZA)


def salvar(wb, caminho):
    wb.save(caminho)
    print(f'planilha salva: {caminho}')


if __name__ == '__main__':
    wb = nova('Exemplo')
    ws = wb.active
    titulo_secao(ws, 'EXEMPLO — apague esta aba', 1, 4)
    cabecalho(ws, ['Item', 'Fornecedor', 'Forma de pagamento', 'Observação'], linha=2)
    menu(ws, ['PIX', 'Boleto', 'Cartão', 'Transferência'], 'C3:C200')
    largura(ws, [30, 24, 22, 40])
    zebra(ws, 3, 20, 4)
    salvar(wb, '/tmp/exemplo-valvic.xlsx')
