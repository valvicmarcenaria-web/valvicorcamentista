#!/usr/bin/env python3
"""Apura o Contas a Pagar por VENCIMENTO e classifica em operação × variáveis.

Base de vencimento porque é a que reconcilia com o número que a empresa já usa:
agosto por vencimento dá R$ 221.208,64, exatamente a despesa total de agosto
informada pela direção. Competência e caixa aparecem no rodapé, para leitura.
"""
import collections
import datetime
import openpyxl

ARQ = ('/root/.claude/uploads/2544489f-df71-5f40-87c6-89025901a0cf/'
       '5a2953b6-Contas_a_Pagar_02_09_2026.xlsx')

# ── mapa de classificação ───────────────────────────────────────────────
FOLHA = {'Salários', 'Pró-labore', 'Adiantamento Salarial', 'Vale-Transporte',
         'Vale-Alimentação', 'Hora extra', 'INSS sobre Pró-labore - GPS'}
FROTA = {'Leasing - Veículos', 'Seguros de Veículos', 'Manutenção de Veículos',
         'Combustíveis'}
# Estrutura = só o que a empresa precisa para funcionar e operar.
ESTRUTURA = {'Aluguel', 'Energia Elétrica', 'Água e Saneamento', 'Taxa de Lixo',
             'Limpeza', 'Materiais de Limpeza e de Higiene',
             'Vigilância e Segurança Patrimonial', 'Materiais de Escritório',
             'Software / Licença de Uso', 'Software/ferramenta de gestao',
             'Honorários Advocatícios', 'Manutenção de Equipamentos', 'Manutenções',
             'Manutenção Predial', 'Ferramentas de Desgastes',
             'Computadores e Periféricos', 'Cursos e Treinamentos'}
IMPOSTO = {'Simples Nacional - DAS'}
# Financiamentos, dívida, comercial e o que não é nem operação nem obra.
DIVERSOS = {'Máquinas, Equipamentos e Instalações Industriais', 'Terrenos',
            'juros e multas', 'Marketing e Publicidade', 'Brindes para Clientes',
            'Lanches e Refeições', 'Viagens e Representações', 'Farmácia'}
RT = {'RT parceiros'}
COMISSAO = {'Comissao de obras', 'Comissões de Vendedores'}
LOGISTICA = {'carretos', 'Transportadoras', 'Transporte Urbano (táxi, Uber)'}
MATERIA = {'Materiais Aplicados na Produção', 'Corte de chapas'}
TERCEIRO = {'Servicos terceirizados', 'Servicos extras'}

# Lançamentos a ignorar: erro de digitação identificado pela direção.
# O pró-labore é fixo em R$ 10 mil para cada sócio; esta linha era duplicidade.
IGNORAR = {'pro labore jonathan - restante'}

# lançamentos sem categoria, classificados pela descrição
POR_DESCRICAO = [
    ('rt ', 'rt'),
    ('bigfer', 'materia'), ('compra 3', 'materia'), ('compra 4', 'materia'),
    ('mgv distribuidora', 'materia'),
    ('douglas marceneiro', 'terceiro'), ('rejane/acabamento', 'terceiro'),
    ('diarias', 'terceiro'), ('carreto', 'logistica'),
    ('conta garantia', 'diversos'), ('raizen', 'diversos'),
    ('eduzz', 'estrutura'), ('consulta protestos', 'diversos'),
    ('bondinho', 'diversos'),
]
GRUPO = {'folha': FOLHA, 'frota': FROTA, 'estrutura': ESTRUTURA, 'imposto': IMPOSTO,
         'diversos': DIVERSOS, 'rt': RT, 'comissao': COMISSAO, 'logistica': LOGISTICA,
         'materia': MATERIA, 'terceiro': TERCEIRO}
OPERACAO = ('folha', 'estrutura', 'frota')
VARIAVEIS = ('materia', 'rt', 'comissao', 'terceiro', 'logistica', 'imposto')
OUTROS = ('diversos',)
ROTULO = {'folha': 'Folha de pagamento', 'estrutura': 'Estrutura e administrativo',
          'frota': 'Frota', 'imposto': 'Impostos sobre a venda',
          'diversos': 'Financiamentos, comercial e diversos',
          'rt': 'RT de parceiros', 'comissao': 'Comissões', 'logistica': 'Logística',
          'materia': 'Matéria-prima', 'terceiro': 'Serviços terceirizados'}


def data(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    if isinstance(v, str) and len(v) == 10:
        dd, mm, aa = v.split('/')
        return datetime.date(int(aa), int(mm), int(dd))
    return None


def classificar(cat, desc):
    for g, cats in GRUPO.items():
        if cat in cats:
            return g
    d = (desc or '').lower()
    for chave, g in POR_DESCRICAO:
        if chave in d:
            return g
    return 'estrutura'          # sobra vai para estrutura, e o teste avisa se for muito


# Lançamentos a ignorar: erro de digitação identificado pela direção.
# O pró-labore é fixo em R$ 10 mil para cada sócio; esta linha era duplicidade.
IGNORAR = {'pro labore jonathan - restante'}


def ler():
    ws = openpyxl.load_workbook(ARQ, data_only=True)['data']
    h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    itens = []
    for r in range(2, ws.max_row + 1):
        g = lambda k: ws.cell(r, h[k]).value
        val = g('Valor')
        if not isinstance(val, (int, float)):
            continue
        cat = g('Categoria') or ''
        if (g('Descrição') or '').strip().lower() in IGNORAR:
            continue
        itens.append({'venc': data(g('Data_Vencimento')), 'comp': data(g('Data_competencia')),
                      'pag': data(g('Data_Pagamento')), 'cat': cat or '(sem categoria)',
                      'desc': g('Descrição') or '', 'status': g('Status'),
                      'grupo': classificar(cat, g('Descrição')), 'val': float(val)})
    return itens


def apurar(itens, ano_mes, campo='venc'):
    sub = [x for x in itens if x[campo] and x[campo].strftime('%Y-%m') == ano_mes]
    por_grupo = collections.Counter()
    por_cat = collections.defaultdict(collections.Counter)
    for x in sub:
        por_grupo[x['grupo']] += x['val']
        por_cat[x['grupo']][x['cat']] += x['val']
    return sub, por_grupo, por_cat


if __name__ == '__main__':
    itens = ler()
    for mes, rot in (('2026-08', 'AGOSTO'), ('2026-09', 'SETEMBRO')):
        sub, gr, cat = apurar(itens, mes)
        tot = sum(gr.values())
        print('═' * 72)
        print(f'{rot} · por vencimento · {len(sub)} lançamentos · TOTAL R$ {tot:,.2f}')
        for bloco, nome in (('OPERAÇÃO', OPERACAO), ('VARIÁVEIS', VARIAVEIS),
                            ('OUTROS', OUTROS)):
            sb = sum(gr[g] for g in nome)
            print(f'\n  {bloco}  R$ {sb:,.2f}   ({sb/tot*100:.1f}%)')
            for g in nome:
                if not gr[g]:
                    continue
                print(f'    {ROTULO[g]:32} {gr[g]:>12,.2f}   {gr[g]/tot*100:>5.1f}%')
                for c, v in sorted(cat[g].items(), key=lambda t: -t[1]):
                    print(f'        {c[:44]:46} {v:>11,.2f}')
        print()
    print('═' * 72)
    for campo, rot in (('venc', 'vencimento'), ('comp', 'competência'), ('pag', 'pagamento')):
        a = sum(x['val'] for x in itens if x[campo] and x[campo].strftime('%Y-%m') == '2026-08')
        s = sum(x['val'] for x in itens if x[campo] and x[campo].strftime('%Y-%m') == '2026-09')
        print(f'  base {rot:14} ago {a:>12,.2f}   set {s:>12,.2f}')
