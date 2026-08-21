#!/usr/bin/env python3
"""Gera o Controle de Custo Direto por Projeto da Valvic.

Uma lâmina por projeto (duplicável) com todo o custo direto e indireto em
ORÇADO x REALIZADO x DESVIO. As compras entram num livro de lançamentos —
uma linha por compra, com forma de pagamento e status — e as categorias se
somam sozinhas. Um painel geral consolida os projetos por INDIRECT.

Uso:  python3 gerar-custo-projeto.py
Saída: Valvic_Custo_por_Projeto.xlsx
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties

NAVY, NAVY2 = '0E2038', '16314F'
GOLD, GOLDS, GOLDBG = 'C2A05A', 'D8BD80', 'F6EDD6'
INK, MUTED = '1B2733', '6C7785'
LINE, LINE2 = 'E8E3D8', 'DFDACD'
OK, BLUE, RED, AMBER = '2F7D4F', '2F5D8C', 'B0413F', 'B57A16'
OKBG, REDBG, AMBBG, BLUEBG = 'E2F0E7', 'FBE7E6', 'FBF1DC', 'E7EEF6'
INPUT, WHITE, CALC = 'FFF9E3', 'FFFFFF', 'F4F6F8'

F = 'Arial'
def font(sz=10, b=False, c=INK, i=False):
    return Font(name=F, size=sz, bold=b, color=c, italic=i)
def fill(c): return PatternFill('solid', fgColor=c)
def side(c=LINE, st='thin'): return Side(style=st, color=c)
GRID = Border(bottom=side(LINE), left=side(LINE), right=side(LINE))
BOTTOM2 = Border(bottom=side(LINE2))
CTR = Alignment(horizontal='center', vertical='center')
CTRW = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
LEFTI = Alignment(horizontal='left', vertical='center', indent=1)
LEFTIW = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
RIGHT = Alignment(horizontal='right', vertical='center', indent=1)
MOEDA, MOEDA0, PCT0, PCT1, DATA = 'R$ #,##0.00', 'R$ #,##0', '0%', '0.0%', 'DD/MM/YYYY'
DIAS = '0 "dias"'

# só quem recebe comissão de produção/montagem/coordenação
EQUIPE = ['Deivson', 'Samuel', 'Cezar', 'Jackson', 'Jomar', 'Joelson',
          'Jonathan Godoy', 'Terceiro / avulso']
VENDEDORES = ['Jonathan', 'Vitor', 'Indicação', 'Arquiteto parceiro', 'Outro']
CAUSAS = ['Erro de projeto', 'Erro de medição', 'Erro de produção', 'Erro de montagem',
          'Falha de material', 'Dano no transporte', 'Mudança pedida pelo cliente',
          'Falha de fornecedor', 'Outro']
PAGAMENTOS = ['PIX', 'Boleto', 'Cartão de crédito', 'Cartão parcelado', 'Dinheiro',
              'Transferência', 'Faturado 30 dias', 'A combinar']
STATUS = ['A comprar', 'Comprado (a pagar)', 'Pago']
# nomes curtos e sem vírgula: cabem no menu suspenso e na largura da coluna.
# o "o que entra" de cada uma fica explicado na aba Listas.
GRUPOS = [
    ('Material', ['MDF e MDP', 'Fita de borda', 'Ferragens', 'Vidros e espelhos',
                  'Esquadrias', 'Lâmina natural', 'Consumíveis']),
    ('Serviços terceirizados', ['Acabamento', 'Serralheria', 'Vidraceiro',
                                'Outro terceirizado']),
    ('Logística', ['Uber e aplicativo', 'Carreto e entrega', 'Deslocamento da equipe',
                   'Frete de material', 'Estacionamento e pedágio']),
]
CATEGORIAS_COMPRA = [c for _, cats in GRUPOS for c in cats]
DESCR_CATEGORIA = {
    'MDF e MDP': 'chapas cruas e revestidas, por cor e espessura',
    'Fita de borda': 'fita de borda de todas as cores e espessuras',
    'Ferragens': 'corrediças, dobradiças, puxadores, pistões, suportes',
    'Vidros e espelhos': 'vidro, espelho e cristal comprados prontos',
    'Esquadrias': 'perfis de alumínio, portas de perfil e box',
    'Lâmina natural': 'lâmina de madeira natural e compostos',
    'Consumíveis': 'parafuso, adesivo, tíner, estopa, lixa, disco, broca',
    'Acabamento': 'pintura, laca, verniz e envernizamento terceirizado',
    'Serralheria': 'estruturas metálicas feitas fora',
    'Vidraceiro': 'corte e instalação de vidro por terceiro',
    'Outro terceirizado': 'qualquer serviço feito fora da fábrica',
    'Uber e aplicativo': 'corrida de medição, conferência e visita',
    'Carreto e entrega': 'transporte do móvel pronto até a obra',
    'Deslocamento da equipe': 'ida e volta da equipe de montagem',
    'Frete de material': 'entrega de material do fornecedor até a fábrica',
    'Estacionamento e pedágio': 'estacionamento, pedágio e miudezas de rota',
}

# nomes definidos: a validação passa a apontar para um NOME, não para um
# intervalo de outra aba. É a forma que sobrevive à conversão para o Google
# Sheets — foi por isso que o menu de Categoria não apareceu lá.
NOMES = {'CATEGORIA_COMPRA': 'D', 'FORMA_PAGAMENTO': 'E', 'STATUS_COMPRA': 'F',
         'EQUIPE_COMISSAO': 'A', 'VENDEDOR': 'B', 'CAUSA_RETRABALHO': 'C'}

NCOL = 10
LP = {c: get_column_letter(c) for c in range(1, 61)}
W_FICHA = [30, 10, 14, 18, 8, 13, 18, 8, 13, 14]
N_AMB, N_COL, N_RETRAB, N_LANC = 12, 12, 12, 60

wb = openpyxl.Workbook()
wb.remove(wb.active)


def faixa_marca(ws, ncols, titulo, sub, linha=1):
    for i, (txt, fo, bg, alt) in enumerate([
            ('VALVIC MARCENARIA', Font(name=F, size=13, bold=True, color=WHITE), NAVY, 26),
            ('Vargas Decor Ltda   ·   CNPJ 17.269.304/0001-51   ·   Belo Horizonte / MG',
             Font(name=F, size=8, color='9FB0C4'), NAVY, 15),
            (titulo, Font(name=F, size=15, bold=True, color=WHITE), NAVY2, 28),
            (sub, Font(name=F, size=8.5, color='7A5B17', italic=True), GOLDBG, 17)]):
        r = linha + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1, txt); c.font = fo; c.alignment = LEFTI
        ws.row_dimensions[r].height = alt
        for cc in range(1, ncols + 1):
            ws.cell(r, cc).fill = fill(bg)
    return linha + 4


def titulo_secao(ws, row, ncols, texto, nota=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, ('  ' + texto.upper()) + (f'          {nota}' if nota else ''))
    c.font = Font(name=F, size=9, bold=True, color=NAVY); c.alignment = LEFT
    for cc in range(1, ncols + 1):
        cel = ws.cell(row, cc)
        cel.fill = fill(GOLDBG); cel.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 20
    return row + 1


def cab(ws, row, pares, alt=26):
    for c0, span, txt in pares:
        if span > 1:
            ws.merge_cells(start_row=row, start_column=c0, end_row=row,
                           end_column=c0 + span - 1)
        c = ws.cell(row, c0, txt)
        c.font = Font(name=F, size=8.5, bold=True, color=WHITE); c.alignment = CTRW
        for k in range(span):
            cel = ws.cell(row, c0 + k); cel.fill = fill(NAVY2)
            cel.border = Border(left=side(NAVY2), right=side(NAVY2),
                                bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = alt
    return row + 1


def print_cfg(ws, area, retrato=False, margens=(0.4, 0.3, 0.4, 0.3)):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait' if retrato else 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.print_area = area
    l, r, t, b = margens
    ws.page_margins.left, ws.page_margins.right = l, r
    ws.page_margins.top, ws.page_margins.bottom = t, b
    ws.sheet_view.showGridLines = False


def dv(ws, formula, rng):
    d = DataValidation(type='list', formula1=formula, allow_blank=True,
                       showDropDown=False)
    d.showInputMessage = False
    d.showErrorMessage = True          # avisa quando o valor está fora da lista
    d.errorStyle = 'warning'
    d.errorTitle = 'Fora da lista'
    d.error = ('Este valor não está na lista. Você pode manter, mas os totais só '
               'somam o que estiver escrito exatamente como na aba Listas.')
    ws.add_data_validation(d); d.add(rng)


def bloco(ws, row, col, span, valor=None, *, f=None, bg=None, al=None, nf=None, bd=True):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row, col, valor)
    if f: c.font = f
    if al: c.alignment = al
    if nf: c.number_format = nf
    for k in range(span):
        cel = ws.cell(row, col + k)
        if bd: cel.border = GRID
        if bg: cel.fill = fill(bg)
    return c


F_ENT = Font(name=F, size=9.5, color=INK)
F_CALC = Font(name=F, size=9.5, color=NAVY2)
F_PCT = Font(name=F, size=9.5, bold=True, color='7A5B17')
F_SUB = Font(name=F, size=10, bold=True, color=WHITE)

# ══════════════ mapa de linhas — tudo que o Painel lê fica até a linha 24
R_ID1, R_ID2, R_ID3, R_ID4 = 6, 7, 8, 9
R_TOP, R_KPI_L, R_KPI_V = 11, 12, 13
R_RES_H = 15
CATEGORIAS = ['Custos de venda', 'Comissões operacionais', 'Material',
              'Serviços terceirizados', 'Logística', 'Retrabalho']
R_RES0 = 16
R_RESF = R_RES0 + len(CATEGORIAS) - 1          # 21
R_CUSTO_TOT = R_RESF + 1                       # 22
R_MC = R_CUSTO_TOT + 1                         # 23
R_ALERTA = R_MC + 1                            # 24
LIMITE_PAINEL = R_ALERTA

R_VEN_T, R_VEN_H, R_VENDA = 26, 27, 28
R_CV_T, R_CV_H = 30, 31
R_IMP, R_MAQ, R_TRX, R_CVEND, R_PROJ, R_RTP = 32, 33, 34, 35, 36, 37
R_CV_SUB, R_LIQ = 38, 39
R_AMB_T, R_AMB_H, R_AMB0 = 41, 42, 43
R_AMBF = R_AMB0 + N_AMB - 1
R_AMB_TOT = R_AMBF + 1
R_CO_T, R_CO_H = R_AMB_TOT + 2, R_AMB_TOT + 3
R_COORD, R_PRODC, R_MONTC, R_CO_SUB = R_CO_H + 1, R_CO_H + 2, R_CO_H + 3, R_CO_H + 4
R_CL_T, R_CL_H, R_CL0 = R_CO_SUB + 2, R_CO_SUB + 3, R_CO_SUB + 4
R_CLF = R_CL0 + N_COL - 1
R_CL_TOT = R_CLF + 1
# orçamento por categoria de compra, com subtotal por grupo
R_OC_T, R_OC_H = R_CL_TOT + 2, R_CL_TOT + 3
LINHAS_OC, R_SUBGRUPO = {}, {}
_r = R_OC_H + 1
for _g, _cats in GRUPOS:
    for _c in _cats:
        LINHAS_OC[_c] = _r; _r += 1
    R_SUBGRUPO[_g] = _r; _r += 1
R_OCF = _r - 1
R_RB_T, R_RB_H, R_RB0 = R_OCF + 2, R_OCF + 3, R_OCF + 4
R_RBF = R_RB0 + N_RETRAB - 1
R_RB_SUB = R_RBF + 1
R_LAN_T, R_LAN_H, R_LAN0 = R_RB_SUB + 2, R_RB_SUB + 3, R_RB_SUB + 4
R_LANF = R_LAN0 + N_LANC - 1
R_LAN_TOT = R_LANF + 1
R_NOTA = R_LAN_TOT + 2

VENDA = '$A$13'
SUBLINHAS = {'Custos de venda': R_CV_SUB, 'Comissões operacionais': R_CO_SUB,
             'Material': R_SUBGRUPO['Material'],
             'Serviços terceirizados': R_SUBGRUPO['Serviços terceirizados'],
             'Logística': R_SUBGRUPO['Logística'], 'Retrabalho': R_RB_SUB}
# faixas do livro de lançamentos
LAN_CAT = f'$B${R_LAN0}:$B${R_LANF}'
LAN_VAL = f'$E${R_LAN0}:$E${R_LANF}'
LAN_STA = f'$I${R_LAN0}:$I${R_LANF}'
print(f'ficha: resumo {R_RES0}-{R_MC} · venda {R_VENDA} · líquido {R_LIQ} · '
      f'ambientes {R_AMB0}-{R_AMBF} · categorias {R_OC_H+1}-{R_OCF} · '
      f'lançamentos {R_LAN0}-{R_LANF} · fim {R_NOTA}')


AMB_PROD_Q = f'$D${R_AMB0}:$D${R_AMBF}'
AMB_PROD_V = f'$F${R_AMB0}:$F${R_AMBF}'
AMB_MONT_Q = f'$G${R_AMB0}:$G${R_AMBF}'
AMB_MONT_V = f'$I${R_AMB0}:$I${R_AMBF}'
CAB_CUSTO = [(1, 1, 'Item'), (2, 1, '%'), (3, 1, 'Orçado (R$)'), (4, 1, 'Realizado (R$)'),
             (5, 2, 'Desvio (R$)'), (7, 1, '% da venda'), (8, 3, 'Observação')]


def linha_custo(ws, r, rotulo, tipo, *, base_o=None, base_r=None, dica=''):
    ws.row_dimensions[r].height = 17
    bloco(ws, r, 1, 1, rotulo, f=Font(name=F, size=9.5, color=INK), bg=WHITE, al=LEFTI)
    if tipo == 'pct':
        bloco(ws, r, 2, 1, None, f=F_PCT, bg=INPUT, al=CTR, nf=PCT1)
        bloco(ws, r, 3, 1, f'=IF($B{r}="","",ROUND($B{r}*{base_o},2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 4, 1, f'=IF($B{r}="","",ROUND($B{r}*{base_r},2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
    else:
        bloco(ws, r, 2, 1, None, bg=WHITE)
        bloco(ws, r, 3, 1, None, f=F_ENT, bg=INPUT, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 4, 1, None, f=F_ENT, bg=INPUT, al=RIGHT, nf=MOEDA)
    bloco(ws, r, 5, 2, f'=IF(OR($C{r}="",$D{r}=""),"",ROUND($D{r}-$C{r},2))',
          f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
    bloco(ws, r, 7, 1, f'=IF(OR($D{r}="",{VENDA}="",{VENDA}=0),"",$D{r}/{VENDA})',
          f=Font(name=F, size=9, color=MUTED), bg=CALC, al=CTR, nf=PCT1)
    bloco(ws, r, 8, 3, dica or None, f=Font(name=F, size=8.5, color=INK), bg=INPUT, al=LEFTI)


def subtotal(ws, r, rotulo, ini, fim):
    ws.row_dimensions[r].height = 20
    bloco(ws, r, 1, 2, rotulo, f=F_SUB, bg=NAVY2, al=RIGHT)
    for col in (3, 4):
        bloco(ws, r, col, 1, f'=ROUND(SUM({LP[col]}{ini}:{LP[col]}{fim}),2)',
              f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, r, 5, 2, f'=ROUND($D{r}-$C{r},2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, r, 7, 1, f'=IF(OR({VENDA}="",{VENDA}=0),"",$D{r}/{VENDA})',
          f=Font(name=F, size=9.5, bold=True, color=GOLDS), bg=NAVY2, al=CTR, nf=PCT1)
    bloco(ws, r, 8, 3, None, bg=NAVY2)


def montar_ficha(ws):
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(W_FICHA, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = faixa_marca(ws, NCOL, 'FICHA DE CUSTO DO PROJETO',
                    'Fundo creme = você preenche · fundo cinza = calculado · duplique esta aba '
                    'a cada projeto e registre o nome dela no Painel Geral')
    ws.row_dimensions[r].height = 6

    # ── identificação
    IDENT = [((1, 3, 'CLIENTE'), (4, 2, 'Nº DO PROJETO'), (6, 2, 'DATA DE ENTRADA'),
              (8, 2, 'ENTREGA PREVISTA'), (10, 1, 'ENTREGA REAL')),
             ((1, 3, 'PROJETO'), (4, 2, 'VENDEDOR'), (6, 2, 'COORDENADOR'),
              (8, 2, 'DIAS DE ATRASO'), (10, 1, 'SITUAÇÃO DA ENTREGA'))]
    for k, linha in enumerate(IDENT):
        rl, rv = R_ID1 + k * 2, R_ID2 + k * 2
        for c0, span, rot in linha:
            bloco(ws, rl, c0, span, rot, f=Font(name=F, size=7.5, bold=True, color=MUTED),
                  bg=WHITE, al=CTR, bd=False)
            calc = (k == 1 and c0 in (8, 10))
            bloco(ws, rv, c0, span, None, f=Font(name=F, size=10, bold=True, color=NAVY2),
                  bg=CALC if calc else INPUT, al=LEFTI if c0 in (1, 4) else CTR)
        ws.row_dimensions[rl].height = 13
        ws.row_dimensions[rv].height = 21
    for ref, nf in ((f'F{R_ID2}', DATA), (f'H{R_ID2}', DATA), (f'J{R_ID2}', DATA),
                    (f'H{R_ID4}', DIAS)):
        ws[ref].number_format = nf
    ws[f'H{R_ID4}'] = f'=IF(OR($H${R_ID2}="",$J${R_ID2}=""),"",$J${R_ID2}-$H${R_ID2})'
    ws[f'J{R_ID4}'] = (f'=IF($H${R_ID2}="","",IF($J${R_ID2}="",'
                       f'IF(TODAY()>$H${R_ID2},"Atrasado","Em produção"),'
                       f'IF($H${R_ID4}<=0,"No prazo","Atrasado")))')
    ws.row_dimensions[R_ID4 + 1].height = 8

    # ── faixa de resultado + resumo por categoria (no topo)
    titulo_secao(ws, R_TOP, NCOL, 'Resultado do projeto',
                 'tudo aqui se atualiza conforme você lança os custos mais abaixo')
    KPIS = [(1, 2, 'VALOR DE VENDA', MOEDA0), (3, 2, 'CUSTO TOTAL', MOEDA0),
            (5, 2, 'MARGEM DE CONTRIBUIÇÃO', MOEDA0), (7, 1, 'MC %', PCT1),
            (8, 1, 'MC % ORÇADA', PCT1), (9, 2, 'DESVIO DE CUSTO', MOEDA0)]
    for c0, span, rot, nf in KPIS:
        bloco(ws, R_KPI_L, c0, span, rot, f=Font(name=F, size=7.5, bold=True, color=MUTED),
              bg=WHITE, al=CTR, bd=False)
        bloco(ws, R_KPI_V, c0, span, None, f=Font(name=F, size=13, bold=True, color=NAVY),
              bg=GOLDBG, al=CTR, nf=nf)
    ws.row_dimensions[R_KPI_L].height = 13
    ws.row_dimensions[R_KPI_V].height = 30
    ws.row_dimensions[R_KPI_V + 1].height = 8

    cab(ws, R_RES_H, [(1, 2, 'Resumo por categoria'), (3, 1, 'Orçado (R$)'),
                      (4, 1, 'Realizado (R$)'), (5, 2, 'Desvio (R$)'), (7, 1, 'Desvio (%)'),
                      (8, 3, '% da venda realizada')])
    for i, categoria in enumerate(CATEGORIAS):
        r = R_RES0 + i
        orig = SUBLINHAS[categoria]
        ws.row_dimensions[r].height = 18
        bloco(ws, r, 1, 2, categoria, f=Font(name=F, size=9.5, color=INK), bg=WHITE, al=LEFTI)
        for col in (3, 4):
            bloco(ws, r, col, 1, f'=IF({LP[col]}{orig}="","",{LP[col]}{orig})',
                  f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 5, 2, f'=IF(OR($C{r}="",$D{r}=""),"",ROUND($D{r}-$C{r},2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 7, 1, f'=IF(OR($C{r}="",$C{r}=0,$D{r}=""),"",$D{r}/$C{r}-1)',
              f=Font(name=F, size=9.5, bold=True, color=NAVY2), bg=CALC, al=CTR,
              nf='+0%;-0%;0%')
        bloco(ws, r, 8, 3, f'=IF(OR($D{r}="",{VENDA}="",{VENDA}=0),"",$D{r}/{VENDA})',
              f=Font(name=F, size=9.5, color=MUTED), bg=CALC, al=CTR, nf=PCT1)
    ws.row_dimensions[R_CUSTO_TOT].height = 22
    bloco(ws, R_CUSTO_TOT, 1, 2, '(=) CUSTO TOTAL DO PROJETO', f=F_SUB, bg=NAVY2, al=RIGHT)
    for col in (3, 4):
        bloco(ws, R_CUSTO_TOT, col, 1,
              f'=ROUND(SUM({LP[col]}{R_RES0}:{LP[col]}{R_RESF}),2)',
              f=Font(name=F, size=11, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, R_CUSTO_TOT, 5, 2, f'=ROUND($D${R_CUSTO_TOT}-$C${R_CUSTO_TOT},2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, R_CUSTO_TOT, 7, 1,
          f'=IF(OR($C${R_CUSTO_TOT}="",$C${R_CUSTO_TOT}=0),"",'
          f'$D${R_CUSTO_TOT}/$C${R_CUSTO_TOT}-1)',
          f=Font(name=F, size=9.5, bold=True, color=GOLDS), bg=NAVY2, al=CTR, nf='+0%;-0%;0%')
    bloco(ws, R_CUSTO_TOT, 8, 3,
          f'=IF(OR({VENDA}="",{VENDA}=0),"",$D${R_CUSTO_TOT}/{VENDA})',
          f=Font(name=F, size=9.5, bold=True, color=GOLDS), bg=NAVY2, al=CTR, nf=PCT1)
    ws.row_dimensions[R_MC].height = 28
    bloco(ws, R_MC, 1, 2, '(=) MARGEM DE CONTRIBUIÇÃO',
          f=Font(name=F, size=11.5, bold=True, color=WHITE), bg=NAVY, al=RIGHT)
    bloco(ws, R_MC, 3, 1, f'=ROUND($C${R_VENDA}-$C${R_CUSTO_TOT},2)',
          f=Font(name=F, size=12, bold=True, color=GOLDS), bg=NAVY, al=RIGHT, nf=MOEDA)
    bloco(ws, R_MC, 4, 1, f'=ROUND({VENDA}-$D${R_CUSTO_TOT},2)',
          f=Font(name=F, size=12, bold=True, color=GOLDS), bg=NAVY, al=RIGHT, nf=MOEDA)
    bloco(ws, R_MC, 5, 2, f'=ROUND($D${R_MC}-$C${R_MC},2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY, al=RIGHT, nf=MOEDA)
    bloco(ws, R_MC, 7, 1,
          f'=IF(OR($C${R_VENDA}="",$C${R_VENDA}=0),"",$C${R_MC}/$C${R_VENDA})',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY, al=CTR, nf=PCT1)
    bloco(ws, R_MC, 8, 3, f'=IF(OR({VENDA}="",{VENDA}=0),"",$D${R_MC}/{VENDA})',
          f=Font(name=F, size=12, bold=True, color=GOLDS), bg=NAVY, al=CTR, nf=PCT1)
    # alerta do que ainda falta comprar
    AC = f'ROUND(SUMIF({LAN_STA},"A comprar",{LAN_VAL}),2)'
    AP = f'ROUND(SUMIF({LAN_STA},"Comprado (a pagar)",{LAN_VAL}),2)'
    ws.row_dimensions[R_ALERTA].height = 22
    bloco(ws, R_ALERTA, 1, 8,
          f'=IF({VENDA}="","",IF({AC}+{AP}=0,'
          f'"   Nenhuma compra pendente lançada — o custo realizado já é o custo final.",'
          f'"   Já comprado e ainda a pagar: "&TEXT({AP},"R$ #,##0")&'
          f'"   ·   custo projetado: "&TEXT($D${R_CUSTO_TOT}+{AC},"R$ #,##0")&'
          f'"   ·   MC projetada: "&TEXT(({VENDA}-$D${R_CUSTO_TOT}-{AC})/{VENDA},"0.0%")))',
          f=Font(name=F, size=9.5, bold=True, color='7A5B17'), bg=GOLDBG, al=LEFTI)
    # valor numérico do que falta comprar — o Painel Geral lê esta célula
    bloco(ws, R_ALERTA, 9, 2, f'=IF({VENDA}="","",{AC})',
          f=Font(name=F, size=10, bold=True, color='7A5B17'), bg=GOLDBG, al=CTR,
          nf='"ainda a comprar  "R$ #,##0')
    ws.row_dimensions[R_ALERTA + 1].height = 8

    # ── 1 · valor de venda
    titulo_secao(ws, R_VEN_T, NCOL, '1 · Valor de venda')
    cab(ws, R_VEN_H, CAB_CUSTO)
    linha_custo(ws, R_VENDA, 'Valor de venda do projeto', 'rs',
                dica='Orçado = valor da proposta · Realizado = valor fechado')
    ws[f'G{R_VENDA}'] = None
    ws.row_dimensions[R_VENDA + 1].height = 8

    # ── 2 · custos de venda
    titulo_secao(ws, R_CV_T, NCOL, '2 · Custos de venda',
                 'tudo que sai por causa da venda · define a RECEITA LÍQUIDA')
    cab(ws, R_CV_H, CAB_CUSTO)
    BRT_O = f'($C${R_VENDA}-SUM($C${R_IMP}:$C${R_TRX}))'
    BRT_R = f'({VENDA}-SUM($D${R_IMP}:$D${R_TRX}))'
    linha_custo(ws, R_IMP, 'Impostos sobre a nota (Simples)', 'pct',
                base_o=f'$C${R_VENDA}', base_r=VENDA, dica='% sobre o valor de venda')
    linha_custo(ws, R_MAQ, 'Taxa de máquina de cartão', 'pct',
                base_o=f'$C${R_VENDA}', base_r=VENDA, dica='% sobre o valor de venda')
    linha_custo(ws, R_TRX, 'Taxas de transação (PIX, boleto, TED)', 'rs', dica='valor em reais')
    linha_custo(ws, R_CVEND, 'Comissão de venda (vendedor)', 'pct',
                base_o=f'$C${R_VENDA}', base_r=VENDA, dica='% sobre o valor de venda')
    linha_custo(ws, R_PROJ, 'Projeto / anteprojeto (externo)', 'rs',
                dica='projetista ou arquiteto contratado para o projeto')
    linha_custo(ws, R_RTP, 'RT do parceiro', 'pct', base_o=BRT_O, base_r=BRT_R,
                dica='% sobre o líquido: venda menos impostos, máquina e taxas')
    subtotal(ws, R_CV_SUB, '(=) SUBTOTAL DOS CUSTOS DE VENDA', R_IMP, R_RTP)
    ws.row_dimensions[R_LIQ].height = 24
    bloco(ws, R_LIQ, 1, 2, '(=) RECEITA LÍQUIDA',
          f=Font(name=F, size=11, bold=True, color=WHITE), bg=NAVY, al=RIGHT)
    bloco(ws, R_LIQ, 3, 1, f'=ROUND($C${R_VENDA}-$C${R_CV_SUB},2)',
          f=Font(name=F, size=11.5, bold=True, color=GOLDS), bg=NAVY, al=RIGHT, nf=MOEDA)
    bloco(ws, R_LIQ, 4, 1, f'=ROUND({VENDA}-$D${R_CV_SUB},2)',
          f=Font(name=F, size=11.5, bold=True, color=GOLDS), bg=NAVY, al=RIGHT, nf=MOEDA)
    bloco(ws, R_LIQ, 5, 2, f'=ROUND($D${R_LIQ}-$C${R_LIQ},2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY, al=RIGHT, nf=MOEDA)
    bloco(ws, R_LIQ, 7, 1, f'=IF(OR({VENDA}="",{VENDA}=0),"",$D${R_LIQ}/{VENDA})',
          f=Font(name=F, size=9.5, bold=True, color=GOLDS), bg=NAVY, al=CTR, nf=PCT1)
    bloco(ws, R_LIQ, 8, 3, 'É esta a base das comissões de coordenação, produção e montagem',
          f=Font(name=F, size=8.5, color=GOLDS, i=True), bg=NAVY, al=LEFTI)
    ws.row_dimensions[R_LIQ + 1].height = 8

    # ── 3 · ambientes
    titulo_secao(ws, R_AMB_T, NCOL, '3 · Ambientes, produção e montagem',
                 'cada ambiente com seu produtor e seu montador · a comissão incide sobre a receita líquida')
    cab(ws, R_AMB_H, [(1, 1, 'Ambiente'), (2, 1, '% do total'), (3, 1, 'Valor do ambiente (R$)'),
                      (4, 1, 'Produção — quem'), (5, 1, '%'), (6, 1, 'Comissão produção (R$)'),
                      (7, 1, 'Montagem — quem'), (8, 1, '%'), (9, 1, 'Comissão montagem (R$)'),
                      (10, 1, 'Total do ambiente')], alt=32)
    for r in range(R_AMB0, R_AMBF + 1):
        ws.row_dimensions[r].height = 18
        bloco(ws, r, 1, 1, None, f=F_ENT, bg=INPUT, al=LEFTI)
        bloco(ws, r, 2, 1, f'=IF(OR($C{r}="",{VENDA}="",{VENDA}=0),"",$C{r}/{VENDA})',
              f=Font(name=F, size=9, color=MUTED), bg=CALC, al=CTR, nf=PCT1)
        bloco(ws, r, 3, 1, None, f=Font(name=F, size=9.5, bold=True, color=NAVY),
              bg=INPUT, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 4, 1, None, f=F_ENT, bg=INPUT, al=LEFTI)
        bloco(ws, r, 5, 1, None, f=F_PCT, bg=INPUT, al=CTR, nf=PCT1)
        bloco(ws, r, 6, 1, f'=IF(OR($B{r}="",$E{r}="",$D${R_LIQ}=""),"",'
                           f'ROUND($B{r}*$D${R_LIQ}*$E{r},2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 7, 1, None, f=F_ENT, bg=INPUT, al=LEFTI)
        bloco(ws, r, 8, 1, None, f=F_PCT, bg=INPUT, al=CTR, nf=PCT1)
        bloco(ws, r, 9, 1, f'=IF(OR($B{r}="",$H{r}="",$D${R_LIQ}=""),"",'
                           f'ROUND($B{r}*$D${R_LIQ}*$H{r},2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 10, 1, f'=IF(AND($F{r}="",$I{r}=""),"",'
                            f'ROUND(IF($F{r}="",0,$F{r})+IF($I{r}="",0,$I{r}),2))',
              f=Font(name=F, size=9.5, bold=True, color=NAVY2), bg=CALC, al=RIGHT, nf=MOEDA)
    dv(ws, '=EQUIPE_COMISSAO', f'D{R_AMB0}:D{R_AMBF}')
    dv(ws, '=EQUIPE_COMISSAO', f'G{R_AMB0}:G{R_AMBF}')
    ws.row_dimensions[R_AMB_TOT].height = 20
    bloco(ws, R_AMB_TOT, 1, 1, 'SOMA DOS AMBIENTES', f=F_SUB, bg=NAVY2, al=RIGHT)
    bloco(ws, R_AMB_TOT, 2, 1, f'=IF({VENDA}=0,"",$C${R_AMB_TOT}/{VENDA})',
          f=Font(name=F, size=9.5, bold=True, color=GOLDS), bg=NAVY2, al=CTR, nf=PCT1)
    for col in (3, 6, 9, 10):
        bloco(ws, R_AMB_TOT, col, 1, f'=ROUND(SUM({LP[col]}{R_AMB0}:{LP[col]}{R_AMBF}),2)',
              f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, R_AMB_TOT, 4, 2,
          f'=IF($C${R_AMB_TOT}=0,"—",IF(ROUND($C${R_AMB_TOT}-{VENDA},2)=0,'
          f'"OK · os ambientes somam o valor de venda",'
          f'"ATENÇÃO: diferença de "&TEXT($C${R_AMB_TOT}-{VENDA},"R$ #,##0.00")))',
          f=Font(name=F, size=8.5, bold=True, color=GOLDS), bg=NAVY2, al=CTR)
    bloco(ws, R_AMB_TOT, 7, 2, None, bg=NAVY2)
    ws.row_dimensions[R_AMB_TOT + 1].height = 8

    # ── 4 · comissões operacionais
    titulo_secao(ws, R_CO_T, NCOL, '4 · Comissões operacionais',
                 'calculadas sobre a RECEITA LÍQUIDA · o realizado vem dos ambientes acima')
    cab(ws, R_CO_H, CAB_CUSTO)
    linha_custo(ws, R_COORD, 'Coordenação de produção', 'pct',
                base_o=f'$C${R_LIQ}', base_r=f'$D${R_LIQ}',
                dica='quem coordena está na identificação, no topo da ficha')
    for rr, rot, orig, dica in (
            (R_PRODC, 'Comissão de produção', f'$F${R_AMB_TOT}',
             'orçado = % único estimado · realizado = soma dos ambientes'),
            (R_MONTC, 'Comissão de montagem', f'$I${R_AMB_TOT}',
             'orçado = % único estimado · realizado = soma dos ambientes')):
        linha_custo(ws, rr, rot, 'pct', base_o=f'$C${R_LIQ}', base_r=f'$D${R_LIQ}', dica=dica)
        bloco(ws, rr, 4, 1, f'=IF({orig}=0,"",{orig})', f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
    subtotal(ws, R_CO_SUB, '(=) SUBTOTAL DAS COMISSÕES', R_COORD, R_MONTC)
    ws.row_dimensions[R_CO_SUB + 1].height = 8

    # ── 5 · comissões por colaborador
    titulo_secao(ws, R_CL_T, NCOL, '5 · Comissões por colaborador',
                 'consolidado automático — só marceneiros, ajudantes e o coordenador')
    cab(ws, R_CL_H, [(1, 1, 'Colaborador'), (2, 2, 'Produção (R$)'), (4, 2, 'Montagem (R$)'),
                     (6, 2, 'Coordenação (R$)'), (8, 3, 'Total no projeto (R$)')])
    for i, r in enumerate(range(R_CL0, R_CLF + 1)):
        ws.row_dimensions[r].height = 17
        bloco(ws, r, 1, 1, EQUIPE[i] if i < len(EQUIPE) else None,
              f=Font(name=F, size=9.5, bold=True, color=NAVY2), bg=INPUT, al=LEFTI)
        bloco(ws, r, 2, 2, f'=IF($A{r}="","",ROUND(SUMIF({AMB_PROD_Q},$A{r},{AMB_PROD_V}),2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 4, 2, f'=IF($A{r}="","",ROUND(SUMIF({AMB_MONT_Q},$A{r},{AMB_MONT_V}),2))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 6, 2, f'=IF($A{r}="","",IF($A{r}=$F${R_ID4},'
                           f'IF($D${R_COORD}="",0,$D${R_COORD}),0))',
              f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 8, 3, f'=IF($A{r}="","",ROUND($B{r}+$D{r}+$F{r},2))',
              f=Font(name=F, size=10, bold=True, color=NAVY), bg=CALC, al=RIGHT, nf=MOEDA)
    dv(ws, '=EQUIPE_COMISSAO', f'A{R_CL0}:A{R_CLF}')
    ws.row_dimensions[R_CL_TOT].height = 20
    bloco(ws, R_CL_TOT, 1, 1, 'TOTAL', f=F_SUB, bg=NAVY2, al=RIGHT)
    for c0, span in ((2, 2), (4, 2), (6, 2), (8, 3)):
        bloco(ws, R_CL_TOT, c0, span, f'=ROUND(SUM({LP[c0]}{R_CL0}:{LP[c0]}{R_CLF}),2)',
              f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    ws.row_dimensions[R_CL_TOT + 1].height = 8

    # ── 6 · orçamento e compras por categoria
    titulo_secao(ws, R_OC_T, NCOL, '6 · Material, terceirizados e logística',
                 'o orçado você digita · o realizado vem sozinho do livro de lançamentos, lá no fim da ficha')
    cab(ws, R_OC_H, [(1, 1, 'Categoria'), (2, 1, ''), (3, 1, 'Orçado (R$)'),
                     (4, 1, 'Realizado (R$)'), (5, 2, 'Desvio (R$)'),
                     (7, 1, 'Ainda a comprar'), (8, 3, 'Comprado, ainda a pagar')], alt=30)
    for grupo, cats in GRUPOS:
        for cat in cats:
            r = LINHAS_OC[cat]
            ws.row_dimensions[r].height = 17
            bloco(ws, r, 1, 1, cat, f=Font(name=F, size=9.5, color=INK), bg=WHITE, al=LEFTI)
            bloco(ws, r, 2, 1, None, bg=WHITE)
            bloco(ws, r, 3, 1, None, f=F_ENT, bg=INPUT, al=RIGHT, nf=MOEDA)
            bloco(ws, r, 4, 1,
                  f'=ROUND(SUMIFS({LAN_VAL},{LAN_CAT},$A{r},{LAN_STA},"Comprado (a pagar)")'
                  f'+SUMIFS({LAN_VAL},{LAN_CAT},$A{r},{LAN_STA},"Pago"),2)',
                  f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
            bloco(ws, r, 5, 2, f'=IF($C{r}="","",ROUND($D{r}-$C{r},2))',
                  f=F_CALC, bg=CALC, al=RIGHT, nf=MOEDA)
            bloco(ws, r, 7, 1,
                  f'=ROUND(SUMIFS({LAN_VAL},{LAN_CAT},$A{r},{LAN_STA},"A comprar"),2)',
                  f=Font(name=F, size=9.5, color=AMBER), bg=CALC, al=RIGHT, nf=MOEDA)
            bloco(ws, r, 8, 3,
                  f'=ROUND(SUMIFS({LAN_VAL},{LAN_CAT},$A{r},{LAN_STA},"Comprado (a pagar)"),2)',
                  f=Font(name=F, size=9.5, color=BLUE), bg=CALC, al=RIGHT, nf=MOEDA)
        rs = R_SUBGRUPO[grupo]
        i0 = LINHAS_OC[cats[0]]
        ws.row_dimensions[rs].height = 20
        bloco(ws, rs, 1, 2, f'(=) SUBTOTAL · {grupo.upper()}', f=F_SUB, bg=NAVY2, al=RIGHT)
        for c0, span in ((3, 1), (4, 1), (5, 2), (7, 1), (8, 3)):
            bloco(ws, rs, c0, span, f'=ROUND(SUM({LP[c0]}{i0}:{LP[c0]}{rs-1}),2)',
                  f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    ws.row_dimensions[R_OCF + 1].height = 8

    # ── 7 · retrabalho
    titulo_secao(ws, R_RB_T, NCOL, '7 · Retrabalho',
                 'o que aconteceu, por que aconteceu e quanto custou — é aqui que o orçamento aprende')
    cab(ws, R_RB_H, [(1, 1, 'O que aconteceu'), (2, 3, 'Causa'),
                     (5, 2, 'Custo estimado (R$)'), (7, 4, 'Providência / responsável')])
    for r in range(R_RB0, R_RBF + 1):
        ws.row_dimensions[r].height = 18
        bloco(ws, r, 1, 1, None, f=F_ENT, bg=INPUT, al=LEFTI)
        bloco(ws, r, 2, 3, None, f=F_ENT, bg=INPUT, al=LEFTI)
        bloco(ws, r, 5, 2, None, f=Font(name=F, size=9.5, bold=True, color=RED),
              bg=INPUT, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 7, 4, None, f=Font(name=F, size=8.5, color=INK), bg=INPUT, al=LEFTI)
    dv(ws, '=CAUSA_RETRABALHO', f'B{R_RB0}:B{R_RBF}')
    ws.row_dimensions[R_RB_SUB].height = 20
    bloco(ws, R_RB_SUB, 1, 2, 'CONTINGÊNCIA PREVISTA NO ORÇAMENTO  →', f=F_SUB,
          bg=NAVY2, al=RIGHT)
    bloco(ws, R_RB_SUB, 3, 1, None, f=Font(name=F, size=10, bold=True, color=GOLDS),
          bg=INPUT, al=RIGHT, nf=MOEDA)
    bloco(ws, R_RB_SUB, 4, 1, f'=ROUND(SUM($E${R_RB0}:$E${R_RBF}),2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, R_RB_SUB, 5, 2, f'=ROUND($D${R_RB_SUB}-IF($C${R_RB_SUB}="",0,$C${R_RB_SUB}),2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, R_RB_SUB, 7, 1, f'=IF(OR({VENDA}="",{VENDA}=0),"",$D${R_RB_SUB}/{VENDA})',
          f=Font(name=F, size=9.5, bold=True, color=GOLDS), bg=NAVY2, al=CTR, nf=PCT1)
    bloco(ws, R_RB_SUB, 8, 3, 'orçado = a contingência prevista · realizado = a soma acima',
          f=Font(name=F, size=8.5, color=GOLDS, i=True), bg=NAVY2, al=LEFTI)
    ws.row_dimensions[R_RB_SUB + 1].height = 8

    # ── 8 · livro de lançamentos
    titulo_secao(ws, R_LAN_T, NCOL, '8 · Livro de compras e despesas do projeto',
                 'uma linha por compra, na ordem em que acontecer · pode inserir linhas à vontade aqui')
    cab(ws, R_LAN_H, [(1, 1, 'Descrição / fornecedor'), (2, 2, 'Categoria'), (4, 1, 'Data'),
                      (5, 2, 'Valor (R$)'), (7, 2, 'Forma de pagamento'), (9, 2, 'Status')],
        alt=28)
    for r in range(R_LAN0, R_LANF + 1):
        ws.row_dimensions[r].height = 17
        bloco(ws, r, 1, 1, None, f=F_ENT, bg=INPUT, al=LEFTI)
        bloco(ws, r, 2, 2, None, f=Font(name=F, size=9, color=INK), bg=INPUT, al=LEFTI)
        bloco(ws, r, 4, 1, None, f=font(9.5), bg=INPUT, al=CTR, nf=DATA)
        bloco(ws, r, 5, 2, None, f=Font(name=F, size=9.5, bold=True, color=NAVY2),
              bg=INPUT, al=RIGHT, nf=MOEDA)
        bloco(ws, r, 7, 2, None, f=Font(name=F, size=9, color=INK), bg=INPUT, al=CTR)
        bloco(ws, r, 9, 2, None, f=Font(name=F, size=9, bold=True, color=NAVY2),
              bg=INPUT, al=CTR)
    dv(ws, '=CATEGORIA_COMPRA', f'B{R_LAN0}:B{R_LANF}')
    dv(ws, '=FORMA_PAGAMENTO', f'G{R_LAN0}:G{R_LANF}')
    dv(ws, '=STATUS_COMPRA', f'I{R_LAN0}:I{R_LANF}')
    for txt, bg, cor in (('Pago', OKBG, OK), ('Comprado (a pagar)', BLUEBG, BLUE),
                         ('A comprar', AMBBG, AMBER)):
        ws.conditional_formatting.add(f'I{R_LAN0}:J{R_LANF}', FormulaRule(
            formula=[f'$I{R_LAN0}="{txt}"'], fill=fill(bg), font=Font(bold=True, color=cor),
            stopIfTrue=True))
    ws.conditional_formatting.add(f'A{R_LAN0}:J{R_LANF}', FormulaRule(
        formula=[f'AND($E{R_LAN0}<>"",$B{R_LAN0}="")'], fill=fill(REDBG)))
    ws.row_dimensions[R_LAN_TOT].height = 22
    bloco(ws, R_LAN_TOT, 1, 4, 'TOTAL LANÇADO NO LIVRO', f=F_SUB, bg=NAVY2, al=RIGHT)
    bloco(ws, R_LAN_TOT, 5, 2, f'=ROUND(SUM({LAN_VAL}),2)',
          f=Font(name=F, size=11, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA)
    bloco(ws, R_LAN_TOT, 7, 4,
          f'=IF(COUNTA({LAN_CAT})=COUNT({LAN_VAL}),'
          f'"lançamentos: "&COUNT({LAN_VAL})&" · todos classificados",'
          f'"ATENÇÃO: "&(COUNT({LAN_VAL})-COUNTA({LAN_CAT}))&" lançamento(s) sem categoria")',
          f=Font(name=F, size=9, bold=True, color=GOLDS), bg=NAVY2, al=CTR)

    # ── KPIs do topo
    ws[f'A{R_KPI_V}'] = f'=IF($D${R_VENDA}="",$C${R_VENDA},$D${R_VENDA})'
    ws[f'C{R_KPI_V}'] = f'=IF($D${R_CUSTO_TOT}=0,"",$D${R_CUSTO_TOT})'
    ws[f'E{R_KPI_V}'] = f'=IF({VENDA}="","",$D${R_MC})'
    ws[f'G{R_KPI_V}'] = f'=IF(OR({VENDA}="",{VENDA}=0),"",$D${R_MC}/{VENDA})'
    ws[f'H{R_KPI_V}'] = f'=IF(OR($C${R_VENDA}="",$C${R_VENDA}=0),"",$C${R_MC}/$C${R_VENDA})'
    ws[f'I{R_KPI_V}'] = (f'=IF(OR($C${R_CUSTO_TOT}=0,$D${R_CUSTO_TOT}=0),"",'
                         f'ROUND($D${R_CUSTO_TOT}-$C${R_CUSTO_TOT},2))')

    # ── nota de rodapé
    ws.row_dimensions[R_NOTA].height = 50
    bloco(ws, R_NOTA, 1, NCOL,
          '  Preencha primeiro a coluna ORÇADO com o que o orçamento previu e vá preenchendo o REALIZADO conforme '
          'o projeto anda — o desvio de cada linha é o que mostra onde o orçamento erra. Compras entram uma a uma '
          'no livro do bloco 8, com categoria, forma de pagamento e status; as categorias do bloco 6 se somam '
          'sozinhas. Você pode inserir linhas em qualquer bloco a partir da linha 26: tudo o que o Painel Geral lê '
          'está acima disso.',
          f=Font(name=F, size=8.5, color='41505D'), bg=GOLDBG, al=LEFTIW)

    # ── formatação condicional
    for rng, ref in ((f'E{R_RES0}:F{R_CUSTO_TOT}', f'$E{R_RES0}'),
                     (f'E{R_OC_H+1}:F{R_OCF}', f'$E{R_OC_H+1}')):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({ref}<>"",{ref}>0)'], font=Font(bold=True, color=RED)))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({ref}<>"",{ref}<0)'], font=Font(bold=True, color=OK)))
    ws.conditional_formatting.add(f'J{R_ID4}', FormulaRule(
        formula=[f'$J${R_ID4}="No prazo"'], fill=fill(OKBG), font=Font(bold=True, color=OK),
        stopIfTrue=True))
    ws.conditional_formatting.add(f'J{R_ID4}', FormulaRule(
        formula=[f'$J${R_ID4}="Atrasado"'], fill=fill(REDBG), font=Font(bold=True, color=RED)))
    ws.conditional_formatting.add(f'A{R_KPI_V}:J{R_KPI_V}', FormulaRule(
        formula=[f'AND($G${R_KPI_V}<>"",$G${R_KPI_V}<0.25)'], fill=fill(REDBG)))
    dv(ws, '=VENDEDOR', f'D{R_ID2}:E{R_ID2}')
    dv(ws, '=EQUIPE_COMISSAO', f'F{R_ID4}:G{R_ID4}')
    ws.freeze_panes = f'A{R_VEN_T}'
    print_cfg(ws, f'A1:J{R_NOTA}')
    return ws


# ══════════════════════════════════════════════════════════════════════
#  ABAS
# ══════════════════════════════════════════════════════════════════════
D = datetime.date
EXEMPLO = {
    'cliente': 'Jonathan Vargas', 'num': 'P-2026-041',
    'projeto': 'Apartamento completo — 4 ambientes',
    'entrada': D(2026, 5, 12), 'prevista': D(2026, 7, 30), 'real': D(2026, 8, 6),
    'vendedor': 'Jonathan', 'coordenador': 'Deivson',
    'venda_o': 90000, 'venda_r': 90000,
    'cv': {R_IMP: 0.075, R_MAQ: 0.02, R_CVEND: 0.03, R_RTP: 0.05},
    'cv_rs': {R_TRX: (120, 148), R_PROJ: (1500, 1500)},
    'ambientes': [('Cozinha', 30000, 'Jackson', 0.03, 'Samuel', 0.02),
                  ('Suíte', 20000, 'Samuel', 0.03, 'Cezar', 0.02),
                  ('Lavanderia', 10000, 'Joelson', 0.03, 'Samuel', 0.02),
                  ('Sala', 30000, 'Deivson', 0.03, 'Jackson', 0.02)],
    'coord': 0.01, 'prod_o': 0.03, 'mont_o': 0.02,
    'orcado_cat': {'MDF e MDP': 14500, 'Fita de borda': 900,
                   'Ferragens': 6800, 'Vidros e espelhos': 2200,
                   'Lâmina natural': 3200,
                   'Consumíveis': 1100,
                   'Acabamento': 4800, 'Serralheria': 1500,
                   'Vidraceiro': 900, 'Uber e aplicativo': 250,
                   'Carreto e entrega': 900,
                   'Deslocamento da equipe': 400,
                   'Frete de material': 350, 'Estacionamento e pedágio': 100},
    'contingencia': 1500,
    'lancamentos': [
        ('MADEGEM — chapas brancas TX', 'MDF e MDP', D(2026, 5, 12), 8200, 'PIX', 'Pago'),
        ('Frete das chapas', 'Frete de material', D(2026, 5, 12), 350, 'Dinheiro', 'Pago'),
        ('Consumíveis — parafuso e cola', 'Consumíveis',
         D(2026, 5, 15), 640, 'Dinheiro', 'Pago'),
        ('Lâmina natural freijó', 'Lâmina natural', D(2026, 5, 19), 3200, 'PIX', 'Pago'),
        ('Bigfer — corrediças e dobradiças', 'Ferragens', D(2026, 5, 20),
         4300, 'Cartão parcelado', 'Pago'),
        ('Uber — medição e conferência', 'Uber e aplicativo', D(2026, 5, 22), 230, 'PIX', 'Pago'),
        ('MADEGEM — complemento de chapas', 'MDF e MDP', D(2026, 5, 28), 5180,
         'Boleto', 'Pago'),
        ('Fita de borda 6 cores', 'Fita de borda', D(2026, 5, 28), 1120, 'PIX', 'Pago'),
        ('Consumíveis — lixa e tíner', 'Consumíveis',
         D(2026, 6, 3), 465, 'PIX', 'Pago'),
        ('JR Ferragens — puxadores', 'Ferragens', D(2026, 6, 11), 3150, 'PIX', 'Pago'),
        ('MADEGEM — reposição pós-retrabalho', 'MDF e MDP', D(2026, 6, 14), 2600,
         'Boleto', 'Comprado (a pagar)'),
        ('Consumíveis — estopa e adesivo', 'Consumíveis',
         D(2026, 6, 25), 360, 'Dinheiro', 'Pago'),
        ('Serralheria — estrutura da bancada', 'Serralheria', D(2026, 6, 26), 1500, 'PIX', 'Pago'),
        ('Espelho da suíte', 'Vidros e espelhos', D(2026, 7, 2), 2200, 'Boleto',
         'Comprado (a pagar)'),
        ('Vidraceiro — instalação', 'Vidraceiro', D(2026, 7, 4), 1250, 'PIX', 'Pago'),
        ('Laqueação das portas', 'Acabamento', D(2026, 7, 10), 5400,
         'Faturado 30 dias', 'Comprado (a pagar)'),
        ('Carreto — 1ª entrega', 'Carreto e entrega', D(2026, 7, 28), 650, 'PIX', 'Pago'),
        ('Uber — acerto final com o cliente', 'Uber e aplicativo', D(2026, 7, 30), 180,
         'PIX', 'Pago'),
        ('Carreto — 2ª entrega e retorno', 'Carreto e entrega', D(2026, 8, 5), 500,
         'PIX', 'Pago'),
        ('Deslocamento da equipe de montagem', 'Deslocamento da equipe',
         D(2026, 8, 5), 620, 'Dinheiro', 'Pago'),
        ('Estacionamento e pedágio', 'Estacionamento e pedágio', D(2026, 8, 6), 185,
         'Dinheiro', 'Pago'),
    ],
    'retrabalho': [
        ('Porta da cozinha empenou depois de instalada', 'Falha de material', 850,
         'Refeita em MDF de outro lote · fornecedor notificado'),
        ('Nicho da suíte 4 cm fora da medida', 'Erro de medição', 1200,
         'Refeito · medição passa a ser conferida por 2 pessoas'),
        ('Risco no tampo da lavanderia no transporte', 'Dano no transporte', 320,
         'Polido no local · exigir manta no carreto'),
    ],
}


def preencher(ws, d):
    ws[f'A{R_ID2}'] = d['cliente']; ws[f'D{R_ID2}'] = d['num']
    ws[f'F{R_ID2}'] = d['entrada']; ws[f'H{R_ID2}'] = d['prevista']; ws[f'J{R_ID2}'] = d['real']
    ws[f'A{R_ID4}'] = d['projeto']; ws[f'D{R_ID4}'] = d['vendedor']
    ws[f'F{R_ID4}'] = d['coordenador']
    ws[f'C{R_VENDA}'] = d['venda_o']; ws[f'D{R_VENDA}'] = d['venda_r']
    for r, pct in d['cv'].items():
        ws[f'B{r}'] = pct
    for r, (o, rr) in d['cv_rs'].items():
        ws[f'C{r}'] = o; ws[f'D{r}'] = rr
    for i, (nome, val, pq, pp, mq, mp) in enumerate(d['ambientes']):
        r = R_AMB0 + i
        ws[f'A{r}'] = nome; ws[f'C{r}'] = val
        ws[f'D{r}'] = pq; ws[f'E{r}'] = pp
        ws[f'G{r}'] = mq; ws[f'H{r}'] = mp
    ws[f'B{R_COORD}'] = d['coord']
    ws[f'B{R_PRODC}'] = d['prod_o']; ws[f'B{R_MONTC}'] = d['mont_o']
    for cat, val in d['orcado_cat'].items():
        ws[f'C{LINHAS_OC[cat]}'] = val
    ws[f'C{R_RB_SUB}'] = d['contingencia']
    for i, (oc, causa, custo, prov) in enumerate(d['retrabalho']):
        r = R_RB0 + i
        ws[f'A{r}'] = oc; ws[f'B{r}'] = causa; ws[f'E{r}'] = custo; ws[f'G{r}'] = prov
    for i, (desc, cat, data, val, forma, status) in enumerate(d['lancamentos']):
        r = R_LAN0 + i
        ws[f'A{r}'] = desc; ws[f'B{r}'] = cat; ws[f'D{r}'] = data
        ws[f'E{r}'] = val; ws[f'G{r}'] = forma; ws[f'I{r}'] = status


ficha = wb.create_sheet('Ficha Modelo')
montar_ficha(ficha)
ex = wb.create_sheet('Exemplo P-2026-041')
montar_ficha(ex)
preencher(ex, EXEMPLO)


# ══════════════════════════════════════════════════════════════════════
#  ABA · Painel Geral
# ══════════════════════════════════════════════════════════════════════
pg = wb.create_sheet('Painel Geral', 0)
pg.sheet_view.showGridLines = False
NC_PG = 18
N_SLOT = 40
C_AUX0 = 34
W_PG = [5, 22, 20, 26, 11, 11, 11, 9, 14, 14, 14, 14, 14, 9, 9, 9, 14, 24]
r = faixa_marca(pg, NC_PG, 'PAINEL GERAL DE CUSTO POR PROJETO',
                'Cada linha puxa uma ficha · digite o nome exato da aba na coluna ABA e o '
                'resto aparece sozinho')
for i, w in enumerate(W_PG, start=1):
    pg.column_dimensions[get_column_letter(i)].width = w
pg.row_dimensions[r].height = 6
r += 1

P0, PF = r + 5, r + 4 + N_SLOT
KPI_PG = [(1, 3, 'PROJETOS', '0'), (4, 3, 'VENDIDO', MOEDA0), (7, 3, 'CUSTO TOTAL', MOEDA0),
          (10, 3, 'AINDA A COMPRAR', MOEDA0), (13, 2, 'MARGEM DE CONTRIB.', MOEDA0),
          (15, 2, 'MC MÉDIA', PCT1), (17, 2, 'ENTREGA NO PRAZO', PCT0)]
FX_KPI = [f'=COUNTA($B${P0}:$B${PF})',
          f'=ROUND(SUM($J${P0}:$J${PF}),2)',
          f'=ROUND(SUM($K${P0}:$K${PF}),2)',
          f'=ROUND(SUM($L${P0}:$L${PF}),2)',
          f'=ROUND(SUM($M${P0}:$M${PF}),2)',
          f'=IF(SUM($J${P0}:$J${PF})=0,"",SUM($M${P0}:$M${PF})/SUM($J${P0}:$J${PF}))',
          f'=IF(COUNTIF($I${P0}:$I${PF},"No prazo")+COUNTIF($I${P0}:$I${PF},"Atrasado")=0,"",'
          f'COUNTIF($I${P0}:$I${PF},"No prazo")/'
          f'(COUNTIF($I${P0}:$I${PF},"No prazo")+COUNTIF($I${P0}:$I${PF},"Atrasado")))']
for (c0, span, rot, nf), fx in zip(KPI_PG, FX_KPI):
    bloco(pg, r, c0, span, rot, f=Font(name=F, size=8, bold=True, color=MUTED), bg=WHITE,
          al=CTR, bd=False)
    bloco(pg, r + 1, c0, span, fx, f=Font(name=F, size=14, bold=True, color=NAVY),
          bg=GOLDBG, al=CTR, nf=nf)
pg.row_dimensions[r].height = 14
pg.row_dimensions[r + 1].height = 32
pg.row_dimensions[r + 2].height = 8
r = titulo_secao(pg, r + 3, NC_PG, 'Projetos',
                 'digite na coluna ABA o nome exato da lâmina de cada projeto')
HDR_PG = ['#', 'Aba do projeto', 'Cliente', 'Projeto', 'Entrada', 'Prevista', 'Entregue',
          'Dias', 'Entrega', 'Venda', 'Custo total', 'A comprar', 'MC (R$)', 'MC %',
          'MC % orç.', 'Δ p.p.', 'Desvio de custo', 'Diagnóstico']
r = cab(pg, r, [(i + 1, 1, h) for i, h in enumerate(HDR_PG)], alt=30)
assert r == P0, f'P0 esperado {P0}, obtido {r}'


def puxa(r, ref):
    return f'=IFERROR(INDIRECT("\'"&$B{r}&"\'!{ref}"),"")'


REFS = [(3, f'A{R_ID2}', None, LEFTI), (4, f'A{R_ID4}', None, LEFTI),
        (5, f'F{R_ID2}', DATA, CTR), (6, f'H{R_ID2}', DATA, CTR),
        (7, f'J{R_ID2}', DATA, CTR), (8, f'H{R_ID4}', '0', CTR),
        (9, f'J{R_ID4}', None, CTR), (10, f'A{R_KPI_V}', MOEDA0, RIGHT),
        (11, f'C{R_KPI_V}', MOEDA0, RIGHT), (12, f'I{R_ALERTA}', MOEDA0, RIGHT),
        (13, f'E{R_KPI_V}', MOEDA0, RIGHT), (14, f'G{R_KPI_V}', PCT1, CTR),
        (15, f'H{R_KPI_V}', PCT1, CTR), (17, f'I{R_KPI_V}', MOEDA0, RIGHT)]

for i, rr in enumerate(range(P0, PF + 1), start=1):
    pg.row_dimensions[rr].height = 17
    bloco(pg, rr, 1, 1, i, f=Font(name=F, size=8.5, color=MUTED), bg=CALC, al=CTR)
    bloco(pg, rr, 2, 1, None, f=Font(name=F, size=9.5, bold=True, color=NAVY),
          bg=INPUT, al=LEFTI)
    for col, ref, nf, al in REFS:
        bloco(pg, rr, col, 1, puxa(rr, ref), f=Font(name=F, size=9.5, color=NAVY2),
              bg=CALC, al=al, nf=nf)
    bloco(pg, rr, 16, 1, f'=IF(OR($N{rr}="",$O{rr}=""),"",ROUND(($N{rr}-$O{rr})*100,1))',
          f=Font(name=F, size=9.5, bold=True, color=NAVY2), bg=CALC, al=CTR,
          nf='+0.0;-0.0;0.0')
    bloco(pg, rr, 18, 1,
          f'=IF($N{rr}="","",IF($N{rr}>=0.35,"Margem boa",'
          f'IF($N{rr}>=0.25,"Margem apertada","Margem crítica")))',
          f=Font(name=F, size=8.5, bold=True, color=NAVY2), bg=CALC, al=CTR)
    for k in range(len(CATEGORIAS)):
        for j, col_f in enumerate(('C', 'D')):
            c = pg.cell(rr, C_AUX0 + k * 2 + j, puxa(rr, f'{col_f}{R_RES0 + k}'))
            c.number_format = MOEDA; c.font = font(8, c=MUTED)

pg[f'B{P0}'] = 'Exemplo P-2026-041'
pg.row_dimensions[PF + 1].height = 20
bloco(pg, PF + 1, 1, 9, 'TOTAL', f=F_SUB, bg=NAVY2, al=RIGHT)
for col in (10, 11, 12, 13):
    bloco(pg, PF + 1, col, 1, f'=ROUND(SUM({LP[col]}{P0}:{LP[col]}{PF}),2)',
          f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA0)
bloco(pg, PF + 1, 14, 1, f'=IF($J${PF+1}=0,"",$M${PF+1}/$J${PF+1})',
      f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=CTR, nf=PCT1)
bloco(pg, PF + 1, 15, 2, None, bg=NAVY2)
bloco(pg, PF + 1, 17, 1, f'=ROUND(SUM($Q${P0}:$Q${PF}),2)',
      f=Font(name=F, size=10, bold=True, color=GOLDS), bg=NAVY2, al=RIGHT, nf=MOEDA0)
bloco(pg, PF + 1, 18, 1, None, bg=NAVY2)
for txt, bg, cor in (('Margem boa', OKBG, OK), ('Margem apertada', AMBBG, AMBER),
                     ('Margem crítica', REDBG, RED)):
    pg.conditional_formatting.add(f'R{P0}:R{PF}', FormulaRule(
        formula=[f'$R{P0}="{txt}"'], fill=fill(bg), font=Font(bold=True, color=cor),
        stopIfTrue=True))
for txt, bg, cor in (('No prazo', OKBG, OK), ('Atrasado', REDBG, RED),
                     ('Em produção', BLUEBG, BLUE)):
    pg.conditional_formatting.add(f'I{P0}:I{PF}', FormulaRule(
        formula=[f'$I{P0}="{txt}"'], fill=fill(bg), font=Font(bold=True, color=cor),
        stopIfTrue=True))
pg.conditional_formatting.add(f'Q{P0}:Q{PF}', FormulaRule(
    formula=[f'AND($Q{P0}<>"",$Q{P0}>0)'], font=Font(bold=True, color=RED)))
pg.conditional_formatting.add(f'Q{P0}:Q{PF}', FormulaRule(
    formula=[f'AND($Q{P0}<>"",$Q{P0}<0)'], font=Font(bold=True, color=OK)))
pg.conditional_formatting.add(f'L{P0}:L{PF}', FormulaRule(
    formula=[f'AND($L{P0}<>"",$L{P0}>0)'], fill=fill(AMBBG), font=Font(bold=True, color=AMBER)))
for c in range(C_AUX0, C_AUX0 + len(CATEGORIAS) * 2):
    pg.column_dimensions[get_column_letter(c)].hidden = True
r = PF + 3

# ── onde o orçamento erra
R_ERR_T = r
r = titulo_secao(pg, r, NC_PG, 'Onde o orçamento erra',
                 'somando todos os projetos lançados · desvio positivo = gastou mais do que previu')
R_ERR_H = r
r = cab(pg, r, [(1, 3, 'Categoria'), (4, 2, 'Orçado (R$)'), (6, 2, 'Realizado (R$)'),
                (8, 2, 'Desvio (R$)'), (10, 2, 'Desvio (%)'), (12, 2, '% da venda'),
                (14, 5, 'Leitura')], alt=26)
R_ERR0 = r
for k, categoria in enumerate(CATEGORIAS):
    co, cr = get_column_letter(C_AUX0 + k * 2), get_column_letter(C_AUX0 + k * 2 + 1)
    pg.row_dimensions[r].height = 19
    bloco(pg, r, 1, 3, categoria, f=Font(name=F, size=9.5, color=INK), bg=WHITE, al=LEFTI)
    bloco(pg, r, 4, 2, f'=ROUND(SUM({co}${P0}:{co}${PF}),2)', f=F_CALC, bg=CALC,
          al=RIGHT, nf=MOEDA0)
    bloco(pg, r, 6, 2, f'=ROUND(SUM({cr}${P0}:{cr}${PF}),2)', f=F_CALC, bg=CALC,
          al=RIGHT, nf=MOEDA0)
    bloco(pg, r, 8, 2, f'=ROUND($F{r}-$D{r},2)',
          f=Font(name=F, size=9.5, bold=True, color=NAVY2), bg=CALC, al=RIGHT, nf=MOEDA0)
    bloco(pg, r, 10, 2, f'=IF($D{r}=0,"",$F{r}/$D{r}-1)',
          f=Font(name=F, size=10, bold=True, color=NAVY2), bg=CALC, al=CTR,
          nf='+0.0%;-0.0%;0.0%')
    bloco(pg, r, 12, 2, f'=IF($J${PF+1}=0,"",$F{r}/$J${PF+1})', f=F_CALC, bg=CALC,
          al=CTR, nf=PCT1)
    bloco(pg, r, 14, 5,
          f'=IF($D{r}=0,"—",IF($J{r}>0.1,"Estourou mais de 10% — revisar a premissa",'
          f'IF($J{r}<-0.1,"Sobrou mais de 10% — orçamento pode estar gordo",'
          f'"Dentro do previsto")))',
          f=Font(name=F, size=8.5, color=INK), bg=CALC, al=LEFTI)
    r += 1
R_ERRF = r - 1
for txt, bg, cor in (('Estourou', REDBG, RED), ('Sobrou', AMBBG, AMBER), ('Dentro', OKBG, OK)):
    pg.conditional_formatting.add(f'N{R_ERR0}:R{R_ERRF}', FormulaRule(
        formula=[f'LEFT($N{R_ERR0},{len(txt)})="{txt}"'], fill=fill(bg),
        font=Font(bold=True, color=cor), stopIfTrue=True))
r += 1

# ── leituras
r = titulo_secao(pg, r, NC_PG, 'Leituras', 'frases montadas a partir dos números acima')
LEITURAS = [
    (f'="Carteira lançada: "&COUNTA($B${P0}:$B${PF})&" projeto(s), "&'
     f'TEXT($J${PF+1},"R$ #,##0")&" de venda e "&TEXT($M${PF+1},"R$ #,##0")&'
     f'" de margem de contribuição — "&TEXT($N${PF+1},"0.0%")&" sobre a venda."'),
    (f'="Categoria que mais estoura: "&INDEX($A${R_ERR0}:$A${R_ERRF},'
     f'MATCH(MAX($H${R_ERR0}:$H${R_ERRF}),$H${R_ERR0}:$H${R_ERRF},0))&'
     f'", com "&TEXT(MAX($H${R_ERR0}:$H${R_ERRF}),"R$ #,##0")&" acima do orçado. "&'
     f'"Categoria que mais sobra: "&INDEX($A${R_ERR0}:$A${R_ERRF},'
     f'MATCH(MIN($H${R_ERR0}:$H${R_ERRF}),$H${R_ERR0}:$H${R_ERRF},0))&"."'),
    (f'="Margem: "&COUNTIF($R${P0}:$R${PF},"Margem crítica")&" projeto(s) abaixo de 25%, "&'
     f'COUNTIF($R${P0}:$R${PF},"Margem apertada")&" entre 25% e 35% e "&'
     f'COUNTIF($R${P0}:$R${PF},"Margem boa")&" acima de 35%."'),
    (f'="Compras pendentes na carteira: "&TEXT(SUM($L${P0}:$L${PF}),"R$ #,##0")&'
     f'" ainda a comprar. Entrega: "&COUNTIF($I${P0}:$I${PF},"No prazo")&" no prazo, "&'
     f'COUNTIF($I${P0}:$I${PF},"Atrasado")&" atrasado(s), "&'
     f'COUNTIF($I${P0}:$I${PF},"Em produção")&" em produção. Atraso médio: "&'
     f'IF(COUNT($H${P0}:$H${PF})=0,"—",TEXT(AVERAGE($H${P0}:$H${PF}),"0")&" dia(s)")&"."'),
]
for txt in LEITURAS:
    pg.row_dimensions[r].height = 22
    bloco(pg, r, 1, NC_PG, txt, f=Font(name=F, size=9.5, color='2A3744'), bg=WHITE,
          al=LEFTIW, bd=False)
    r += 1
pg.row_dimensions[r].height = 8
r += 1
pg.row_dimensions[r].height = 46
bloco(pg, r, 1, NC_PG,
      '  Como alimentar: a cada projeto novo, botão direito na aba "Ficha Modelo" → Mover ou copiar → Criar uma '
      'cópia, renomeie (ex.: P-2026-042 Maria) e escreva esse mesmo nome numa linha livre da coluna ABA aqui. '
      'Todo o resto se preenche sozinho. Nome errado ou aba inexistente deixa a linha em branco — não quebra nada.',
      f=Font(name=F, size=8.5, color='41505D'), bg=GOLDBG, al=LEFTIW)
R_FIM_PG = r


def estilo(ch, titulo):
    ch.style = None
    ch.y_axis.majorGridlines = ChartLines()
    ch.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(
        defRPr=CharacterProperties(sz=800, solidFill=MUTED)),
        endParaRPr=CharacterProperties(sz=800))])
    ch.x_axis.txPr = ch.y_axis.txPr
    ch.legend.position = 'b'; ch.legend.overlay = False
    ch.title = titulo


g1 = BarChart(); g1.type = 'col'
g1.add_data(Reference(pg, min_col=14, min_row=P0 - 1, max_row=PF), titles_from_data=True)
g1.set_categories(Reference(pg, min_col=3, min_row=P0, max_row=PF))
estilo(g1, 'Margem de contribuição por projeto')
g1.y_axis.numFmt = '0%'
g1.series[0].graphicalProperties.solidFill = GOLD
g1.width, g1.height = 17, 8
pg.add_chart(g1, f'T{R_ERR_T - 20}')

g2 = BarChart(); g2.type = 'bar'
g2.add_data(Reference(pg, min_col=8, min_row=R_ERR_H, max_row=R_ERRF), titles_from_data=True)
g2.set_categories(Reference(pg, min_col=1, min_row=R_ERR0, max_row=R_ERRF))
estilo(g2, 'Desvio do orçamento por categoria (R$)')
g2.y_axis.numFmt = 'R$ #,##0'
g2.series[0].graphicalProperties.solidFill = NAVY2
g2.width, g2.height = 17, 8
pg.add_chart(g2, f'T{R_ERR_T}')
pg.freeze_panes = f'C{P0}'
print_cfg(pg, f'A1:R{R_FIM_PG}')


# ══════════════════════════════════════════════════════════════════════
#  ABA · Listas
# ══════════════════════════════════════════════════════════════════════
ls = wb.create_sheet('Listas')
ls.sheet_view.showGridLines = False
for col, titulo, vals, larg in (('A', 'Equipe (comissões)', EQUIPE, 24),
                                ('B', 'Vendedor', VENDEDORES, 22),
                                ('C', 'Causa do retrabalho', CAUSAS, 30),
                                ('D', 'Categoria de compra', CATEGORIAS_COMPRA, 42),
                                ('E', 'Forma de pagamento', PAGAMENTOS, 22),
                                ('F', 'Status da compra', STATUS, 22)):
    ls.column_dimensions[col].width = larg
    c = ls[f'{col}1']; c.value = titulo
    c.font = font(9, True, WHITE); c.fill = fill(NAVY2); c.alignment = CTR
    for i, v in enumerate(vals, start=2):
        cc = ls[f'{col}{i}']; cc.value = v; cc.font = font(10); cc.border = BOTTOM2
# coluna G: o que entra em cada categoria de compra
ls.column_dimensions['G'].width = 46
c = ls['G1']; c.value = 'O que entra nesta categoria'
c.font = font(9, True, WHITE); c.fill = fill(NAVY2); c.alignment = CTR
for i, cat in enumerate(CATEGORIAS_COMPRA, start=2):
    cc = ls.cell(i, 7, DESCR_CATEGORIA.get(cat, '')); cc.font = font(9, c=MUTED)
    cc.border = BOTTOM2

ls['I1'] = 'Para que servem'
ls['I1'].font = font(9, True, WHITE); ls['I1'].fill = fill(NAVY2); ls['I1'].alignment = CTR
ls.column_dimensions['I'].width = 80
for i, t in enumerate([
    'Coluna A — só quem recebe comissão: marceneiros, ajudantes e o coordenador. Alimenta produção, montagem, coordenação e o consolidado por colaborador.',
    'Se acrescentar alguém aqui, inclua também na tabela "Comissões por colaborador" da ficha, senão a pessoa não aparece no consolidado.',
    'Coluna D — as categorias de compra. Se você renomear uma categoria aqui, renomeie também na tabela do bloco 6 da ficha: o SUMIFS casa pelo texto.',
    'Coluna F — status. "A comprar" NÃO entra no custo realizado, entra na coluna "ainda a comprar". "Comprado (a pagar)" e "Pago" entram no custo.',
    'Coluna C — causas do retrabalho. É a lista mais estratégica da planilha: ela é que vai dizer se o problema está no projeto, na medição, na produção ou no fornecedor.',
], start=2):
    c = ls.cell(i, 9, t); c.font = font(9, c='41505D')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ls.row_dimensions[i].height = 32


# ══════════════════════════════════════════════════════════════════════
#  ABA · Instruções
# ══════════════════════════════════════════════════════════════════════
ins = wb.create_sheet('Instruções', 0)
ins.sheet_view.showGridLines = False
NC_IN = 6
r = faixa_marca(ins, NC_IN, 'COMO USAR ESTA PLANILHA',
                'Custo direto por projeto · orçado × realizado · eficiência do orçamento')
for i, w in enumerate([4, 34, 30, 30, 30, 12], start=1):
    ins.column_dimensions[get_column_letter(i)].width = w
r += 1


def par(row, texto, f=None, h=None, bg=None):
    ins.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC_IN)
    c = ins.cell(row, 1, '   ' + texto)
    c.font = f or font(10, c='2A3744')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if bg:
        for k in range(1, NC_IN + 1):
            ins.cell(row, k).fill = fill(bg)
    ins.row_dimensions[row].height = h or 18
    return row + 1


BLOCOS_IN = [
    ('PARA QUE SERVE', [
        'Medir se o seu orçamento acerta. Cada linha de custo tem ORÇADO e REALIZADO lado a lado, e o desvio entre os dois é a resposta.',
        'Sem a coluna ORÇADO isto seria só um controle de gasto. Com ela, vira uma régua: você descobre que sempre erra ferragem para menos, ou que a logística custa o dobro do que imagina.',
    ]),
    ('O PASSO A PASSO', [
        '1 · Botão direito na aba "Ficha Modelo" → Mover ou copiar → marcar "Criar uma cópia". Renomeie (ex.: P-2026-042 Maria).',
        '2 · Preencha a identificação: cliente, nº, projeto, data de entrada, entrega prevista, vendedor e coordenador.',
        '3 · Lance o valor de venda ORÇADO (proposta) e REALIZADO (fechado). A diferença entre os dois já é informação: é o desconto que você deu.',
        '4 · Preencha as premissas de custo de venda (% de imposto, máquina, comissão e RT). Elas se aplicam sozinhas aos dois cenários.',
        '5 · Liste os ambientes com o valor de cada um e, para cada ambiente, quem produziu e quem montou, com o % de comissão.',
        '6 · No bloco 6, digite o ORÇADO de cada categoria de compra. O realizado virá sozinho do livro.',
        '7 · Conforme for comprando, lance no livro (bloco 8). Uma linha por compra.',
        '8 · Vá ao Painel Geral e escreva o nome da aba nova numa linha livre da coluna ABA.',
    ]),
    ('O LIVRO DE COMPRAS — a parte que dá o dinamismo', [
        'Hoje você compra chapa, amanhã compra outra, semana que vem compra ferragem. Cada compra é UMA LINHA no bloco 8, com: descrição/fornecedor, categoria, data, valor, forma de pagamento e status.',
        'O bloco 6 (Material, terceirizados e logística) NÃO se digita: ele soma o livro por categoria, sozinho. Você só digita o orçado de cada categoria.',
        'O STATUS é o que separa o que já é custo do que ainda não é: "A comprar" NÃO entra no custo realizado — vai para a coluna "Ainda a comprar". "Comprado (a pagar)" e "Pago" entram no custo.',
        'Isso responde três perguntas de uma vez: quanto já custou, quanto ainda vou gastar e quanto devo aos fornecedores. A faixa dourada logo abaixo da margem mostra o custo projetado e a MC projetada.',
        'Você pode inserir quantas linhas quiser no livro. Ele fica no fim da ficha exatamente por isso: nada que o Painel Geral lê está abaixo dele.',
        'Categoria, Forma de pagamento e Status são MENUS SUSPENSOS: clique na célula e escolha. Digitar por fora é permitido, mas a planilha avisa — e o que não estiver escrito igual à lista não entra na soma da categoria.',
        'Lançamento com valor e sem categoria fica destacado em vermelho, e o total do livro avisa quantos estão sem classificar.',
    ]),
    ('A CASCATA DE CÁLCULO — a ordem importa', [
        'Valor de venda  −  custos de venda (impostos, máquina, taxas, comissão de venda, projeto, RT)  =  RECEITA LÍQUIDA',
        'A receita líquida é a base das comissões de coordenação, produção e montagem. Não é o valor de venda. O marceneiro não deve ser comissionado sobre o imposto nem sobre o RT do arquiteto.',
        'O RT tem base própria: incide sobre a venda menos impostos, máquina e taxas de transação — como você pediu.',
        'Depois da receita líquida saem, na ordem: comissões, material, terceirizados, logística e retrabalho. O que sobra é a MARGEM DE CONTRIBUIÇÃO.',
    ]),
    ('COMO FUNCIONAM AS COMISSÕES POR AMBIENTE', [
        'Cada ambiente tem um valor em reais. A planilha calcula o peso dele no projeto e aplica esse peso sobre a receita líquida — essa é a base do ambiente.',
        'Sobre a base do ambiente incidem dois percentuais: o de quem produziu e o de quem montou. Pessoas diferentes em ambientes diferentes, cada uma com seu percentual.',
        'Exemplo: projeto de R$ 90.000 com cozinha R$ 30.000 (33,3%). Se a receita líquida for R$ 76.000, a base da cozinha é R$ 25.333. A 3% de produção, o Jackson recebe R$ 760 por aquele ambiente.',
        'A tabela "Comissões por colaborador" soma tudo sozinha. Ela lista só marceneiros, ajudantes e o coordenador — quem não recebe comissão não aparece.',
        'A soma dos ambientes precisa fechar com o valor de venda. A linha de total avisa quando não fecha.',
    ]),
    ('RETRABALHO — a parte que mais ensina', [
        'Registre o que aconteceu, a causa (menu suspenso) e o custo estimado. Não precisa ser exato: o valor aproximado já serve.',
        'No orçamento você lança a CONTINGÊNCIA prevista. O desvio entre ela e o retrabalho real diz se a sua provisão está no tamanho certo.',
        'Com o tempo, a coluna Causa vira o dado mais valioso da planilha: se metade dos retrabalhos é "erro de medição", o problema não é preço, é processo.',
    ]),
    ('O PAINEL GERAL', [
        'Sete indicadores no topo: projetos, vendido, custo total, ainda a comprar, margem em reais, margem média e percentual de entrega no prazo.',
        'A tabela de projetos puxa cada ficha pelo nome da aba. Nome errado deixa a linha em branco — não quebra.',
        'O bloco "Onde o orçamento erra" soma todos os projetos por categoria e mostra em qual delas você mais estoura. É o retorno direto para ajustar a base de orçamento.',
        'Dois gráficos: margem por projeto e desvio do orçamento por categoria.',
    ]),
    ('CUIDADOS', [
        'A regra de ouro: pode inserir e apagar linhas em qualquer bloco A PARTIR DA LINHA 26. Tudo o que o Painel Geral lê está entre as linhas 6 e 24.',
        'Não mexa nas linhas 1 a 25 da ficha. Se elas mudarem de lugar, o Painel passa a ler a célula errada.',
        'A margem calculada aqui é de CONTRIBUIÇÃO: ela ainda não paga o custo fixo da empresa. Um projeto com 30% de MC não deu 30% de lucro.',
        'O custo de retrabalho é estimado por você. Não sai de nota fiscal — é a sua leitura do prejuízo, e vale mais registrada por alto do que não registrada.',
    ]),
]
for titulo, linhas in BLOCOS_IN:
    r = titulo_secao(ins, r, NC_IN, titulo)
    ins.row_dimensions[r].height = 4
    r += 1
    for t in linhas:
        r = par(r, t, h=30 if len(t) > 115 else (18 if len(t) < 80 else 24))
    ins.row_dimensions[r].height = 8
    r += 1
r = par(r, 'Abas:   Instruções  ·  Painel Geral  ·  Ficha Modelo (duplique esta)  ·  '
           'Exemplo P-2026-041 (preenchido)  ·  Listas',
        f=Font(name=F, size=9, bold=True, color='7A5B17'), h=26, bg=GOLDBG)
print_cfg(ins, f'A1:F{r - 1}', retrato=True)

# ── nomes definidos que alimentam os menus suspensos
from openpyxl.workbook.defined_name import DefinedName
TAM = {'A': len(EQUIPE), 'B': len(VENDEDORES), 'C': len(CAUSAS),
       'D': len(CATEGORIAS_COMPRA), 'E': len(PAGAMENTOS), 'F': len(STATUS)}
for nome, col in NOMES.items():
    wb.defined_names.add(
        DefinedName(nome, attr_text=f"Listas!${col}$2:${col}${TAM[col] + 1}"))
print('nomes definidos:', ', '.join(sorted(NOMES)))

wb.active = 0
SAIDA = '/home/user/valvicorcamentista/painel/planilhas/Valvic_Custo_por_Projeto.xlsx'
wb.save(SAIDA)
print('OK →', SAIDA)
print(f'   painel: projetos {P0}-{PF} · erros {R_ERR0}-{R_ERRF} · fim {R_FIM_PG}')
