#!/usr/bin/env python3
"""Gera o Controle de Patrimônio da Valvic.

Abas: Dashboard · Patrimônio Geral · Movimentações · Conferência Mensal ·
um Termo por colaborador (layout A4) · Modelo em branco · Listas.

Uso:  python3 gerar-controle-patrimonio.py
Saída: Valvic_Controle_Patrimonio.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.comments import Comment

# ─────────────────────────────────────────────── paleta Valvic
NAVY   = '0E2038'
NAVY2  = '16314F'
GOLD   = 'C2A05A'
GOLDS  = 'D8BD80'
GOLDBG = 'F6EDD6'
CREAM  = 'FBFAF7'
INK    = '1B2733'
MUTED  = '6C7785'
LINE   = 'E8E3D8'
LINE2  = 'DFDACD'
OK     = '2F7D4F'
BLUE   = '2F5D8C'
RED    = 'B0413F'
INPUT  = 'FFF9E3'          # célula a preencher
WHITE  = 'FFFFFF'

F = 'Arial'
def font(sz=10, b=False, c=INK, i=False):
    return Font(name=F, size=sz, bold=b, color=c, italic=i)
def fill(c):
    return PatternFill('solid', fgColor=c)
def side(c=LINE, st='thin'):
    return Side(style=st, color=c)
BOX      = Border(left=side(), right=side(), top=side(), bottom=side())
BOTTOM   = Border(bottom=side())
BOTTOM2  = Border(bottom=side(LINE2))
NOBORDER = Border()
GRID   = Border(bottom=side(LINE), left=side(LINE), right=side(LINE))
F9     = Font(name=F, size=9, color=INK)
F9CALC = Font(name=F, size=9, color=NAVY2)
F9COD  = Font(name=F, size=9, bold=True, color=NAVY)
FILL_IN   = PatternFill('solid', fgColor=INPUT)
FILL_CALC = PatternFill('solid', fgColor='F4F6F8')
CTR   = Alignment(horizontal='center', vertical='center')
CTRW  = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT  = Alignment(horizontal='left', vertical='center')
LEFTW = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

MOEDA = 'R$ #,##0.00'
MOEDA0 = 'R$ #,##0'
PCT   = '0.0%'
DATA  = 'DD/MM/YYYY'

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════ dados-base
PESSOAS = [
    # nome, documento, função, vínculo, instrumento
    ('Deivson',        '', 'Coordenador de Produção',   'PJ',  'comodato'),
    ('Samuel',         '', 'Marceneiro Sênior',         'PJ',  'comodato'),
    ('Jackson',        '', 'Marceneiro Pleno',          'PJ',  'comodato'),
    ('Cezar',          '', 'Marceneiro',                'CLT', 'responsabilidade'),
    ('Jomar',          '', 'Marceneiro',                'CLT', 'responsabilidade'),
    ('Davi',           '', 'Ajudante',                  'CLT', 'responsabilidade'),
    ('Jonathan Godoy', '', 'Ajudante Geral',            'CLT', 'responsabilidade'),
    ('Joelson',        '', 'Operador CNC & Máquinas',   'CLT', 'responsabilidade'),
]
CATEGORIAS = ['Máquina fixa', 'Ferramenta elétrica', 'Ferramenta manual',
              'Instrumento de medição', 'Veículo', 'Informática', 'Mobiliário', 'Infraestrutura']
SITUACOES  = ['Em uso', 'Em estoque', 'Em manutenção', 'Extraviado', 'Baixado']
LOCAIS     = ['Fábrica', 'Obra', 'Veículo', 'Escritório', 'Fornecedor (manutenção)']
SIMNAO     = ['Sim', 'Não']

# parque instalado conhecido — valores em branco para preencher
PARQUE = [
    ('Máquina fixa', 'Centro de usinagem CNC Raizen Solid TAF · 380V · 3 eixos', 'Raizen', 'Solid TAF', 10),
    ('Máquina fixa', 'Coladeira de bordas SCM Minimax ME 25 · fita 0,4 mm · cola EVA', 'SCM', 'Minimax ME 25', 10),
    ('Máquina fixa', 'Esquadrejadeira Raizen RZN 3200P', 'Raizen', 'RZN 3200P', 10),
    ('Máquina fixa', 'Compressor de ar', '', '', 10),
    ('Máquina fixa', 'Coletor de pó / exaustão', '', '', 10),
    ('Veículo',      'Chevrolet Montana 2020 · placa QUY4166', 'Chevrolet', 'Montana 2020', 5),
    ('Informática',  'Computador de escritório', '', '', 5),
    ('Informática',  'Tablet do chão de fábrica', '', '', 5),
    ('Ferramenta elétrica', 'Furador de coluna', '', '', 10),
    ('Ferramenta elétrica', 'Etiquetadora', '', '', 10),
]
# kit de referência — exemplos com valores realistas
KIT = [
    ('Ferramenta elétrica', 'Parafusadeira / furadeira de impacto a bateria 18V com 2 baterias', 'Makita', 'DHP482', 890.00, 5, 'PARAF-MAK-DHP482'),
    ('Ferramenta elétrica', 'Serra circular manual 7.1/4"', 'Bosch', 'GKS 150', 690.00, 5, 'SERRA-BOS-GKS150'),
    ('Ferramenta elétrica', 'Tupia manual 1/4"', 'Makita', 'RT0700C', 780.00, 5, 'TUPIA-MAK-RT0700C'),
    ('Ferramenta elétrica', 'Lixadeira orbital 1/4 de folha', 'Bosch', 'GSS 140', 380.00, 5, 'LIXA-BOS-GSS140'),
    ('Ferramenta manual',   'Jogo de formões 4 peças', 'Tramontina', '', 180.00, 10, 'FORMAO-TRA-4P'),
    ('Ferramenta manual',   'Martelo unha 27 mm cabo de fibra', 'Tramontina', '', 75.00, 10, 'MARTELO-TRA-27'),
    ('Ferramenta manual',   'Jogo de chaves de fenda e philips 6 peças', 'Gedore', '', 130.00, 10, 'CHAVES-GED-6P'),
    ('Ferramenta manual',   'Grampo sargento 300 mm — par', 'Vonder', '', 160.00, 10, 'SARGENTO-VON-300'),
    ('Instrumento de medição', 'Trena 5 m', 'Starrett', '', 55.00, 5, 'TRENA-STA-5M'),
    ('Instrumento de medição', 'Esquadro de precisão 300 mm', 'Starrett', '', 120.00, 10, 'ESQUADRO-STA-300'),
    ('Instrumento de medição', 'Nível a laser de linha', 'Bosch', 'GLL 2-10', 520.00, 5, 'LASER-BOS-GLL210'),
    ('Ferramenta manual',   'Maleta / caixa organizadora de ferramentas', 'Tramontina', '', 210.00, 5, 'MALETA-TRA-ORG'),
]
TAXA = {'Máquina fixa': 0.10, 'Ferramenta elétrica': 0.20, 'Ferramenta manual': 0.10,
        'Instrumento de medição': 0.20, 'Veículo': 0.20, 'Informática': 0.20,
        'Mobiliário': 0.10, 'Infraestrutura': 0.10}

# ══════════════════════════════════════════════ helpers de layout
def faixa_marca(ws, ncols, titulo, sub, linha=1):
    """Cabeçalho navy padrão Valvic."""
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
    c = ws.cell(linha, 1, 'VALVIC MARCENARIA')
    c.font = Font(name=F, size=13, bold=True, color=WHITE)
    c.fill = fill(NAVY); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha].height = 26

    ws.merge_cells(start_row=linha+1, start_column=1, end_row=linha+1, end_column=ncols)
    c = ws.cell(linha+1, 1, 'Vargas Decor Ltda   ·   CNPJ 17.269.304/0001-51   ·   Belo Horizonte / MG')
    c.font = Font(name=F, size=8, color='9FB0C4')
    c.fill = fill(NAVY); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha+1].height = 15

    ws.merge_cells(start_row=linha+2, start_column=1, end_row=linha+2, end_column=ncols)
    c = ws.cell(linha+2, 1, titulo)
    c.font = Font(name=F, size=15, bold=True, color=WHITE)
    c.fill = fill(NAVY2); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha+2].height = 28

    ws.merge_cells(start_row=linha+3, start_column=1, end_row=linha+3, end_column=ncols)
    c = ws.cell(linha+3, 1, sub)
    c.font = Font(name=F, size=8.5, color='7A5B17', italic=True)
    c.fill = fill(GOLDBG); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha+3].height = 17
    return linha + 4

def titulo_secao(ws, row, ncols, texto, nota=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, ('  ' + texto.upper()) + (f'          {nota}' if nota else ''))
    c.font = Font(name=F, size=9, bold=True, color=NAVY)
    c.fill = fill(GOLDBG)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 20
    return row + 1

def cab_tabela(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(name=F, size=8.5, bold=True, color=WHITE)
        c.fill = fill(NAVY2); c.alignment = CTRW
        c.border = Border(left=side(NAVY2), right=side(NAVY2), bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1

def print_a4(ws, area, retrato=True, fit_w=1, margens=(0.45, 0.35, 0.4, 0.35)):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait' if retrato else 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = fit_w
    ws.page_setup.fitToHeight = 0
    ws.print_area = area
    l, r, t, b = margens
    ws.page_margins.left, ws.page_margins.right = l, r
    ws.page_margins.top, ws.page_margins.bottom = t, b
    ws.sheet_view.showGridLines = False

# ══════════════════════════════════════════════ ABA · Listas
ls = wb.active; ls.title = 'Listas'
ls.sheet_view.showGridLines = False
ls['A1'] = 'Colaborador'; ls['B1'] = 'Vínculo'; ls['C1'] = 'Função'; ls['D1'] = 'Documento'
for col, w in zip('ABCD', (24, 12, 30, 22)):
    ls.column_dimensions[col].width = w
for c in ls['A1:D1'][0]:
    c.font = font(9, True, WHITE); c.fill = fill(NAVY2); c.alignment = CTR
for i, (nome, doc, func, vinc, _) in enumerate(PESSOAS, start=2):
    ls.cell(i, 1, nome).font = font(10)
    ls.cell(i, 2, vinc).font = font(10)
    ls.cell(i, 3, func).font = font(10)
    ls.cell(i, 4, doc).font = font(10); ls.cell(i, 4).fill = fill(INPUT)
    for j in range(1, 5):
        ls.cell(i, j).border = BOTTOM2

def col_lista(ws, col, titulo, valores, larg=26):
    ws.column_dimensions[col].width = larg
    c = ws[f'{col}1']; c.value = titulo
    c.font = font(9, True, WHITE); c.fill = fill(NAVY2); c.alignment = CTR
    for i, v in enumerate(valores, start=2):
        cc = ws[f'{col}{i}']; cc.value = v; cc.font = font(10); cc.border = BOTTOM2

col_lista(ls, 'F', 'Categoria', CATEGORIAS, 26)
col_lista(ls, 'G', 'Situação', SITUACOES, 18)
col_lista(ls, 'H', 'Localização', LOCAIS, 22)
col_lista(ls, 'I', 'Sim / Não', SIMNAO, 12)
col_lista(ls, 'J', 'Estado de conservação', ['Novo', 'Bom', 'Regular', 'Ruim'], 22)
col_lista(ls, 'P', 'Modelo padrão (SKU)', [k[6] for k in KIT] + [''] * 18, 26)

ls['F13'] = 'Taxas de depreciação sugeridas (% ao ano)'
ls['F13'].font = font(9, True, NAVY)
for i, (cat, tx) in enumerate(TAXA.items(), start=14):
    ls.cell(i, 6, cat).font = font(9)
    cc = ls.cell(i, 7, tx); cc.font = font(9); cc.number_format = PCT
ls['N1'] = 'Data de referência do cálculo de depreciação'
ls['N1'].font = font(9, True, NAVY)
ls['N2'] = '=TODAY()'
ls['N2'].font = font(11, True, NAVY2); ls['N2'].number_format = DATA
ls.column_dimensions['N'].width = 38
ls['F23'] = 'Confirmar as taxas com o contador antes de fechar o balanço.'
ls['F23'].font = font(8, False, MUTED, True)

# ══════════════════════════════════════════════ ABA · Patrimônio Geral
pg = wb.create_sheet('Patrimônio Geral')
pg.sheet_view.showGridLines = False
HDR = ['Código', 'Categoria', 'Descrição completa do item', 'Marca', 'Modelo',
       'Modelo padrão (SKU)', 'Código de etiqueta / nº de série', 'Data de aquisição', 'Fornecedor',
       'Nota fiscal', 'Valor de aquisição', 'Vida útil (anos)', 'Taxa deprec. (% a.a.)', 'Meses de uso',
       'Depreciação acumulada', 'Valor contábil atual', 'Situação', 'Localização',
       'Responsável', 'Vínculo', 'Estado na entrega', 'Data de entrega', 'Termo assinado', 'Observações']
W   = [11, 20, 44, 14, 16, 20, 22, 13, 18, 12, 15, 10, 12, 10, 16, 15, 14, 15, 17, 9, 14, 13, 11, 34]
# mapa de colunas — mudar aqui reflete em todo o arquivo
C_COD, C_CAT, C_DESC, C_MARCA, C_MOD, C_SKU, C_ETIQ = 1, 2, 3, 4, 5, 6, 7
C_DTAQ, C_FORN, C_NF, C_VALOR, C_VIDA, C_TAXA = 8, 9, 10, 11, 12, 13
C_MESES, C_DEPR, C_CONT = 14, 15, 16
C_SIT, C_LOC, C_RESP, C_VINC, C_EST, C_DTENT, C_TERMO, C_OBS = 17, 18, 19, 20, 21, 22, 23, 24

r = faixa_marca(pg, len(HDR), 'CONTROLE DE PATRIMÔNIO',
                'Cadastro único de todos os bens da empresa · preencha apenas as células de fundo creme')
r += 1
HEAD_ROW = r
r = cab_tabela(pg, r, HDR, W)
FIRST = r
LAST  = FIRST + 149          # 150 linhas de cadastro

def linha_item(row, cod, cat, desc, marca, modelo, valor, vida, sku=''):
    pg.cell(row, C_COD, cod).font = font(9, True, NAVY)
    pg.cell(row, C_CAT, cat).font = font(9)
    pg.cell(row, C_DESC, desc).font = font(9)
    pg.cell(row, C_MARCA, marca).font = font(9)
    pg.cell(row, C_MOD, modelo).font = font(9)
    if sku:
        pg.cell(row, C_SKU, sku).font = font(9, True, '7A5B17')
    if valor is not None:
        v = pg.cell(row, C_VALOR, valor); v.font = font(9)
    pg.cell(row, C_VIDA, vida).font = font(9)

# semeia parque instalado + kit de referência
row = FIRST
for i, (cat, desc, marca, modelo, vida) in enumerate(PARQUE, start=1):
    linha_item(row, f'VLV-{i:04d}', cat, desc, marca, modelo, None, vida)
    pg.cell(row, C_SIT, 'Em uso'); pg.cell(row, C_LOC, 'Fábrica')
    row += 1
for j, (cat, desc, marca, modelo, valor, vida, sku) in enumerate(KIT, start=len(PARQUE)+1):
    linha_item(row, f'VLV-{j:04d}', cat, desc, marca, modelo, valor, vida, sku)
    pg.cell(row, C_SIT, 'Em estoque'); pg.cell(row, C_LOC, 'Fábrica')
    row += 1
EXEMPLO_ATE = row - 1

# formata todas as linhas + fórmulas
for row in range(FIRST, LAST + 1):
    pg.row_dimensions[row].height = 22
    for col in range(1, len(HDR) + 1):
        c = pg.cell(row, col)
        if c.value is None:
            c.font = font(9)
        c.border = GRID
        c.alignment = LEFT if col in (C_CAT, C_DESC, C_MARCA, C_MOD, C_ETIQ, C_FORN,
                                      C_SIT, C_LOC, C_RESP, C_EST, C_OBS) else CTR
    for col in (C_COD, C_CAT, C_DESC, C_MARCA, C_MOD, C_SKU, C_ETIQ, C_DTAQ, C_FORN, C_NF,
                C_VALOR, C_VIDA, C_SIT, C_LOC, C_RESP, C_EST, C_DTENT, C_TERMO, C_OBS):
        pg.cell(row, col).fill = FILL_IN
    pg.cell(row, C_DTAQ).number_format = DATA
    pg.cell(row, C_VALOR).number_format = MOEDA
    pg.cell(row, C_DTENT).number_format = DATA
    # fórmulas
    pg.cell(row, C_TAXA, f'=IFERROR(INDEX(Listas!$G$14:$G$21,MATCH($B{row},Listas!$F$14:$F$21,0)),"")')
    pg.cell(row, C_MESES, f'=IF(OR($H{row}="",$L{row}=""),"",MIN($L{row}*12,MAX(0,DATEDIF($H{row},Listas!$N$2,"m"))))')
    pg.cell(row, C_DEPR, f'=IF(OR($K{row}="",$N{row}=""),"",MIN($K{row},ROUND($K{row}*($M{row}/12)*$N{row},2)))')
    pg.cell(row, C_CONT, f'=IF($K{row}="","",$K{row}-IF($O{row}="",0,$O{row}))')
    pg.cell(row, C_VINC, f'=IFERROR(INDEX(Listas!$B$2:$B$40,MATCH($S{row},Listas!$A$2:$A$40,0)),"")')
    for col in (C_TAXA, C_MESES, C_DEPR, C_CONT, C_VINC):
        pg.cell(row, col).font = F9CALC
        pg.cell(row, col).fill = FILL_CALC
    pg.cell(row, C_TAXA).number_format = PCT
    pg.cell(row, C_DEPR).number_format = MOEDA
    pg.cell(row, C_CONT).number_format = MOEDA

# validações
def dv(ws, formula, rng):
    d = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(d); d.add(rng)
DVS = {
    'B': f'=Listas!$F$2:$F${1+len(CATEGORIAS)}',           # categoria
    'F': f'=Listas!$P$2:$P${1+len(KIT)+18}',               # modelo padrão (SKU)
    'Q': f'=Listas!$G$2:$G${1+len(SITUACOES)}',            # situação
    'R': f'=Listas!$H$2:$H${1+len(LOCAIS)}',               # localização
    'S': f'=Listas!$A$2:$A${1+len(PESSOAS)}',              # responsável
    'U': '=Listas!$J$2:$J$5',                              # estado na entrega
    'W': '=Listas!$I$2:$I$3',                              # termo assinado
}
for col, f_ in DVS.items():
    dv(pg, f_, f'{col}{FIRST}:{col}{LAST}')

pg.freeze_panes = f'D{FIRST}'
pg.auto_filter.ref = f'A{HEAD_ROW}:X{LAST}'
pg.cell(HEAD_ROW, C_COD).comment = Comment(
    'Cada item recebe um código sequencial VLV-0001, VLV-0002…\n'
    'Nunca reaproveitar código de item baixado.', 'Valvic', width=260, height=70)
pg.cell(HEAD_ROW, C_TAXA).comment = Comment(
    'Preenchida automaticamente pela categoria, com base nas taxas da aba Listas.\n'
    'Confirmar as taxas com o contador.', 'Valvic', width=260, height=70)

# legenda
lg = LAST + 2
pg.merge_cells(start_row=lg, start_column=1, end_row=lg, end_column=8)
c = pg.cell(lg, 1, '  COMO USAR ESTA ABA')
c.font = font(9, True, NAVY); c.fill = fill(GOLDBG)
for k, texto in enumerate([
    'Células de fundo creme são de preenchimento manual. Células de fundo cinza-azulado são calculadas — não editar.',
    'As dez primeiras linhas são o parque instalado já conhecido: complete data de aquisição, valor, fornecedor e nota fiscal.',
    'Uma linha por unidade física, sempre — cinco parafusadeiras iguais são cinco linhas, com cinco códigos. É o que permite saber quem está com qual.',
    'Itens idênticos compartilham o mesmo Modelo padrão (SKU). É por ele que o Dashboard e o Kit Padrão agrupam as unidades.',
    'As doze linhas seguintes são um kit de referência com valores de exemplo — ajuste conforme a compra real.',
    'Depreciação linear: valor de aquisição × taxa anual ÷ 12 × meses de uso, limitada ao valor de aquisição.',
    'Ao ceder um item, preencha Responsável, Estado na entrega e Data de entrega, e gere o termo na aba da pessoa.',
], start=1):
    cc = pg.cell(lg + k, 1, f'{k}.  {texto}')
    cc.font = font(9, c='41505D')
    pg.merge_cells(start_row=lg + k, start_column=1, end_row=lg + k, end_column=8)

print_a4(pg, f'A1:X{LAST}', retrato=False, fit_w=1)

# ══════════════════════════════════════════════ ABA · Dashboard
db = wb.create_sheet('Dashboard', 0)
db.sheet_view.showGridLines = False
NC = 12
for i, w in enumerate([3, 26, 15, 15, 4, 26, 15, 15, 4, 24, 14, 14], start=1):
    db.column_dimensions[get_column_letter(i)].width = w

r = faixa_marca(db, NC, 'PAINEL DE PATRIMÔNIO',
                'Números calculados automaticamente a partir da aba Patrimônio Geral')
r += 1

RNG = lambda col: f"'Patrimônio Geral'!${col}${FIRST}:${col}${LAST}"

def kpi(row, col, rotulo, formula, fmt, cor=NAVY, larg=3):
    db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + larg - 1)
    c = db.cell(row, col, rotulo.upper())
    c.font = font(8, True, MUTED); c.fill = fill(WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = Border(top=side(LINE), left=side(GOLD, 'medium'), right=side(LINE))
    db.row_dimensions[row].height = 17
    db.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + larg - 1)
    v = db.cell(row + 1, col, formula)
    v.font = Font(name=F, size=18, bold=True, color=cor); v.fill = fill(WHITE)
    v.number_format = fmt
    v.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    v.border = Border(bottom=side(LINE), left=side(GOLD, 'medium'), right=side(LINE))
    db.row_dimensions[row + 1].height = 30

r = titulo_secao(db, r, NC, 'Posição patrimonial')
r += 1
kpi(r, 2, 'Itens cadastrados', f'=COUNTIF({RNG("A")},"?*")', '#,##0')
kpi(r, 6, 'Valor de aquisição', f'=SUM({RNG("K")})', MOEDA0, NAVY)
kpi(r, 10, 'Valor contábil atual', f'=SUM({RNG("P")})', MOEDA0, OK)
r += 3
kpi(r, 2, 'Depreciação acumulada', f'=SUM({RNG("O")})', MOEDA0, RED)
kpi(r, 6, 'Itens cedidos a colaboradores', f'=COUNTIF({RNG("S")},"?*")', '#,##0', BLUE)
kpi(r, 10, '% do valor já depreciado',
    f'=IF(SUM({RNG("K")})=0,0,SUM({RNG("O")})/SUM({RNG("K")}))', PCT, RED)
r += 3

r = titulo_secao(db, r, NC, 'Termos de responsabilidade')
r += 1
kpi(r, 2, 'Termos assinados', f'=COUNTIF({RNG("W")},"Sim")', '#,##0', OK)
kpi(r, 6, 'Itens cedidos sem termo',
    f'=COUNTIFS({RNG("S")},"?*",{RNG("W")},"<>Sim")', '#,##0', RED)
kpi(r, 10, 'Itens em manutenção ou extraviados',
    f'=COUNTIF({RNG("Q")},"Em manutenção")+COUNTIF({RNG("Q")},"Extraviado")', '#,##0', RED)
r += 3

# ── tabelas
def tabela(row, col, titulo, chaves, formulas, larg=(26, 15, 15), cabs=('', 'Itens', 'Valor')):
    db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    c = db.cell(row, col, '  ' + titulo.upper())
    c.font = font(8.5, True, WHITE); c.fill = fill(NAVY2); c.alignment = LEFT
    db.row_dimensions[row].height = 19
    for i, h in enumerate(cabs):
        cc = db.cell(row + 1, col + i, h)
        cc.font = font(8, True, NAVY); cc.fill = fill(GOLDBG)
        cc.alignment = CTR if i else LEFT
        cc.border = Border(bottom=side(GOLD))
    for k, chave in enumerate(chaves):
        rr = row + 2 + k
        a = db.cell(rr, col, chave); a.font = font(9); a.alignment = LEFT
        b = db.cell(rr, col + 1, formulas[0].format(k=chave, r=rr, c=col)); b.font = font(9); b.alignment = CTR
        b.number_format = '#,##0'
        d = db.cell(rr, col + 2, formulas[1].format(k=chave, r=rr, c=col)); d.font = font(9, True, NAVY)
        d.number_format = MOEDA0
        for cc in (a, b, d):
            cc.border = BOTTOM2
        db.row_dimensions[rr].height = 17
    return row + 2, row + 1 + len(chaves)      # (primeira, última) linha de dados

r = titulo_secao(db, r, NC, 'Distribuição do patrimônio')
r += 1
blocos = [
    (2,  CATEGORIAS,              tabela(r, 2, 'Por categoria', CATEGORIAS,
           [f'=COUNTIF({RNG("B")},$B{{r}})', f'=SUMIF({RNG("B")},$B{{r}},{RNG("K")})'])),
    (6,  [p[0] for p in PESSOAS], tabela(r, 6, 'Por responsável', [p[0] for p in PESSOAS],
           [f'=COUNTIF({RNG("S")},$F{{r}})', f'=SUMIF({RNG("S")},$F{{r}},{RNG("K")})'])),
    (10, SITUACOES,               tabela(r, 10, 'Por situação', SITUACOES,
           [f'=COUNTIF({RNG("Q")},$J{{r}})', f'=SUMIF({RNG("Q")},$J{{r}},{RNG("K")})'])),
]
r = max(fim for _, _, (_, fim) in blocos) + 1

# totais — cada bloco soma exatamente a sua própria faixa de linhas
for col, chaves, (ini, fim) in blocos:
    a = db.cell(r, col, 'Total'); a.font = font(9, True, WHITE); a.fill = fill(NAVY)
    a.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    Lq = get_column_letter(col + 1)
    b = db.cell(r, col + 1, f'=SUM({Lq}{ini}:{Lq}{fim})')
    b.font = font(9, True, GOLDS); b.fill = fill(NAVY); b.alignment = CTR; b.number_format = '#,##0'
    Lv = get_column_letter(col + 2)
    d = db.cell(r, col + 2, f'=SUM({Lv}{ini}:{Lv}{fim})')
    d.font = font(9, True, GOLDS); d.fill = fill(NAVY); d.number_format = MOEDA0
    d.alignment = RIGHT
    db.row_dimensions[r].height = 19
r += 2

r = titulo_secao(db, r, NC, 'Itens por modelo padrão', 'unidades idênticas agrupadas pelo SKU')
r += 1
db.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = db.cell(r, 2, '  MODELO PADRÃO'); c.font = font(8.5, True, WHITE); c.fill = fill(NAVY2); c.alignment = LEFT
for i, h in enumerate(('Unidades', 'Cedidas', 'Disponíveis', 'Valor total'), start=6):
    cc = db.cell(r, i, h); cc.font = font(8, True, NAVY); cc.fill = fill(GOLDBG); cc.alignment = CTR
    cc.border = Border(bottom=side(GOLD))
db.row_dimensions[r].height = 19
MFI = r + 1
for i, k in enumerate(KIT):
    rr = MFI + i
    db.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
    a = db.cell(rr, 2, k[6]); a.font = font(9, c='7A5B17'); a.alignment = LEFT
    db.cell(rr, 6, f'=COUNTIF({RNG("F")},$B{rr})').number_format = '#,##0'
    db.cell(rr, 7, f'=COUNTIFS({RNG("F")},$B{rr},{RNG("S")},"?*")').number_format = '#,##0'
    db.cell(rr, 8, f'=$F{rr}-$G{rr}').number_format = '#,##0'
    db.cell(rr, 9, f'=SUMIF({RNG("F")},$B{rr},{RNG("K")})').number_format = MOEDA0
    for col in range(6, 10):
        cc = db.cell(rr, col); cc.font = font(9, c=NAVY2 if col < 9 else NAVY)
        cc.alignment = CTR if col < 9 else RIGHT; cc.border = BOTTOM2
    a.border = BOTTOM2
    db.row_dimensions[rr].height = 17
MLA = MFI + len(KIT) - 1
r = MLA + 1
db.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
a = db.cell(r, 2, 'Total do kit de referência'); a.font = font(9, True, WHITE); a.fill = fill(NAVY)
a.alignment = Alignment(horizontal='left', vertical='center', indent=1)
for col, L in ((6, 'F'), (7, 'G'), (8, 'H'), (9, 'I')):
    cc = db.cell(r, col, f'=SUM({L}{MFI}:{L}{MLA})')
    cc.font = font(9, True, GOLDS); cc.fill = fill(NAVY)
    cc.alignment = CTR if col < 9 else RIGHT
    cc.number_format = MOEDA0 if col == 9 else '#,##0'
for col in range(2, 10):
    db.cell(r, col).fill = fill(NAVY)
db.row_dimensions[r].height = 20
r += 2

r = titulo_secao(db, r, NC, 'Como este painel funciona')
r += 1
for texto in [
    'Todos os números vêm da aba Patrimônio Geral e se atualizam sozinhos — nada aqui precisa ser digitado.',
    'A depreciação é linear e calculada mês a mês desde a data de aquisição, limitada ao valor do bem.',
    '"Itens cedidos sem termo" é o indicador de risco: todo item nas mãos de alguém precisa de termo assinado.',
    'A conferência mensal do Deivson alimenta a aba Conferência Mensal e mantém este painel confiável.',
    'Cada unidade física tem a sua própria linha no cadastro. O agrupamento de itens iguais é feito pelo Modelo padrão (SKU), nunca por uma coluna de quantidade — é o que preserva o rastreio de quem está com qual.',
]:
    db.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
    c = db.cell(r, 2, '•  ' + texto); c.font = font(9, c='41505D'); c.alignment = LEFT
    db.row_dimensions[r].height = 16
    r += 1

print_a4(db, f'A1:{get_column_letter(NC)}{r}', retrato=False)

# ══════════════════════════════════════════════ ABA · Movimentações
mv = wb.create_sheet('Movimentações')
mv.sheet_view.showGridLines = False
H = ['Data', 'Código do item', 'Descrição', 'Movimento', 'De', 'Para',
     'Estado do item', 'Documento / termo', 'Registrado por', 'Observações']
Wm = [13, 14, 44, 18, 20, 20, 16, 20, 18, 40]
r = faixa_marca(mv, len(H), 'MOVIMENTAÇÕES',
                'Histórico de entregas, devoluções, manutenções e baixas — uma linha por evento')
r += 1
r = cab_tabela(mv, r, H, Wm)
MFIRST = r; MLAST = r + 99
for row in range(MFIRST, MLAST + 1):
    mv.row_dimensions[row].height = 20
    for col in range(1, len(H) + 1):
        c = mv.cell(row, col); c.font = F9; c.fill = FILL_IN
        c.border = GRID
        c.alignment = LEFT if col in (3, 5, 6, 8, 9, 10) else CTR
    mv.cell(row, 1).number_format = DATA
    mv.cell(row, 3, f'=IFERROR(INDEX({RNG("C")},MATCH($B{row},{RNG("A")},0)),"")')
    mv.cell(row, 3).font = F9CALC; mv.cell(row, 3).fill = FILL_CALC
dv(mv, '=Listas!$L$2:$L$7', f'D{MFIRST}:D{MLAST}')
col_lista(ls, 'L', 'Movimento', ['Entrega', 'Devolução', 'Transferência',
                                 'Envio a manutenção', 'Retorno de manutenção', 'Baixa'], 24)
dv(mv, f'=Listas!$A$2:$A${1+len(PESSOAS)}', f'E{MFIRST}:F{MLAST}')
mv.freeze_panes = f'A{MFIRST}'
mv.auto_filter.ref = f'A{MFIRST-1}:J{MLAST}'
print_a4(mv, f'A1:J{MLAST}', retrato=False)

# ══════════════════════════════════════════════ ABA · Conferência Mensal
cf = wb.create_sheet('Conferência Mensal')
cf.sheet_view.showGridLines = False
H = ['Código', 'Descrição', 'Responsável', 'Localização', 'Conferido?',
     'Estado encontrado', 'Divergência', 'Ação tomada', 'Data', 'Conferido por']
Wc = [11, 46, 18, 16, 12, 18, 26, 30, 13, 18]
r = faixa_marca(cf, len(H), 'CONFERÊNCIA MENSAL',
                'Checagem física do patrimônio · responsável pela conferência: Deivson · periodicidade mensal')
r += 1
r = cab_tabela(cf, r, H, Wc)
CFIRST = r; CLAST = r + 79
for row in range(CFIRST, CLAST + 1):
    cf.row_dimensions[row].height = 20
    for col in range(1, len(H) + 1):
        c = cf.cell(row, col); c.font = F9; c.fill = FILL_IN
        c.border = GRID
        c.alignment = LEFT if col in (2, 3, 4, 6, 7, 8, 10) else CTR
    for col, src in ((2, 'C'), (3, 'S'), (4, 'R')):   # descrição · responsável · localização
        cf.cell(row, col, f'=IFERROR(INDEX({RNG(src)},MATCH($A{row},{RNG("A")},0)),"")')
        cf.cell(row, col).font = F9CALC; cf.cell(row, col).fill = FILL_CALC
    cf.cell(row, 9).number_format = DATA
dv(cf, '=Listas!$I$2:$I$3', f'E{CFIRST}:E{CLAST}')
dv(cf, '=Listas!$J$2:$J$5', f'F{CFIRST}:F{CLAST}')
cf.freeze_panes = f'A{CFIRST}'
print_a4(cf, f'A1:J{CLAST}', retrato=False)

# ══════════════════════════════════════════════ ABA · Kit Padrão
kp = wb.create_sheet('Kit Padrão')
kp.sheet_view.showGridLines = False
H = ['Modelo padrão (SKU)', 'Item', 'Categoria', 'Qtd por kit', 'Valor unitário',
     'Valor por kit', 'Unidades cadastradas', 'Já cedidas', 'Disponíveis', 'Kits completos possíveis']
Wk = [24, 44, 20, 11, 14, 14, 14, 11, 12, 15]
r = faixa_marca(kp, len(H), 'KIT PADRÃO',
                'O que compõe um kit completo · e quantos kits o estoque atual permite montar')
r += 1
r = cab_tabela(kp, r, H, Wk)
KFIRST = r
SKU_R = lambda col: f"'Patrimônio Geral'!${col}${FIRST}:${col}${LAST}"
for i, (cat, desc, marca, modelo, valor, vida, sku) in enumerate(KIT):
    row = r + i
    kp.row_dimensions[row].height = 21
    kp.cell(row, 1, sku).font = font(9, True, '7A5B17')
    kp.cell(row, 2, desc).font = F9
    kp.cell(row, 3, cat).font = F9
    kp.cell(row, 4, 1).font = F9
    kp.cell(row, 5, valor).font = F9
    kp.cell(row, 6, f'=IF(OR($D{row}="",$E{row}=""),"",$D{row}*$E{row})').font = F9CALC
    kp.cell(row, 7, f'=COUNTIF({SKU_R("F")},$A{row})').font = F9CALC
    kp.cell(row, 8, f'=COUNTIFS({SKU_R("F")},$A{row},{SKU_R("S")},"?*")').font = F9CALC
    kp.cell(row, 9, f'=$G{row}-$H{row}').font = F9CALC
    kp.cell(row, 10, f'=IF($D{row}=0,"",INT($G{row}/$D{row}))').font = F9CALC
    for col in range(1, len(H) + 1):
        c = kp.cell(row, col)
        c.border = GRID
        c.alignment = LEFT if col in (1, 2, 3) else CTR
        c.fill = FILL_IN if col in (1, 2, 3, 4, 5) else FILL_CALC
    kp.cell(row, 5).number_format = MOEDA
    kp.cell(row, 6).number_format = MOEDA
KLAST = r + len(KIT) - 1
r = KLAST + 1
kp.cell(r, 1, 'TOTAL').font = font(9, True, WHITE); kp.cell(r, 1).fill = fill(NAVY)
kp.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
kp.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
for col, formula, fmt in ((6, f'=SUM(F{KFIRST}:F{KLAST})', MOEDA),
                          (7, f'=SUM(G{KFIRST}:G{KLAST})', '#,##0'),
                          (8, f'=SUM(H{KFIRST}:H{KLAST})', '#,##0'),
                          (9, f'=SUM(I{KFIRST}:I{KLAST})', '#,##0'),
                          (10, f'=MIN(J{KFIRST}:J{KLAST})', '#,##0')):
    c = kp.cell(r, col, formula)
    c.font = font(9, True, GOLDS); c.fill = fill(NAVY); c.alignment = CTR; c.number_format = fmt
for col in range(1, len(H) + 1):
    kp.cell(r, col).fill = fill(NAVY)
kp.row_dimensions[r].height = 22
r += 2
kp.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(H))
c = kp.cell(r, 1, '  COMO LER')
c.font = font(9, True, NAVY); c.fill = fill(GOLDBG)
for k, texto in enumerate([
    'Esta aba define o que é um kit completo. A quantidade vive aqui — no desenho do kit — e não no cadastro de bens.',
    '"Unidades cadastradas" conta quantas unidades daquele SKU existem na aba Patrimônio Geral. "Já cedidas" conta as que têm responsável.',
    '"Kits completos possíveis" é quantos kits inteiros o estoque atual permite montar — o menor valor da coluna manda no total.',
    'Para incluir um item novo no kit: cadastre o SKU na aba Listas, acrescente a linha aqui e registre cada unidade física no Patrimônio Geral.',
], start=1):
    kp.merge_cells(start_row=r + k, start_column=1, end_row=r + k, end_column=len(H))
    cc = kp.cell(r + k, 1, f'{k}.  {texto}')
    cc.font = font(9, c='41505D')
print_a4(kp, f'A1:J{r + 4}', retrato=False)

# ══════════════════════════════════════════════ TERMOS
CLAUSULAS_CLT = [
    ('Objeto', 'A EMPRESA entrega ao COLABORADOR, em caráter de cessão para uso exclusivamente '
     'profissional, as ferramentas e equipamentos relacionados no quadro acima, que permanecem sendo '
     'de propriedade integral da EMPRESA.'),
    ('Uso', 'Os itens destinam-se exclusivamente à execução das atividades de trabalho. É vedado '
     'emprestá-los, cedê-los, aliená-los ou utilizá-los em serviços particulares ou para terceiros.'),
    ('Guarda', 'Os itens NÃO podem ser levados para a residência do COLABORADOR. A guarda deve ocorrer '
     'nas dependências da EMPRESA ou no canteiro de obra, em local seguro indicado pela coordenação.'),
    ('Desgaste natural', 'O desgaste decorrente do uso normal e a manutenção preventiva correm por conta '
     'da EMPRESA, sem qualquer ônus para o COLABORADOR.'),
    ('Dano por uso indevido', 'Havendo dano decorrente de uso indevido, imprudência ou negligência '
     'comprovada, o COLABORADOR se compromete a REPOR o item por outro de mesma especificação e em '
     'estado equivalente, no prazo de 30 (trinta) dias, adquirindo-o por sua própria conta e da forma '
     'que lhe for mais conveniente. Não havendo reposição no prazo, e MEDIANTE AUTORIZAÇÃO EXPRESSA E '
     'ESPECÍFICA do COLABORADOR, o valor poderá ser descontado em folha, nos termos do art. 462, §1º, da CLT.'),
    ('Furto ou roubo', 'Em caso de furto ou roubo, o COLABORADOR deve comunicar imediatamente a '
     'coordenação e registrar Boletim de Ocorrência em até 48 (quarenta e oito) horas. Comprovado o '
     'fato, não haverá qualquer ônus para o COLABORADOR.'),
    ('Extravio', 'O extravio sem registro de Boletim de Ocorrência equipara-se, para os efeitos deste '
     'termo, ao dano por negligência previsto na cláusula 5.'),
    ('Devolução', 'Os itens devem ser devolvidos sempre que solicitados pela EMPRESA e, '
     'obrigatoriamente, no ato do desligamento, mediante registro no quadro de devolução deste termo.'),
    ('Registro fotográfico', 'O registro fotográfico do estado dos itens, feito na data da entrega, '
     'integra este termo para todos os efeitos.'),
]
CLAUSULAS_PJ = [
    ('Objeto e natureza', 'A COMODANTE entrega à COMODATÁRIA, em COMODATO GRATUITO, nos termos dos '
     'artigos 579 e seguintes do Código Civil, as ferramentas e equipamentos relacionados no quadro '
     'acima, que permanecem sendo de propriedade integral da COMODANTE.'),
    ('Uso', 'Os itens destinam-se exclusivamente à execução dos serviços contratados. É vedado '
     'emprestá-los, sublocá-los, cedê-los ou aliená-los, sob qualquer título.'),
    ('Guarda', 'Os itens NÃO podem ser levados para a residência da COMODATÁRIA. A guarda deve ocorrer '
     'nas dependências da COMODANTE ou no canteiro de obra, em local seguro indicado pela coordenação.'),
    ('Conservação', 'A COMODATÁRIA obriga-se a conservar os itens como se seus fossem, respondendo pela '
     'guarda e pelo uso adequado, nos termos do art. 582 do Código Civil.'),
    ('Desgaste natural', 'O desgaste decorrente do uso normal e a manutenção preventiva correm por conta '
     'da COMODANTE, sem ônus para a COMODATÁRIA.'),
    ('Dano ou perda', 'Havendo dano por uso indevido, negligência ou perda, a COMODATÁRIA se compromete a '
     'REPOR o item por outro de mesma especificação e em estado equivalente, no prazo de 30 (trinta) dias, '
     'adquirindo-o por sua própria conta e da forma que lhe for mais conveniente.'),
    ('Furto ou roubo', 'Em caso de furto ou roubo, a COMODATÁRIA deve comunicar imediatamente a '
     'coordenação e registrar Boletim de Ocorrência em até 48 (quarenta e oito) horas. Comprovado o '
     'fato, não haverá ônus de reposição.'),
    ('Restituição', 'Os itens devem ser restituídos sempre que solicitados pela COMODANTE e, '
     'obrigatoriamente, ao término da prestação de serviços, mediante registro no quadro de devolução.'),
    ('Registro fotográfico', 'O registro fotográfico do estado dos itens, feito na data da entrega, '
     'integra este termo para todos os efeitos.'),
]

def aba_termo(nome_aba, pessoa, funcao, vinculo, instrumento):
    ws = wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False
    NCT = 8
    for i, w in enumerate([9, 34, 12, 15, 9, 6, 5, 6], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    clt = instrumento == 'responsabilidade'
    titulo = 'TERMO DE ENTREGA E RESPONSABILIDADE' if clt else 'TERMO DE COMODATO DE FERRAMENTAS'
    sub = ('Ferramentas e equipamentos cedidos para uso profissional · vínculo CLT' if clt
           else 'Ferramentas e equipamentos cedidos em comodato · prestador de serviços PJ')
    r = faixa_marca(ws, NCT, titulo, sub)
    r += 1

    # identificação
    r = titulo_secao(ws, r, NCT, 'Identificação')
    campos = [('Nome completo', pessoa, 'B'), ('CPF / CNPJ', '', 'B'),
              ('Função', funcao, 'B'), ('Vínculo', vinculo, 'B'),
              ('Data da entrega', '', 'B'), ('Local de guarda', 'Fábrica / canteiro de obra', 'B')]
    for i in range(0, len(campos), 2):
        for j, (rot, val, _) in enumerate(campos[i:i + 2]):
            col = 1 + j * 4
            c = ws.cell(r, col, rot)
            c.font = font(8, True, MUTED); c.alignment = LEFT
            ws.merge_cells(start_row=r, start_column=col + 1, end_row=r, end_column=col + 3)
            v = ws.cell(r, col + 1, val)
            v.font = font(10, True, NAVY); v.alignment = LEFT
            v.fill = fill(INPUT if not val else WHITE)
            v.border = Border(bottom=side(LINE2))
        ws.row_dimensions[r].height = 21
        r += 1
    r += 1

    # itens
    r = titulo_secao(ws, r, NCT, 'Relação de itens entregues',
                     'preencha o código — descrição, marca e valor vêm da aba Patrimônio Geral')
    HT = ['Código', 'Descrição do item', 'Marca', 'Etiqueta / nº série', 'Estado', 'Valor de referência']
    for i, h in enumerate(HT, start=1):
        cc = ws.cell(r, i if i < 6 else 6, h)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    for i in range(1, NCT + 1):
        cc = ws.cell(r, i)
        cc.font = font(8.5, True, WHITE); cc.fill = fill(NAVY2); cc.alignment = CTRW
        cc.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[r].height = 24
    r += 1
    TFIRST = r
    for k in range(14):
        row = r + k
        ws.row_dimensions[row].height = 20
        ws.cell(row, 1).fill = FILL_IN
        ws.cell(row, 1).font = F9COD; ws.cell(row, 1).alignment = CTR
        ws.cell(row, 2, f'=IFERROR(INDEX({RNG("C")},MATCH($A{row},{RNG("A")},0)),"")')
        ws.cell(row, 3, f'=IFERROR(INDEX({RNG("D")},MATCH($A{row},{RNG("A")},0)),"")')
        ws.cell(row, 4, f'=IFERROR(INDEX({RNG("G")},MATCH($A{row},{RNG("A")},0)),"")')
        ws.cell(row, 5).fill = FILL_IN; ws.cell(row, 5).alignment = CTR
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        ws.cell(row, 6, f'=IFERROR(INDEX({RNG("K")},MATCH($A{row},{RNG("A")},0)),"")')
        ws.cell(row, 6).number_format = MOEDA
        for col in range(1, NCT + 1):
            cc = ws.cell(row, col)
            if col in (2, 3, 4, 6):
                cc.font = F9CALC
            elif col != 1:
                cc.font = F9
            cc.border = GRID
            if col in (2, 3, 4):
                cc.alignment = LEFT
            elif col >= 6:
                cc.alignment = RIGHT
    TLAST = r + 13
    dv(ws, f"='Patrimônio Geral'!$A${FIRST}:$A${LAST}", f'A{TFIRST}:A{TLAST}')
    dv(ws, '=Listas!$J$2:$J$5', f'E{TFIRST}:E{TLAST}')
    r = TLAST + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, 'TOTAL DE REFERÊNCIA DOS ITENS ENTREGUES')
    c.font = font(9, True, WHITE); c.fill = fill(NAVY); c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    t = ws.cell(r, 6, f'=SUM(F{TFIRST}:F{TLAST})')
    t.font = font(11, True, GOLDS); t.fill = fill(NAVY); t.number_format = MOEDA; t.alignment = RIGHT
    ws.row_dimensions[r].height = 22
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCT)
    c = ws.cell(r, 1, 'O valor de referência serve apenas para identificar o bem. Não constitui caução, '
                      'depósito, garantia nem autorização prévia de desconto.')
    c.font = font(8, False, MUTED, True); c.alignment = LEFTW
    ws.row_dimensions[r].height = 15
    r += 2

    # cláusulas
    r = titulo_secao(ws, r, NCT, 'Condições')
    clausulas = CLAUSULAS_CLT if clt else CLAUSULAS_PJ
    for i, (tit, txt) in enumerate(clausulas, start=1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCT)
        c = ws.cell(r, 1, f'{i}.  {tit.upper()} — {txt}')
        c.font = font(8.5, c='33414F'); c.alignment = LEFTW
        n = len(f'{i}. {tit} — {txt}')
        ws.row_dimensions[r].height = max(24, 11.5 * ((n // 132) + 1))
        r += 1
    r += 1

    # declaração e assinaturas
    r = titulo_secao(ws, r, NCT, 'Declaração e assinaturas')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCT)
    quem = 'o COLABORADOR' if clt else 'a COMODATÁRIA'
    c = ws.cell(r, 1, f'Declaro ter recebido os itens acima relacionados, conferido seu estado de '
                      f'conservação e estar ciente das condições deste termo, que assino em duas vias '
                      f'de igual teor. Fica {quem} responsável pela guarda e pelo uso adequado dos itens '
                      f'enquanto estiverem sob sua posse.')
    c.font = font(9, c='33414F'); c.alignment = LEFTW
    ws.row_dimensions[r].height = 34
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    for col in (1, 5):
        cc = ws.cell(r, col); cc.border = Border(top=side(INK))
    ws.row_dimensions[r].height = 8
    r += 1
    a = ws.cell(r, 1, pessoa); a.font = font(9, True, NAVY); a.alignment = CTR
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    b = ws.cell(r, 5, 'Valvic Marcenaria — Vargas Decor Ltda'); b.font = font(9, True, NAVY); b.alignment = CTR
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    r += 1
    a = ws.cell(r, 1, 'CPF / CNPJ' if clt else 'CNPJ'); a.font = font(8, c=MUTED); a.alignment = CTR
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    b = ws.cell(r, 5, 'Representante legal'); b.font = font(8, c=MUTED); b.alignment = CTR
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCT)
    c = ws.cell(r, 1, 'Belo Horizonte / MG,  ______  de  ____________________  de  __________')
    c.font = font(9, c='33414F'); c.alignment = CTR
    ws.row_dimensions[r].height = 24
    r += 2

    # devolução
    r = titulo_secao(ws, r, NCT, 'Registro de devolução', 'preencher no ato da devolução ou do desligamento')
    HD = ['Data', 'Itens devolvidos', 'Estado na devolução', 'Pendências', 'Recebido por']
    for i, h in enumerate(HD, start=1):
        col = {1: 1, 2: 2, 3: 4, 4: 5, 5: 7}[i]
        ws.cell(r, col, h)
    for merge in [(2, 3), (5, 6), (7, 8)]:
        ws.merge_cells(start_row=r, start_column=merge[0], end_row=r, end_column=merge[1])
    for i in range(1, NCT + 1):
        cc = ws.cell(r, i)
        cc.font = font(8.5, True, WHITE); cc.fill = fill(NAVY2); cc.alignment = CTRW
        cc.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[r].height = 22
    r += 1
    for k in range(3):
        row = r + k
        ws.row_dimensions[row].height = 22
        for merge in [(2, 3), (5, 6), (7, 8)]:
            ws.merge_cells(start_row=row, start_column=merge[0], end_row=row, end_column=merge[1])
        for col in range(1, NCT + 1):
            cc = ws.cell(row, col)
            cc.fill = FILL_IN; cc.font = F9
            cc.border = GRID
            cc.alignment = LEFT
        ws.cell(row, 1).number_format = DATA
    r += 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCT)
    c = ws.cell(r, 1, 'Anexo obrigatório: registro fotográfico do estado dos itens na data da entrega. '
                      'Sem o anexo, este termo não sustenta pedido de reposição.')
    c.font = font(8, False, RED, True); c.alignment = LEFTW
    ws.row_dimensions[r].height = 15

    print_a4(ws, f'A1:H{r}', retrato=True, fit_w=1)
    return ws

for nome, doc, func, vinc, instr in PESSOAS:
    curto = nome.split()[0] if nome != 'Jonathan Godoy' else 'J. Godoy'
    aba_termo(f'Termo · {curto}', nome, func, vinc, instr)
aba_termo('Termo · MODELO CLT', '', '', 'CLT', 'responsabilidade')
aba_termo('Termo · MODELO PJ', '', '', 'PJ', 'comodato')

# ordem final das abas
ordem = ['Dashboard', 'Patrimônio Geral', 'Kit Padrão', 'Movimentações', 'Conferência Mensal'] + \
        [ws.title for ws in wb.worksheets if ws.title.startswith('Termo')] + ['Listas']
wb._sheets = [wb[t] for t in ordem]
wb.active = 0

out = 'Valvic_Controle_Patrimonio.xlsx'
wb.save(out)
print(f'gerado · {len(wb.worksheets)} abas · {out}')
