#!/usr/bin/env python3
"""Testes da planilha de cotação: calcula as fórmulas de verdade e confere
os resultados contra um modelo independente escrito em Python.

Uso:  python3 testar-cotacao.py
"""
import re
import sys
import datetime
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import formulas

def xround(x, n=2):
    """ROUND do Excel: meio para cima, não o arredondamento bancário do Python."""
    # o Excel normaliza para 15 dígitos significativos ANTES de arredondar:
    # 6*125.50*1.065 vale 801.9449999999999 em ponto flutuante, mas o Excel
    # trata como 801.945 e devolve 801.95 — o repr() do Python devolveria 801.94.
    q = Decimal(1).scaleb(-n)
    return float(Decimal(f'{float(x):.15g}').quantize(q, rounding=ROUND_HALF_UP))

ARQ = 'Valvic_Cotacao_Fornecedores.xlsx'
TRIOS = [(5, 'E', 'F', 'G'), (8, 'H', 'I', 'J'), (11, 'K', 'L', 'M'), (14, 'N', 'O', 'P')]
R_FORN, R_BAR, R_IT0, R_ITF = 10, 12, 14, 43
AP = {k: 46 + i for i, k in enumerate([
    'itens', 'situacao', 'subsem', 'imposto', 'subtotal', 'frete', 'totbruto',
    'pdesc', 'avista', 'condprazo', 'pacresc', 'aprazo', 'custoprazo',
    'entrega', 'validade', 'subcompleto', 'completo'])}
R_V0 = 65
C_MUNIT, C_MTOT, C_MFORN = 'Q', 'R', 'S'

falhas, testes = [], 0
def ck(nome, obtido, esperado, tol=0.005):
    global testes
    testes += 1
    if isinstance(esperado, (int, float)) and isinstance(obtido, (int, float)):
        ok = abs(obtido - esperado) <= tol
    else:
        ok = str(obtido).strip() == str(esperado).strip()
    if not ok:
        falhas.append(f'{nome}\n     obtido   = {obtido!r}\n     esperado = {esperado!r}')
    return ok


# ══════════════════════════════════════════ 1 · calcular a planilha
print('Calculando a planilha com o motor de fórmulas...')
xl = formulas.ExcelModel().loads(ARQ).finish()
sol = xl.calculate()
VAL = {}
pat = re.compile(r"^'\[" + re.escape(ARQ) + r"\]([^']+)'!([A-Z]+[0-9]+)$", re.I)
for k, v in sol.items():
    m = pat.match(str(k))
    if not m:
        continue
    try:
        val = v.value[0, 0]
    except Exception:
        try:
            val = v.value
        except Exception:
            val = v
    if hasattr(val, 'item'):
        try: val = val.item()
        except Exception: pass
    VAL[(m.group(1).upper(), m.group(2).upper())] = val

def C(aba, ref):
    v = VAL.get((aba.upper(), ref.upper()), '<ausente>')
    if isinstance(v, str) and v == '':
        return ''
    return v

print(f'{len(VAL)} células resolvidas em {len({k[0] for k in VAL})} abas\n')


# ══════════════════════════════════════════ 2 · modelo independente
wbv = openpyxl.load_workbook(ARQ)
ex = wbv['Exemplo']

itens, pad = [], []
for i, (c0, u, p, t) in enumerate(TRIOS):
    pad.append(ex.cell(R_BAR, c0 + 1).value or 0)
for r in range(R_IT0, R_ITF + 1):
    desc = ex.cell(r, 2).value
    if not desc:
        continue
    qtd = ex.cell(r, 4).value
    linha = []
    for i, (c0, u, p, t) in enumerate(TRIOS):
        unit = ex.cell(r, c0).value
        pct = ex.cell(r, c0 + 1).value
        linha.append((unit, pad[i] if pct is None else pct))
    itens.append((desc, qtd, linha))

NB = len(itens)
apu = []
for i, (c0, u, p, t) in enumerate(TRIOS):
    frete = ex.cell(AP['frete'], c0).value or 0
    pdesc = ex.cell(AP['pdesc'], c0).value or 0
    pacr = ex.cell(AP['pacresc'], c0).value or 0
    tot_i = [xround(q * un * (1 + pc)) for _, q, ln in itens
             for (un, pc) in [ln[i]] if un is not None]
    n = len(tot_i)
    if n == 0:
        apu.append(None); continue
    subtotal = xround(sum(tot_i))
    # SUMPRODUCT não arredonda parcela a parcela
    subsem = xround(sum(q * ln[i][0] for _, q, ln in itens if ln[i][0] is not None))
    imposto = xround(subtotal - subsem)
    totbruto = xround(subtotal + frete)
    avista = xround(totbruto * (1 - pdesc))
    aprazo = xround(totbruto * (1 + pacr))
    apu.append(dict(n=n, subsem=subsem, imposto=imposto, subtotal=subtotal, frete=frete,
                    totbruto=totbruto, avista=avista, aprazo=aprazo,
                    custoprazo=xround(aprazo - avista),
                    situacao=('COMPLETO' if n == NB else f'PARCIAL — faltam {NB-n} item(ns)'),
                    completo=(avista if n == NB else ''),
                    subcompleto=(subtotal if n == NB else '')))

nomes = [ex.cell(R_FORN, c0).value for c0, *_ in TRIOS]


# ══════════════════════════════════════════ 3 · testes — aba Exemplo
print('─' * 74)
print('TESTE 1 · totais por item (quantidade × unitário × imposto)')
for j, (desc, qtd, linha) in enumerate(itens):
    r = R_IT0 + j
    for i, (c0, u, p, t) in enumerate(TRIOS):
        un, pc = linha[i]
        esp = '' if un is None else xround(qtd * un * (1 + pc))
        ck(f'  Exemplo!{t}{r} ({nomes[i]} · {desc[:28]})', C('Exemplo', f'{t}{r}'), esp)
print(f'  {len(itens)*4} células conferidas')

print('\nTESTE 2 · menor total, menor unitário e fornecedor vencedor de cada item')
for j, (desc, qtd, linha) in enumerate(itens):
    r = R_IT0 + j
    tots = [xround(qtd * un * (1 + pc)) for (un, pc) in linha if un is not None]
    mn = min(tots)
    venc = next(nomes[i] for i, (un, pc) in enumerate(linha)
                if un is not None and xround(qtd * un * (1 + pc)) == mn)
    ck(f'  Exemplo!{C_MTOT}{r} menor total', C('Exemplo', f'{C_MTOT}{r}'), mn)
    ck(f'  Exemplo!{C_MUNIT}{r} menor unit', C('Exemplo', f'{C_MUNIT}{r}'), xround(mn / qtd))
    ck(f'  Exemplo!{C_MFORN}{r} vencedor', C('Exemplo', f'{C_MFORN}{r}'), venc)
print(f'  {len(itens)*3} células conferidas')

print('\nTESTE 3 · apuração por fornecedor')
for i, (c0, u, p, t) in enumerate(TRIOS):
    a = apu[i]
    ck(f'  {nomes[i]} · itens cotados', C('Exemplo', f'{u}{AP["itens"]}'), f'{a["n"]} de {NB}')
    ck(f'  {nomes[i]} · situação', C('Exemplo', f'{u}{AP["situacao"]}'), a['situacao'])
    for chave in ('subsem', 'imposto', 'subtotal', 'totbruto', 'avista', 'aprazo',
                  'custoprazo', 'subcompleto', 'completo'):
        ck(f'  {nomes[i]} · {chave}', C('Exemplo', f'{u}{AP[chave]}'), a[chave])
print(f'  {len(TRIOS)*11} células conferidas')

print('\nTESTE 4 · barra fixa (situação curta e total do pedido)')
for i, (c0, u, p, t) in enumerate(TRIOS):
    ck(f'  {nomes[i]} · barra situação', C('Exemplo', f'{u}{R_BAR}'), f'{apu[i]["n"]}/{NB}')
    ck(f'  {nomes[i]} · barra total', C('Exemplo', f'{t}{R_BAR}'), apu[i]['totbruto'])
print(f'  {len(TRIOS)*2} células conferidas')

print('\nTESTE 5 · veredito')
av = [a['avista'] for a in apu]
cp = [a['completo'] for a in apu if a['completo'] != '']
pz = [a['aprazo'] for a in apu]
mn_av, mn_cp, mn_pz = min(av), min(cp), min(pz)
ck('  melhor à vista · fornecedor', C('Exemplo', f'E{R_V0}'), nomes[av.index(mn_av)])
ck('  melhor à vista · valor', C('Exemplo', f'K{R_V0}'), mn_av)
ck('  melhor à vista · situação', C('Exemplo', f'Q{R_V0}'), apu[av.index(mn_av)]['situacao'])
idx_cp = next(i for i, a in enumerate(apu) if a['completo'] == mn_cp)
ck('  melhor entre completos · fornecedor', C('Exemplo', f'E{R_V0+1}'), nomes[idx_cp])
ck('  melhor entre completos · valor', C('Exemplo', f'K{R_V0+1}'), mn_cp)
ck('  melhor a prazo · fornecedor', C('Exemplo', f'E{R_V0+2}'), nomes[pz.index(mn_pz)])
ck('  melhor a prazo · valor', C('Exemplo', f'K{R_V0+2}'), mn_pz)
frac = xround(sum(min(xround(q * un * (1 + pc)) for (un, pc) in ln if un is not None)
                  for _, q, ln in itens))
ck('  compra fracionada · total', C('Exemplo', f'K{R_V0+3}'), frac)
mn_sub = min(a['subcompleto'] for a in apu if a['subcompleto'] != '')
# TEXT() usa o separador do Windows/Excel do usuário; o motor de teste é en-US.
# Aceitamos as duas formas — em pt-BR sai "103,80", em en-US sai "103.80".
nota_obt = str(C('Exemplo', f'Q{R_V0+3}'))
cauda = ' sobre o subtotal do melhor fornecedor completo — mas some frete de cada um.'
esp_br = f'Economia de R$ {mn_sub - frac:,.2f}'.replace(',', '#').replace('.', ',').replace('#', '.')
esp_us = f'Economia de R$ {mn_sub - frac:,.2f}'
ck('  compra fracionada · nota', nota_obt,
   esp_br + cauda if nota_obt.startswith(esp_br) else esp_us + cauda)
print('  9 células conferidas')

print('\nTESTE 6 · aba Cotação em branco não pode gerar lixo nem erro')
brancos = ([f'{u}{AP[k]}' for c0, u, p, t in TRIOS
            for k in ('itens', 'situacao', 'subsem', 'subtotal', 'totbruto', 'avista',
                      'aprazo', 'completo')] +
           [f'{u}{R_BAR}' for c0, u, p, t in TRIOS] +
           [f'{t}{R_BAR}' for c0, u, p, t in TRIOS] +
           [f'{C_MTOT}{R_IT0}', f'{C_MUNIT}{R_IT0}', f'{C_MFORN}{R_IT0}',
            f'{C_MTOT}{R_ITF}', f'G{R_IT0}', f'K{R_V0+3}'])
for ref in brancos:
    ck(f'  Cotação!{ref} vazia', C('Cotação', ref), '')
ck('  Cotação!E65 veredito sem dados', C('Cotação', f'E{R_V0}'), '—')
ck('  Cotação!Q66 aviso de completos', C('Cotação', f'Q{R_V0+1}'),
   'Nenhum fornecedor tem o pedido completo — você vai precisar dividir a compra')
print(f'  {len(brancos)+2} células conferidas')

print('\nTESTE 7 · Mapa de Cotações (economia e KPIs)')
mp = wbv['Mapa de Cotações']
MP0 = 16
linhas = []
for r in range(MP0, MP0 + 10):
    if mp.cell(r, 1).value is None:
        continue
    maior, fechado = mp.cell(r, 10).value, mp.cell(r, 11).value
    linhas.append((r, maior, fechado, mp.cell(r, 14).value))
for r, maior, fechado, sit in linhas:
    esp_e = '' if (maior is None or fechado is None) else xround(maior - fechado)
    ck(f'  Mapa!L{r} economia (R$)', C('Mapa de Cotações', f'L{r}'), esp_e)
    maior_p = mp.cell(r, 10).value
    esp_p = '' if (esp_e == '' or not maior_p) else esp_e / maior_p
    ck(f'  Mapa!M{r} economia (%)', C('Mapa de Cotações', f'M{r}'), esp_p, tol=1e-9)
fechados = [f for _, _, f, _ in linhas if f is not None]
econ = [xround(m - f) for _, m, f, _ in linhas if m is not None and f is not None]
pcts = [xround(m - f) / m for _, m, f, _ in linhas if m is not None and f is not None]
ck('  KPI cotações registradas', C('Mapa de Cotações', 'A13'), len(linhas))
ck('  KPI em aberto', C('Mapa de Cotações', 'D13'),
   sum(1 for *_, sit_ in linhas if sit_ in ('Em cotação', 'Aguardando fornecedor', 'Em análise')))
ck('  KPI valor fechado', C('Mapa de Cotações', 'G13'), xround(sum(fechados)))
ck('  KPI economia acumulada', C('Mapa de Cotações', 'J13'), xround(sum(econ)))
ck('  KPI economia média', C('Mapa de Cotações', 'M13'), sum(pcts)/len(pcts), tol=1e-9)
print(f'  {len(linhas)*2+5} células conferidas')

print('\nTESTE 8 · Pedido de Cotação puxa da aba escolhida')
ck('  identificação · demanda', C('Pedido de Cotação', 'D10'), '')   # aponta p/ Cotação (vazia)
ck('  item 1 · descrição', C('Pedido de Cotação', 'B25'), '')
print('  2 células conferidas (aba de origem = "Cotação", que está em branco)')


# ══════════════════════════════════════════ 4 · simulações de cenário
print('\n' + '─' * 74)
print('SIMULAÇÕES — o que a planilha responde em cada situação')
print('─' * 74)
for i, (c0, u, p, t) in enumerate(TRIOS):
    a = apu[i]
    ipi = f'{pad[i]:.1%}'.replace('.', ',') if pad[i] else '—'
    print(f'  {nomes[i]:<20} {a["situacao"]:<26} imposto padrão {ipi:>6}')
    print(f'  {"":<20} sem imposto {a["subsem"]:>10,.2f}  imposto {a["imposto"]:>8,.2f}  '
          f'frete {a["frete"]:>7,.2f}')
    print(f'  {"":<20} TOTAL DO PEDIDO {a["totbruto"]:>10,.2f}   à vista {a["avista"]:>10,.2f}'
          f'   a prazo {a["aprazo"]:>10,.2f}')
print(f'\n  Melhor à vista (qualquer)      {nomes[av.index(mn_av)]:<20} R$ {mn_av:>10,.2f}   '
      f'← {apu[av.index(mn_av)]["situacao"]}')
print(f'  Melhor à vista (só completos)  {nomes[idx_cp]:<20} R$ {mn_cp:>10,.2f}')
print(f'  Compra fracionada              {"item a item":<20} R$ {frac:>10,.2f}   '
      f'(economia de R$ {mn_sub-frac:,.2f} sobre o subtotal do melhor completo)')

print('\n  Efeito do IPI de 6,5% da MGV:')
mgv = apu[2]
print(f'    preço de tabela (sem IPI)   R$ {mgv["subsem"]:>10,.2f}  ← era por aqui que a')
print(f'    IPI 6,5% sobre os itens     R$ {mgv["imposto"]:>10,.2f}     comparação enganava')
print(f'    subtotal real               R$ {mgv["subtotal"]:>10,.2f}')

print('\n' + '═' * 74)
print(f'{testes} verificações · {len(falhas)} falha(s)')
if falhas:
    print('═' * 74)
    for f in falhas:
        print('  ✗ ' + f)
    sys.exit(1)
print('TODOS OS TESTES PASSARAM')
