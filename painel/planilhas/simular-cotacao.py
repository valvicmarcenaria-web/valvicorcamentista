#!/usr/bin/env python3
"""Simulações de borda da planilha de cotação.

Monta variantes da aba Cotação, calcula as fórmulas de verdade e mostra o que
a planilha responde em cada situação difícil.

Uso:  python3 simular-cotacao.py
"""
import os
import re
import shutil
import openpyxl
import formulas

BASE = 'Valvic_Cotacao_Fornecedores.xlsx'
TMP = '/tmp/claude-0/-home-user-valvicorcamentista/2544489f-df71-5f40-87c6-89025901a0cf/scratchpad/sim'
os.makedirs(TMP, exist_ok=True)

TRIOS = [(5, 'E', 'F', 'G'), (8, 'H', 'I', 'J'), (11, 'K', 'L', 'M'), (14, 'N', 'O', 'P')]
R_FORN, R_BAR, R_IT0 = 10, 12, 14
AP = {k: 46 + i for i, k in enumerate([
    'itens', 'situacao', 'subsem', 'imposto', 'subtotal', 'frete', 'totbruto',
    'pdesc', 'avista', 'condprazo', 'pacresc', 'aprazo', 'custoprazo',
    'entrega', 'validade', 'subcompleto', 'completo'])}
R_V0 = 65


def montar(nome, itens, padrao=(None,)*4, frete=(0,)*4, desc=(0,)*4, acresc=(0,)*4):
    """Grava uma variante da aba Cotação e devolve o caminho do arquivo."""
    caminho = os.path.join(TMP, f'{nome}.xlsx')
    shutil.copy(BASE, caminho)
    wb = openpyxl.load_workbook(caminho)
    ws = wb['Cotação']
    for i, (c0, u, p, t) in enumerate(TRIOS):
        ws.cell(R_FORN, c0).value = f'Forn {i+1}'
        if padrao[i] is not None:
            ws.cell(R_BAR, c0 + 1).value = padrao[i]
        ws.cell(AP['frete'], c0).value = frete[i]
        ws.cell(AP['pdesc'], c0).value = desc[i]
        ws.cell(AP['pacresc'], c0).value = acresc[i]
    for j, (d, q, precos, pcts) in enumerate(itens):
        r = R_IT0 + j
        ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value = d, 'un', q
        for i, v in enumerate(precos):
            if v is not None:
                ws.cell(r, TRIOS[i][0]).value = v
        for i, v in enumerate(pcts or [None]*4):
            if v is not None:
                ws.cell(r, TRIOS[i][0] + 1).value = v
    wb.save(caminho)
    return caminho


def calcular(caminho):
    xl = formulas.ExcelModel().loads(caminho).finish()
    sol = xl.calculate()
    arq = os.path.basename(caminho)
    pat = re.compile(r"^'\[" + re.escape(arq) + r"\]([^']+)'!([A-Z]+[0-9]+)$", re.I)
    V = {}
    for k, v in sol.items():
        m = pat.match(str(k))
        if not m:
            continue
        try:
            val = v.value[0, 0]
        except Exception:
            val = getattr(v, 'value', v)
        if hasattr(val, 'item'):
            try: val = val.item()
            except Exception: pass
        V[(m.group(1).upper(), m.group(2).upper())] = val
    return lambda ref, aba='COTAÇÃO': V.get((aba, ref), '<ausente>')


def moeda(v):
    if not isinstance(v, (int, float)):
        return '—'
    return f'R$ {v:,.2f}'.replace(',', '#').replace('.', ',').replace('#', '.')


def relatar(titulo, pergunta, caminho):
    C = calcular(caminho)
    print('\n' + '═' * 78)
    print(f'  {titulo}')
    print(f'  Pergunta: {pergunta}')
    print('═' * 78)
    print(f'  {"":16} {"barra":>9} {"total pedido":>15} {"à vista":>13} {"situação"}')
    for i, (c0, u, p, t) in enumerate(TRIOS):
        barra = str(C(u + str(R_BAR)))
        tot = moeda(C(t + str(R_BAR)))
        av = moeda(C(u + str(AP['avista'])))
        sit = C(u + str(AP['situacao']))
        print(f'  Forn {i+1:<11} {barra:>9} {tot:>15} {av:>13} {sit}')
    print('  ' + '─' * 74)
    for rot, lin in (('Melhor à vista', R_V0), ('Melhor entre completos', R_V0 + 1),
                     ('Melhor a prazo', R_V0 + 2), ('Compra fracionada', R_V0 + 3)):
        print(f'  {rot:<24} {str(C(f"E{lin}")):<14} {moeda(C(f"K{lin}")):>14}   '
              f'{str(C(f"Q{lin}"))[:60]}')


# ══════════════════════════════════════════════════════════════════════
# A · fornecedor que não respondeu nada
relatar('A · UM FORNECEDOR NÃO RESPONDEU',
        'a planilha marca "NÃO COTOU" e o exclui do veredito?',
        montar('a_mudo', [
            ('Chapa MDF 18 mm branco TX', 20, [180.00, 175.00, 189.00, None], None),
            ('Fita de borda 22 mm branca', 5, [42.00, 45.00, 39.90, None], None),
            ('Cola PVA 5 kg', 4, [68.00, 71.00, 64.50, None], None),
        ], frete=(150, 0, 200, 0)))

# B · empate exato entre dois fornecedores
relatar('B · EMPATE EXATO',
        'com dois fornecedores no mesmo preço, quem a planilha aponta?',
        montar('b_empate', [
            ('Dobradiça caneco 35 mm', 100, [8.00, 8.00, 9.50, 10.00], None),
            ('Corrediça 45 cm', 20, [40.00, 40.00, 44.00, 45.00], None),
        ]))

# C · item que ninguém cotou
relatar('C · ITEM QUE NINGUÉM COTOU',
        'a compra fracionada avisa que a lista está furada?',
        montar('c_furo', [
            ('Puxador perfil preto 3 m', 10, [60.00, 58.00, 62.00, 59.00], None),
            ('Perfil LED embutir 2 m — ninguém tem', 8, [None]*4, None),
            ('Fecho magnético', 40, [3.00, 3.20, 2.90, 3.10], None),
        ]))

# D · todos completos
relatar('D · OS QUATRO ATENDEM 100%',
        '"melhor à vista" e "melhor entre completos" apontam o mesmo?',
        montar('d_todos', [
            ('Chapa MDF 15 mm', 30, [160.00, 158.00, 165.00, 162.00], None),
            ('Cola de contato 3 L', 6, [95.00, 99.00, 92.00, 97.00], None),
        ], frete=(120, 180, 90, 150), desc=(0.05, 0.03, 0.06, 0.04)))

# E · fracionar não compensa
relatar('E · FRACIONAR NÃO COMPENSA',
        'quando um fornecedor é o mais barato em tudo, a planilha desaconselha dividir?',
        montar('e_semganho', [
            ('Chapa MDF 18 mm', 40, [200.00, 185.00, 210.00, 205.00], None),
            ('Fita de borda 22 mm', 10, [50.00, 44.00, 52.00, 51.00], None),
            ('Parafuso 4x40 — cx 500', 12, [38.00, 33.00, 40.00, 39.00], None),
        ]))

# F · imposto só na linha, sem padrão do fornecedor
relatar('F · IPI DIGITADO SÓ NA LINHA',
        'sem % padrão na barra, o % da linha sozinho já corrige a comparação?',
        montar('f_linha', [
            ('Ferragem importada — ST 12%', 10, [100.00, 100.00, 100.00, 100.00],
             [0.12, None, None, None]),
            ('Item nacional sem imposto', 10, [50.00, 50.00, 50.00, 50.00], None),
        ]))

# G · linha do item sobrepõe o padrão do fornecedor
relatar('G · LINHA SOBREPÕE O PADRÃO DO FORNECEDOR',
        'Forn 1 tem 6,5% padrão; um item dele é isento (0%). O 0% prevalece?',
        montar('g_override', [
            ('Item tributado — usa o padrão 6,5%', 10, [100.00]*4, None),
            ('Item isento — 0% na linha', 10, [100.00]*4, [0.0, None, None, None]),
        ], padrao=(0.065, None, None, None)))

# H · item com descrição e sem quantidade
relatar('H · ITEM SEM QUANTIDADE',
        'linha digitada pela metade quebra a apuração?',
        montar('h_semqtd', [
            ('Chapa MDF 18 mm', 20, [180.00, 175.00, 189.00, 182.00], None),
            ('Item lançado sem quantidade', None, [90.00, 88.00, 92.00, 91.00], None),
        ]))

print('\n' + '═' * 78)
print('  Fim das simulações.')
