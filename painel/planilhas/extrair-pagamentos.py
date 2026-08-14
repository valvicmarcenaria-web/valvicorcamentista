#!/usr/bin/env python3
"""Lê a planilha original de Controle de Pagamentos e devolve os dados
normalizados, com as datas corrigidas.

Diagnóstico que justifica a correção de datas (comprovado na auditoria):
o arquivo foi editado por uma ferramenta em locale en-US, que leu "DD/MM"
como "MM/DD". Sempre que os dois números eram <= 12 ela converteu para data
com dia e mês TROCADOS; quando o primeiro número era > 12 não conseguiu
converter e deixou como texto. A partição é perfeita nas 4 abas:
  · 67 células viraram data — TODAS com dia <= 12 e mês <= 12
  · 113 células ficaram texto — TODAS com o primeiro número > 12
  · nenhuma data gravada com dia > 12, nenhum texto com primeiro número <= 12
Logo: toda data gravada precisa ter dia e mês trocados, e todo texto "DD/MM"
precisa virar data de verdade.

Uso:  python3 extrair-pagamentos.py         (imprime o relatório da extração)
"""
import datetime
import re
import unicodedata
import openpyxl

ORIGEM = ('/root/.claude/uploads/2544489f-df71-5f40-87c6-89025901a0cf/'
          '1e190286-Controle_pagamentos_de_projetos.xlsx')

MESES = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
         'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
NUM_MES = {}
for i, m in enumerate(MESES, start=1):
    NUM_MES[m.lower()] = i
    NUM_MES[unicodedata.normalize('NFKD', m.lower()).encode('ascii', 'ignore').decode()] = i

# (aba, linha do cabeçalho, colunas dos blocos de pagamento)
CFG = {'2026': (8, [8, 12, 16, 20, 24]),
       '2025': (8, [9, 13, 17, 21, 25]),
       '2024': (7, [8, 12, 16, 20, 24]),
       '2023': (7, [8, 12, 16, 20, 24])}

FORMAS_CONHECIDAS = {'pix', 'cartao', 'cartão', 'dinheiro', 'especie', 'espécie',
                     'ted', 'transferencia', 'transferência', 'boleto', 'permuta',
                     'cheque', 'material', 'deposito', 'depósito'}


def norm(s):
    s = str(s or '').strip().lower()
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()


class Ocorrencia:
    """Registro do que foi mexido, para a aba de notas de migração."""
    def __init__(self):
        self.itens = []
    def add(self, aba, linha, cliente, tipo, antes, depois):
        self.itens.append((aba, linha, cliente, tipo, antes, depois))
    def por_tipo(self):
        d = {}
        for it in self.itens:
            d.setdefault(it[3], []).append(it)
        return d


OC = Ocorrencia()


def resolver_data(valor, ano, aba, linha, cliente):
    """Devolve (data, texto_residual). Corrige a inversão dia/mês."""
    if valor is None:
        return None, None
    if isinstance(valor, (datetime.datetime, datetime.date)):
        d = valor.date() if isinstance(valor, datetime.datetime) else valor
        if d.day <= 12:
            nova = datetime.date(d.year, d.day, d.month)
            OC.add(aba, linha, cliente, 'data invertida',
                   d.strftime('%d/%m/%Y'), nova.strftime('%d/%m/%Y'))
            return nova, None
        # dia > 12: a ferramenta não teria como ter invertido — mantém
        return d, None

    s = str(valor).strip()
    if not s:
        return None, None
    m = re.fullmatch(r'(\d{1,2})[/\-.](\d{1,2})(?:[/\-.](\d{2,4}))?', s)
    if m:
        dia, mes, an = int(m.group(1)), int(m.group(2)), m.group(3)
        if an:
            an = int(an)
            an += 2000 if an < 100 else 0
        else:
            an = ano
        try:
            nova = datetime.date(an, mes, dia)
            OC.add(aba, linha, cliente, 'texto virou data', s, nova.strftime('%d/%m/%Y'))
            return nova, None
        except ValueError:
            OC.add(aba, linha, cliente, 'data impossível', s, '(mantida como texto)')
            return None, s
    # não é data: costuma ser a forma de pagamento digitada na coluna errada
    OC.add(aba, linha, cliente, 'texto fora de lugar', s, 'movido para Descrição')
    return None, s


def ler():
    wb = openpyxl.load_workbook(ORIGEM)
    anos = {}
    for aba, (hr, blocos) in CFG.items():
        ws = wb[aba]
        ano = int(aba)
        atual = {'mes': 'Anterior', 'projetos': []}
        blocos_ano = [atual]
        ultimo = None                      # último projeto lido, p/ sub-linhas
        for r in range(hr + 1, ws.max_row + 1):
            cli = ws.cell(r, 1).value
            proj = ws.cell(r, 2).value
            val = ws.cell(r, 3).value
            forma = ws.cell(r, 4).value

            # divisor de mês: coluna A é nome de mês e não há projeto na linha
            eh_divisor = (norm(cli) in NUM_MES and
                          (proj is None or (isinstance(proj, str) and proj.startswith('='))))
            if eh_divisor:
                atual = {'mes': MESES[NUM_MES[norm(cli)] - 1], 'projetos': []}
                blocos_ano.append(atual)
                ultimo = None
                continue

            # linhas de resumo repetidas no meio da planilha (fórmulas em A/B)
            def eh_formula(x):
                return isinstance(x, str) and x.startswith('=')
            if eh_formula(cli) or eh_formula(proj) or norm(cli).startswith('total'):
                continue

            pags = []
            for c in blocos:
                pv = ws.cell(r, c).value
                pd = ws.cell(r, c + 1).value
                pdsc = ws.cell(r, c + 2).value
                if isinstance(pv, str) and pv.startswith('='):
                    pv = None              # fórmula-resumo de sub-linhas
                if pv is None and pd is None and pdsc is None:
                    continue
                data, residual = resolver_data(pd, ano, aba, r, cli or proj)
                desc = str(pdsc).strip() if pdsc else ''
                if desc.startswith('='):
                    desc = ''              # sobra de fórmula, não é descrição
                if residual and not residual.startswith('='):
                    desc = (residual + (' · ' + desc if desc else '')).strip()
                pags.append({'valor': pv if isinstance(pv, (int, float)) else None,
                             'data': data, 'desc': desc})

            tem_id = (cli is not None) or (proj is not None)
            tem_valor = isinstance(val, (int, float))

            # sub-linha: sem identificação e sem valor, mas com pagamentos
            if not tem_id and not tem_valor:
                if pags and ultimo is not None:
                    ultimo['pagamentos'].extend(pags)
                continue
            # projeto de verdade precisa ter valor ou ao menos um pagamento;
            # isso descarta o bloco de anotações no rodapé da aba
            if not tem_valor and not pags:
                continue

            p = {'cliente': str(cli).strip() if cli else '',
                 'projeto': str(proj).strip() if proj else '',
                 'valor': val if tem_valor else None,
                 'forma': str(forma).strip() if forma else '',
                 'pagamentos': pags, 'linha': r}
            atual['projetos'].append(p)
            ultimo = p
        blocos_ano = [b for b in blocos_ano if b['projetos']]
        anos[aba] = blocos_ano
    return anos


def consolidar(p, limite=5):
    """Garante no máximo `limite` pagamentos, juntando o excedente no último."""
    pags = [x for x in p['pagamentos'] if x['valor'] is not None or x['data'] or x['desc']]
    if len(pags) <= limite:
        p['pagamentos'] = pags
        return False
    manter, sobra = pags[:limite - 1], pags[limite - 1:]
    soma = sum(x['valor'] or 0 for x in sobra)
    detalhe = ' + '.join(
        f"{(x['valor'] or 0):,.0f}".replace(',', '.') + (f" ({x['data']:%d/%m})" if x['data'] else '')
        for x in sobra)
    manter.append({'valor': soma, 'data': sobra[0]['data'],
                   'desc': f'{len(sobra)} parcelas: {detalhe}'})
    p['pagamentos'] = manter
    return True


if __name__ == '__main__':
    anos = ler()
    print('EXTRAÇÃO DA PLANILHA ORIGINAL')
    print('=' * 78)
    tot_proj = tot_val = tot_rec = 0
    excedentes = []
    for aba in ('2026', '2025', '2024', '2023'):
        blocos = anos[aba]
        np_ = sum(len(b['projetos']) for b in blocos)
        v = sum(p['valor'] or 0 for b in blocos for p in b['projetos'])
        rec = sum(x['valor'] or 0 for b in blocos for p in b['projetos']
                  for x in p['pagamentos'])
        maxp = max(len(p['pagamentos']) for b in blocos for p in b['projetos'])
        print(f'  {aba}  {len(blocos):2} blocos · {np_:3} projetos · '
              f'vendido {v:>12,.0f} · recebido {rec:>12,.0f} · '
              f'saldo {v-rec:>11,.0f} · máx. pagamentos {maxp}')
        for b in blocos:
            for p in b['projetos']:
                if len(p['pagamentos']) > 5:
                    excedentes.append((aba, p['linha'], p['cliente'], len(p['pagamentos'])))
        tot_proj += np_; tot_val += v; tot_rec += rec
    print('  ' + '-' * 74)
    print(f'  TOTAL {tot_proj:3} projetos · vendido {tot_val:>12,.0f} · '
          f'recebido {tot_rec:>12,.0f} · saldo {tot_val-tot_rec:>11,.0f}')

    print('\n  Blocos de mês encontrados por ano:')
    for aba in ('2026', '2025', '2024', '2023'):
        print(f'    {aba}: ' + ', '.join(f"{b['mes']}({len(b['projetos'])})"
                                         for b in anos[aba]))

    print('\n  Projetos com mais de 5 pagamentos (serão consolidados):')
    print('   ', excedentes or 'nenhum')

    print('\nCORREÇÕES APLICADAS')
    print('=' * 78)
    for tipo, itens in sorted(OC.por_tipo().items()):
        print(f'  {tipo}: {len(itens)}')
        for it in itens[:4]:
            print(f'      {it[0]}!{it[1]:<4} {str(it[2])[:22]:24} {it[4]!r} → {it[5]!r}')
        if len(itens) > 4:
            print(f'      ... e mais {len(itens)-4}')
