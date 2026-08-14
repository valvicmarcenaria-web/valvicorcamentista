#!/usr/bin/env python3
"""Reconcilia o Controle de Pagamentos novo contra a planilha original.

Calcula as fórmulas de verdade (biblioteca `formulas`) e confere, projeto a
projeto e mês a mês, se nenhum contrato e nenhum pagamento se perdeu na
migração — e se as colunas calculadas (Recebido, Saldo, % e Situação) batem
com o modelo independente escrito em Python.

Uso:  python3 testar-pagamentos.py
"""
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import formulas

import importlib.util as _il, pathlib as _pl
_spec = _il.spec_from_file_location(
    'extrair_pagamentos', _pl.Path(__file__).with_name('extrair-pagamentos.py'))
_mod = _il.module_from_spec(_spec); _spec.loader.exec_module(_mod)
ler = _mod.ler
consolidar = _mod.consolidar
MESES = _mod.MESES

ARQ = 'Valvic_Controle_Pagamentos.xlsx'
ANOS = ['2026', '2025', '2024', '2023']
ANTERIOR = 'Anterior (2022)'
BLOCOS = ['I', 'L', 'O', 'R', 'U']
R_IT0 = 10


def xround(x, n=2):
    q = Decimal(1).scaleb(-n)
    return float(Decimal(f'{float(x):.15g}').quantize(q, rounding=ROUND_HALF_UP))


falhas, testes = [], 0
def ck(nome, obtido, esperado, tol=0.005):
    global testes
    testes += 1
    if isinstance(esperado, (int, float)) and isinstance(obtido, (int, float)):
        ok = abs(obtido - esperado) <= tol
    else:
        ok = str(obtido).strip() == str(esperado).strip()
    if not ok:
        falhas.append(f'{nome}\n     obtido   = {obtido!r}\n     esperado = {esperado!r}')
    return ok


# ═════════════════════════════════ 1 · calcular a planilha nova
print('Calculando a planilha com o motor de fórmulas...')
xl = formulas.ExcelModel().loads(ARQ).finish()
sol = xl.calculate()
pat = re.compile(r"^'\[" + re.escape(ARQ) + r"\]([^']+)'!([A-Z]+[0-9]+)$", re.I)
VAL = {}
for k, v in sol.items():
    m = pat.match(str(k))
    if not m:
        continue
    try:
        val = v.value[0, 0]
    except Exception:
        val = getattr(v, 'value', v)
    if hasattr(val, 'item'):
        try: val = val.item()
        except Exception: pass
    VAL[(m.group(1).upper(), m.group(2).upper())] = val

def C(aba, ref):
    return VAL.get((aba.upper(), ref.upper()), '<ausente>')

print(f'{len(VAL)} células resolvidas em {len({k[0] for k in VAL})} abas\n')


# ═════════════════════════════════ 2 · verdade vinda do original
dados = ler()
verdade = {}
for aba in ANOS:
    por_mes = {b['mes']: b for b in dados[aba]}
    ordenados = []
    if 'Anterior' in por_mes:
        b = por_mes.pop('Anterior'); b['mes'] = ANTERIOR
        ordenados.append(b)
    for m in MESES:
        ordenados.append(por_mes.get(m, {'mes': m, 'projetos': []}))
    for b in ordenados:
        for p in b['projetos']:
            consolidar(p, 5)
    verdade[aba] = ordenados

wbv = openpyxl.load_workbook(ARQ)


# ═════════════════════════════════ 3 · testes
print('─' * 76)
print('TESTE 1 · nenhum projeto e nenhum pagamento se perdeu na migração')
for aba in ANOS:
    ws = wbv[aba]
    # coleta o que está na planilha nova
    na_planilha = []
    for r in range(R_IT0, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if not isinstance(v, (int, float)):
            continue
        if isinstance(ws.cell(r, 3).value, str):
            continue
        pags = [ws.cell(r, openpyxl.utils.column_index_from_string(b)).value
                for b in BLOCOS]
        pags = [p for p in pags if isinstance(p, (int, float))]
        na_planilha.append((ws.cell(r, 1).value or '', ws.cell(r, 2).value or '',
                            v, sorted(pags)))
    # o que deveria estar
    esperado = []
    for b in verdade[aba]:
        for p in b['projetos']:
            if p['valor'] is None:
                continue
            pg = sorted(x['valor'] for x in p['pagamentos'] if x['valor'] is not None)
            esperado.append((p['cliente'], p['projeto'], p['valor'], pg))
    # divisores entram como linha com valor -> remover os que são SUM (fórmula)
    ck(f'  {aba} · nº de projetos', len(na_planilha), len(esperado))
    ck(f'  {aba} · soma dos contratos', xround(sum(x[2] for x in na_planilha)),
       xround(sum(x[2] for x in esperado)))
    ck(f'  {aba} · soma dos pagamentos', xround(sum(sum(x[3]) for x in na_planilha)),
       xround(sum(sum(x[3]) for x in esperado)))
    def assina(x):
        return (str(x[0]).strip(), str(x[1]).strip(), xround(x[2]),
                tuple(xround(v) for v in x[3]))
    for a, b_ in zip(na_planilha, esperado):
        ck(f'  {aba} · {str(b_[0])[:22]}', assina(a), assina(b_))
print(f'  {testes} verificações até aqui')

n0 = testes
print('\nTESTE 2 · colunas calculadas linha a linha (Recebido, Saldo, %, Situação)')
for aba in ANOS:
    ws = wbv[aba]
    for r in range(R_IT0, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if not isinstance(v, (int, float)):
            continue
        pags = [ws.cell(r, openpyxl.utils.column_index_from_string(b)).value
                for b in BLOCOS]
        rec = xround(sum(p for p in pags if isinstance(p, (int, float))))
        saldo = xround(v - rec)
        sit = ('A receber' if rec == 0 else
               'Quitado' if saldo == 0 else
               'Recebido a mais' if saldo < 0 else 'Parcial')
        ck(f'  {aba}!E{r} recebido', C(aba, f'E{r}'), rec)
        ck(f'  {aba}!F{r} saldo', C(aba, f'F{r}'), saldo)
        ck(f'  {aba}!G{r} %', C(aba, f'G{r}'), rec / v if v else '')
        ck(f'  {aba}!H{r} situação', C(aba, f'H{r}'), sit)
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 3 · totais de mês e do ano')
for aba in ANOS:
    ws = wbv[aba]
    divisores = [(ws.cell(r, 1).value, r) for r in range(R_IT0, ws.max_row + 1)
                 if isinstance(ws.cell(r, 3).value, str)
                 and str(ws.cell(r, 3).value).startswith('=SUM(')]
    esperado = {b['mes']: b for b in verdade[aba]}
    tot_v = tot_r = 0
    for nome, d in divisores:
        b = esperado[nome]
        v = xround(sum(p['valor'] or 0 for p in b['projetos']))
        rec = xround(sum(x['valor'] or 0 for p in b['projetos']
                         for x in p['pagamentos'] if x['valor'] is not None))
        ck(f'  {aba} · {nome} vendido', C(aba, f'C{d}'), v)
        ck(f'  {aba} · {nome} recebido', C(aba, f'E{d}'), rec)
        ck(f'  {aba} · {nome} saldo', C(aba, f'F{d}'), xround(v - rec))
        if nome != ANTERIOR:
            tot_v += v; tot_r += rec
    ck(f'  {aba} · VENDIDO no ano', C(aba, 'A7'), xround(tot_v))
    ck(f'  {aba} · RECEBIDO no ano', C(aba, 'D7'), xround(tot_r))
    ck(f'  {aba} · A RECEBER no ano', C(aba, 'F7'), xround(tot_v - tot_r))
    ck(f'  {aba} · conferência', str(C(aba, 'R7'))[:2], 'OK')
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 4 · o Painel espelha as abas de ano')
for i, aba in enumerate(ANOS):
    r = 11 + i
    for col, orig in (('C', 'A7'), ('D', 'D7'), ('E', 'F7')):
        ck(f'  Painel!{col}{r} ({aba})', C('Painel', f'{col}{r}'), C(aba, orig))
for col, orig in (('A7', 'A7'), ('D7', 'D7'), ('G7', 'F7')):
    ck(f'  Painel!{col} (KPI)', C('Painel', col), C('2026', orig))
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 5 · ranking dos maiores saldos de 2026')
ws = wbv['2026']
abertos = []
for r in range(R_IT0, ws.max_row + 1):
    v = ws.cell(r, 3).value
    if not isinstance(v, (int, float)):
        continue
    pags = [ws.cell(r, openpyxl.utils.column_index_from_string(b)).value for b in BLOCOS]
    saldo = xround(v - sum(p for p in pags if isinstance(p, (int, float))))
    if saldo > 0:
        abertos.append((saldo, ws.cell(r, 1).value or ''))
abertos.sort(reverse=True)
for k in range(1, 13):
    ck(f'  Painel · {k}º saldo', C('Painel', f'F{32+k}'), abertos[k-1][0])
    ck(f'  Painel · {k}º cliente', C('Painel', f'B{32+k}'), abertos[k-1][1])
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 6 · datas — todas viraram data de verdade')
sem_data = texto = 0
for aba in ANOS:
    ws = wbv[aba]
    import datetime
    for r in range(R_IT0, ws.max_row + 1):
        for b in BLOCOS:
            c = ws.cell(r, openpyxl.utils.column_index_from_string(b) + 1)
            if c.value is None:
                continue
            if isinstance(c.value, (datetime.datetime, datetime.date)):
                sem_data += 1
            else:
                texto += 1
ck('  células de data que sobraram como texto', texto, 0)
print(f'  {sem_data} datas gravadas como data · {texto} ainda como texto')

n0 = testes
print('\nTESTE 7 · abas Aporte Walton e Crédito Samuel')
wa = wbv['Aporte Walton']
tot = {}
for r in range(1, wa.max_row + 1):
    d = wa.cell(r, 2).value
    if isinstance(d, str) and d.startswith('Total · '):
        tot[d.replace('Total · ', '')] = r
    if isinstance(d, str) and d.startswith('SALDO DO APORTE'):
        tot['saldo'] = r
somas = {}
for nome, r in tot.items():
    somas[nome] = C('Aporte Walton', f'A{r}')
ck('  Walton · entradas', somas['entradas'], 239860)
ck('  Walton · saídas', somas['saídas'], 106678)
ck('  Walton · investimentos', somas['investimentos'], 93182)
ck('  Walton · saldo', somas['saldo'], 40000)
cs = wbv['Crédito Samuel']
lin = [r for r in range(1, cs.max_row + 1) if cs.cell(r, 1).value == 'TOTAL ACUMULADO']
ck('  Samuel · total acumulado', C('Crédito Samuel', f'B{lin[0]}'), 300)
print(f'  {testes - n0} verificações')

print('\n' + '═' * 76)
print(f'{testes} verificações · {len(falhas)} falha(s)')
if falhas:
    print('═' * 76)
    for f in falhas[:30]:
        print('  ✗ ' + f)
    if len(falhas) > 30:
        print(f'  ... e mais {len(falhas)-30}')
    sys.exit(1)
print('TODOS OS TESTES PASSARAM — nada se perdeu na migração')
