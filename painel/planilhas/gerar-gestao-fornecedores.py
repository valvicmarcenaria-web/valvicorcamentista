#!/usr/bin/env python3
"""Gera a planilha de Gestão de Fornecedores da Valvic.

Abas: Dashboard · Fornecedores · Compras · Ocorrências · Ficha do Fornecedor (A4) · Listas

Uso:  python3 gerar-gestao-fornecedores.py
Saída: Valvic_Gestao_Fornecedores.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

# ─────────────────────────────────────────────── paleta Valvic
NAVY, NAVY2 = '0E2038', '16314F'
GOLD, GOLDS, GOLDBG = 'C2A05A', 'D8BD80', 'F6EDD6'
INK, MUTED = '1B2733', '6C7785'
LINE, LINE2 = 'E8E3D8', 'DFDACD'
OK, BLUE, RED, AMBER = '2F7D4F', '2F5D8C', 'B0413F', 'B57A16'
INPUT, WHITE, CALC = 'FFF9E3', 'FFFFFF', 'F4F6F8'

F = 'Arial'
def font(sz=10, b=False, c=INK, i=False):
    return Font(name=F, size=sz, bold=b, color=c, italic=i)
def fill(c): return PatternFill('solid', fgColor=c)
def side(c=LINE, st='thin'): return Side(style=st, color=c)
GRID    = Border(bottom=side(LINE), left=side(LINE), right=side(LINE))
BOTTOM2 = Border(bottom=side(LINE2))
F9      = Font(name=F, size=9, color=INK)
F9CALC  = Font(name=F, size=9, color=NAVY2)
F9COD   = Font(name=F, size=9, bold=True, color=NAVY)
FILL_IN, FILL_CALC = PatternFill('solid', fgColor=INPUT), PatternFill('solid', fgColor=CALC)
CTR   = Alignment(horizontal='center', vertical='center')
CTRW  = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT  = Alignment(horizontal='left', vertical='center')
LEFTW = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
MOEDA, MOEDA0, PCT, PCT1, DATA, N2 = 'R$ #,##0.00', 'R$ #,##0', '0%', '0.0%', 'DD/MM/YYYY', '0.00'

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════ dados-base
SEGMENTOS = ['Chapas e madeiras', 'Ferragens e acessórios', 'Fitas de borda', 'Colas e adesivos',
             'Iluminação', 'Vidros e espelhos', 'Pedras e superfícies', 'Tintas e acabamentos',
             'Ferramentas e consumíveis', 'Serviços terceirizados', 'Máquinas e equipamentos',
             'Logística e transporte', 'Outros']
METODOS   = ['PIX', 'Boleto', 'Cartão de crédito', 'Transferência', 'Dinheiro', 'PIX ou boleto',
             'Boleto ou cartão', 'Diversos']
CONDICOES = ['À vista', '7 dias', '14 dias', '28 dias', '30 dias', '30/60', '30/60/90',
             'Parcelado no cartão', 'A combinar']
SITUACOES = ['Ativo', 'Em avaliação', 'Preferencial', 'Restrito', 'Bloqueado', 'Inativo']
TIPO_PIX  = ['CNPJ', 'CPF', 'Telefone', 'E-mail', 'Chave aleatória']
OCORRENCIAS = ['Atraso na entrega', 'Item errado', 'Quantidade divergente', 'Avaria no transporte',
               'Qualidade abaixo do padrão', 'Cobrança indevida', 'Nota fiscal incorreta',
               'Falta de retorno do vendedor', 'Outro']
GRAVIDADE = ['Baixa', 'Média', 'Alta']
SIMNAO    = ['Sim', 'Não']

# fornecedores já identificados no financeiro
SEED = [
    ('MADEGEM — Comércio de Madeiras Ltda', 'Chapas e madeiras',
     'MDF, MDP, compensado, chapas cruas e revestidas'),
    ('Bigfer', 'Ferragens e acessórios',
     'Ferragens, corrediças, dobradiças, puxadores'),
    ('JR Ferragens', 'Ferragens e acessórios',
     'Ferragens em geral e acessórios de montagem'),
    ('MGV Distribuidora de Ferragens e Parafusos', 'Ferragens e acessórios',
     'Parafusos, buchas, fixadores e ferragens'),
    ('Ferragens Ipê', 'Ferragens e acessórios',
     'Ferragens e acessórios diversos'),
]

# ══════════════════════════════════════════════ helpers
def faixa_marca(ws, ncols, titulo, sub, linha=1):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
    c = ws.cell(linha, 1, 'VALVIC MARCENARIA')
    c.font = Font(name=F, size=13, bold=True, color=WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha].height = 26
    ws.merge_cells(start_row=linha+1, start_column=1, end_row=linha+1, end_column=ncols)
    c = ws.cell(linha+1, 1, 'Vargas Decor Ltda   ·   CNPJ 17.269.304/0001-51   ·   Belo Horizonte / MG')
    c.font = Font(name=F, size=8, color='9FB0C4'); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha+1].height = 15
    ws.merge_cells(start_row=linha+2, start_column=1, end_row=linha+2, end_column=ncols)
    c = ws.cell(linha+2, 1, titulo)
    c.font = Font(name=F, size=15, bold=True, color=WHITE); c.fill = fill(NAVY2)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha+2].height = 28
    ws.merge_cells(start_row=linha+3, start_column=1, end_row=linha+3, end_column=ncols)
    c = ws.cell(linha+3, 1, sub)
    c.font = Font(name=F, size=8.5, color='7A5B17', italic=True); c.fill = fill(GOLDBG)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha+3].height = 17
    return linha + 4

def titulo_secao(ws, row, ncols, texto, nota=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, ('  ' + texto.upper()) + (f'          {nota}' if nota else ''))
    c.font = Font(name=F, size=9, bold=True, color=NAVY); c.fill = fill(GOLDBG)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 20
    return row + 1

def cab_tabela(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(name=F, size=8.5, bold=True, color=WHITE); c.fill = fill(NAVY2)
        c.alignment = CTRW
        c.border = Border(left=side(NAVY2), right=side(NAVY2), bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 32
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1

def print_cfg(ws, area, retrato=True, margens=(0.45, 0.35, 0.4, 0.35)):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait' if retrato else 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.print_area = area
    l, r, t, b = margens
    ws.page_margins.left, ws.page_margins.right = l, r
    ws.page_margins.top, ws.page_margins.bottom = t, b
    ws.sheet_view.showGridLines = False

def dv(ws, formula, rng):
    d = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(d); d.add(rng)

def col_lista(ws, col, titulo, valores, larg=26):
    ws.column_dimensions[col].width = larg
    c = ws[f'{col}1']; c.value = titulo
    c.font = font(9, True, WHITE); c.fill = fill(NAVY2); c.alignment = CTR
    for i, v in enumerate(valores, start=2):
        cc = ws[f'{col}{i}']; cc.value = v; cc.font = font(10); cc.border = BOTTOM2

# ══════════════════════════════════════════════ ABA · Listas
ls = wb.active; ls.title = 'Listas'
ls.sheet_view.showGridLines = False
col_lista(ls, 'A', 'Segmento', SEGMENTOS, 28)
col_lista(ls, 'B', 'Método de pagamento', METODOS, 22)
col_lista(ls, 'C', 'Condição de pagamento', CONDICOES, 22)
col_lista(ls, 'D', 'Situação', SITUACOES, 16)
col_lista(ls, 'E', 'Tipo de chave PIX', TIPO_PIX, 20)
col_lista(ls, 'F', 'Tipo de ocorrência', OCORRENCIAS, 30)
col_lista(ls, 'G', 'Gravidade', GRAVIDADE, 14)
col_lista(ls, 'H', 'Sim / Não', SIMNAO, 12)
col_lista(ls, 'I', 'Nota (1 a 5)', [1, 2, 3, 4, 5], 14)
ls['K1'] = 'Período de análise'
ls['K1'].font = font(9, True, WHITE); ls['K1'].fill = fill(NAVY2); ls['K1'].alignment = CTR
ls.column_dimensions['K'].width = 30; ls.column_dimensions['L'].width = 16
for i, (rot, val) in enumerate([('Data inicial', '=DATE(YEAR(TODAY()),1,1)'),
                                ('Data final', '=TODAY()')], start=2):
    ls.cell(i, 11, rot).font = font(9)
    c = ls.cell(i, 12, val); c.font = font(10, True, NAVY2); c.number_format = DATA
    c.fill = fill(INPUT)
ls['K5'] = 'Altere estas duas datas para recortar o período de todo o painel.'
ls['K5'].font = font(8, False, MUTED, True)
ls['K7'] = 'Como o score é calculado'
ls['K7'].font = font(9, True, NAVY)
for i, txt in enumerate([
    'Score = (Preço × 35% + Atendimento × 35% + Qualidade × 30%) × (1 − índice de problemas)',
    'As três notas são de 1 a 5 e você atribui. O índice de problemas é medido, não opinado.',
    'Um fornecedor nota 5 em tudo com 20% de problemas cai para 4,00 — o histórico pesa.',
], start=8):
    ls.cell(i, 11, txt).font = font(8.5, c='41505D')

# ══════════════════════════════════════════════ ABA · Fornecedores
fo = wb.create_sheet('Fornecedores')
fo.sheet_view.showGridLines = False
HDR = ['Código', 'Empresa', 'Segmento', 'Insumos que fornece', 'CNPJ', 'Vendedor',
       'Contato / WhatsApp', 'E-mail', 'Chave PIX', 'Tipo da chave', 'Métodos de pagamento',
       'Condição padrão', 'Situação', 'Preço (1–5)', 'Atendimento (1–5)', 'Qualidade (1–5)',
       'Compras no período', '% de compra', 'Nº de compras', 'Ocorrências',
       'Índice de problemas', 'Entregas no prazo', 'Score geral', 'Última compra',
       'Pontos de atenção']
W = [10, 34, 24, 40, 20, 20, 20, 26, 26, 14, 22, 16, 14, 11, 13, 12, 17, 11, 11, 11, 13, 13, 11, 13, 44]
# mapa de colunas
C_COD, C_EMP, C_SEG, C_INS, C_CNPJ, C_VEND, C_CONT, C_MAIL = 1, 2, 3, 4, 5, 6, 7, 8
C_PIX, C_TPIX, C_MET, C_CPAG, C_SIT = 9, 10, 11, 12, 13
C_PRE, C_ATD, C_QUA = 14, 15, 16
C_COMP, C_PCT, C_NCOMP, C_OCOR, C_IPROB, C_PRAZO, C_SCORE, C_ULT, C_ATEN = 17, 18, 19, 20, 21, 22, 23, 24, 25

r = faixa_marca(fo, len(HDR), 'GESTÃO DE FORNECEDORES',
                'Cadastro único · fundo creme = preencher · fundo cinza-azulado = calculado a partir das abas Compras e Ocorrências')
r += 1
FHEAD = r
r = cab_tabela(fo, r, HDR, W)
FFIRST = r; FLAST = FFIRST + 59          # 60 fornecedores

CMP = lambda c: f"Compras!${c}$7:${c}$306"
OCR = lambda c: f"Ocorrências!${c}$7:${c}$156"
DT0, DT1 = 'Listas!$L$2', 'Listas!$L$3'

for i, (emp, seg, ins) in enumerate(SEED):
    row = FFIRST + i
    fo.cell(row, C_COD, f'FOR-{i+1:03d}').font = F9COD
    fo.cell(row, C_EMP, emp).font = font(9, True, NAVY)
    fo.cell(row, C_SEG, seg).font = F9
    fo.cell(row, C_INS, ins).font = F9
    fo.cell(row, C_SIT, 'Ativo').font = F9

for row in range(FFIRST, FLAST + 1):
    fo.row_dimensions[row].height = 22
    for col in range(1, len(HDR) + 1):
        c = fo.cell(row, col)
        if c.value is None: c.font = F9
        c.border = GRID
        c.alignment = LEFT if col in (C_EMP, C_SEG, C_INS, C_VEND, C_MAIL, C_PIX,
                                      C_MET, C_CPAG, C_SIT, C_ATEN) else CTR
    for col in (C_COD, C_EMP, C_SEG, C_INS, C_CNPJ, C_VEND, C_CONT, C_MAIL, C_PIX, C_TPIX,
                C_MET, C_CPAG, C_SIT, C_PRE, C_ATD, C_QUA, C_ATEN):
        fo.cell(row, col).fill = FILL_IN
    # ── métricas medidas ──
    fo.cell(row, C_COMP,  f'=SUMIFS({CMP("F")},{CMP("B")},$A{row},{CMP("A")},">="&{DT0},{CMP("A")},"<="&{DT1})')
    fo.cell(row, C_PCT,   f'=IF($Q${FLAST+1}=0,"",$Q{row}/$Q${FLAST+1})')
    fo.cell(row, C_NCOMP, f'=COUNTIFS({CMP("B")},$A{row},{CMP("A")},">="&{DT0},{CMP("A")},"<="&{DT1})')
    fo.cell(row, C_OCOR,  f'=COUNTIFS({OCR("B")},$A{row},{OCR("A")},">="&{DT0},{OCR("A")},"<="&{DT1})')
    fo.cell(row, C_IPROB, f'=IF($S{row}=0,"",$T{row}/$S{row})')
    # entregas no prazo — denominador contado explicitamente ("<>" contaria fórmula que devolve "")
    per = f'{CMP("A")},">="&{DT0},{CMP("A")},"<="&{DT1}'
    no_prazo   = f'COUNTIFS({CMP("B")},$A{row},{CMP("I")},"Sim",{per})'
    fora_prazo = f'COUNTIFS({CMP("B")},$A{row},{CMP("I")},"Não",{per})'
    fo.cell(row, C_PRAZO, f'=IF({no_prazo}+{fora_prazo}=0,"",{no_prazo}/({no_prazo}+{fora_prazo}))')
    fo.cell(row, C_SCORE, f'=IF(COUNT($N{row}:$P{row})<3,"",'
                          f'ROUND(($N{row}*0.35+$O{row}*0.35+$P{row}*0.3)*(1-IF($U{row}="",0,$U{row})),2))')
    fo.cell(row, C_ULT,   f'=IF($S{row}=0,"",SUMPRODUCT(MAX(({CMP("B")}=$A{row})*{CMP("A")})))')
    for col in (C_COMP, C_PCT, C_NCOMP, C_OCOR, C_IPROB, C_PRAZO, C_SCORE, C_ULT):
        fo.cell(row, col).font = F9CALC; fo.cell(row, col).fill = FILL_CALC
    fo.cell(row, C_COMP).number_format = MOEDA0
    fo.cell(row, C_PCT).number_format = PCT1
    fo.cell(row, C_IPROB).number_format = PCT1
    fo.cell(row, C_PRAZO).number_format = PCT
    fo.cell(row, C_SCORE).number_format = N2
    fo.cell(row, C_ULT).number_format = DATA

# linha de total
rt = FLAST + 1
fo.merge_cells(start_row=rt, start_column=1, end_row=rt, end_column=C_QUA)
c = fo.cell(rt, 1, '   TOTAL DO PERÍODO')
c.font = font(9.5, True, WHITE); c.fill = fill(NAVY); c.alignment = LEFT
for col, formula, fmt in ((C_COMP, f'=SUM(Q{FFIRST}:Q{FLAST})', MOEDA0),
                          (C_PCT, f'=IF($Q${rt}=0,"",1)', PCT),
                          (C_NCOMP, f'=SUM(S{FFIRST}:S{FLAST})', '#,##0'),
                          (C_OCOR, f'=SUM(T{FFIRST}:T{FLAST})', '#,##0'),
                          (C_IPROB, f'=IF($S${rt}=0,"",$T${rt}/$S${rt})', PCT1)):
    cc = fo.cell(rt, col, formula)
    cc.font = font(9.5, True, GOLDS); cc.fill = fill(NAVY); cc.alignment = CTR; cc.number_format = fmt
for col in range(1, len(HDR) + 1):
    fo.cell(rt, col).fill = fill(NAVY)
fo.row_dimensions[rt].height = 22

DVF = {'C': f'=Listas!$A$2:$A${1+len(SEGMENTOS)}', 'J': f'=Listas!$E$2:$E${1+len(TIPO_PIX)}',
       'K': f'=Listas!$B$2:$B${1+len(METODOS)}',   'L': f'=Listas!$C$2:$C${1+len(CONDICOES)}',
       'M': f'=Listas!$D$2:$D${1+len(SITUACOES)}',
       'N': '=Listas!$I$2:$I$6', 'O': '=Listas!$I$2:$I$6', 'P': '=Listas!$I$2:$I$6'}
for col, f_ in DVF.items():
    dv(fo, f_, f'{col}{FFIRST}:{col}{FLAST}')

# semáforo no score e no índice de problemas
fo.conditional_formatting.add(f'W{FFIRST}:W{FLAST}',
    ColorScaleRule(start_type='num', start_value=1, start_color='F5C6C4',
                   mid_type='num', mid_value=3, mid_color='FDF3D0',
                   end_type='num', end_value=5, end_color='CBE7D5'))
fo.conditional_formatting.add(f'U{FFIRST}:U{FLAST}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.2'],
               font=Font(name=F, size=9, bold=True, color=RED), fill=fill('FDF0EF')))
fo.conditional_formatting.add(f'R{FFIRST}:R{FLAST}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.4'],
               font=Font(name=F, size=9, bold=True, color=AMBER), fill=fill('FCF4E2')))

fo.freeze_panes = f'C{FFIRST}'
fo.auto_filter.ref = f'A{FHEAD}:Y{FLAST}'
fo.cell(FHEAD, C_PCT).comment = Comment(
    'Percentual do valor comprado no período que foi com este fornecedor.\n'
    'Acima de 40% fica destacado — é concentração que vira risco.', 'Valvic', width=280, height=80)
fo.cell(FHEAD, C_SCORE).comment = Comment(
    'Score = (Preço×35% + Atendimento×35% + Qualidade×30%) × (1 − índice de problemas).\n'
    'Só aparece quando as três notas estiverem preenchidas.', 'Valvic', width=300, height=80)

lg = FLAST + 3
for k, texto in enumerate([
    'COMO USAR ESTA ABA',
    'Células de fundo creme são de preenchimento manual. As de fundo cinza-azulado são calculadas — não editar.',
    'Preço, Atendimento e Qualidade são notas de 1 a 5 que você atribui. Todo o resto é medido a partir dos lançamentos.',
    '% de compra, índice de problemas, entregas no prazo e última compra saem das abas Compras e Ocorrências.',
    'O período de análise é controlado por duas datas na aba Listas — mude lá e o painel inteiro se recorta.',
    'Concentração acima de 40% num único fornecedor fica destacada em âmbar. Índice de problemas acima de 20%, em vermelho.',
]):
    fo.merge_cells(start_row=lg + k, start_column=1, end_row=lg + k, end_column=8)
    cc = fo.cell(lg + k, 1, texto if k == 0 else f'{k}.  {texto}')
    cc.font = font(9, True, NAVY) if k == 0 else font(9, c='41505D')
    if k == 0: cc.fill = fill(GOLDBG)
print_cfg(fo, f'A1:Y{FLAST}', retrato=False)

# ══════════════════════════════════════════════ ABA · Compras
cp = wb.create_sheet('Compras')
cp.sheet_view.showGridLines = False
H = ['Data', 'Cód. fornecedor', 'Fornecedor', 'Nota fiscal / pedido', 'Projeto ou obra',
     'Valor', 'Prazo prometido', 'Data de entrega', 'No prazo?', 'Forma de pagamento', 'Observações']
Wc = [12, 14, 34, 20, 26, 15, 15, 15, 11, 22, 40]
r = faixa_marca(cp, len(H), 'COMPRAS',
                'Uma linha por compra · é daqui que saem o volume, o percentual e a pontualidade de cada fornecedor')
r += 1
r = cab_tabela(cp, r, H, Wc)
CFIRST = r; CLAST = r + 299
FOR_A = f"Fornecedores!$A${FFIRST}:$A${FLAST}"
FOR_B = f"Fornecedores!$B${FFIRST}:$B${FLAST}"
for row in range(CFIRST, CLAST + 1):
    cp.row_dimensions[row].height = 20
    for col in range(1, len(H) + 1):
        c = cp.cell(row, col); c.font = F9; c.fill = FILL_IN; c.border = GRID
        c.alignment = LEFT if col in (3, 4, 5, 10, 11) else CTR
    cp.cell(row, 3, f'=IFERROR(INDEX({FOR_B},MATCH($B{row},{FOR_A},0)),"")')
    cp.cell(row, 9, f'=IF(OR($G{row}="",$H{row}=""),"",IF($H{row}<=$G{row},"Sim","Não"))')
    for col in (3, 9):
        cp.cell(row, col).font = F9CALC; cp.cell(row, col).fill = FILL_CALC
    for col in (1, 7, 8): cp.cell(row, col).number_format = DATA
    cp.cell(row, 6).number_format = MOEDA
dv(cp, f'={FOR_A}', f'B{CFIRST}:B{CLAST}')
dv(cp, f'=Listas!$B$2:$B${1+len(METODOS)}', f'J{CFIRST}:J{CLAST}')
cp.conditional_formatting.add(f'I{CFIRST}:I{CLAST}',
    CellIsRule(operator='equal', formula=['"Não"'],
               font=Font(name=F, size=9, bold=True, color=RED), fill=fill('FDF0EF')))
cp.freeze_panes = f'A{CFIRST}'
cp.auto_filter.ref = f'A{CFIRST-1}:K{CLAST}'
print_cfg(cp, f'A1:K{CLAST}', retrato=False)

# ══════════════════════════════════════════════ ABA · Ocorrências
oc = wb.create_sheet('Ocorrências')
oc.sheet_view.showGridLines = False
H = ['Data', 'Cód. fornecedor', 'Fornecedor', 'Nota fiscal / pedido', 'Tipo de ocorrência',
     'Gravidade', 'O que aconteceu', 'Providência tomada', 'Resolvido?', 'Data da solução']
Wo = [12, 14, 32, 20, 28, 12, 46, 40, 12, 14]
r = faixa_marca(oc, len(H), 'OCORRÊNCIAS',
                'Todo problema registrado aqui alimenta o índice de problemas do fornecedor')
r += 1
r = cab_tabela(oc, r, H, Wo)
OFIRST = r; OLAST = r + 149
for row in range(OFIRST, OLAST + 1):
    oc.row_dimensions[row].height = 20
    for col in range(1, len(H) + 1):
        c = oc.cell(row, col); c.font = F9; c.fill = FILL_IN; c.border = GRID
        c.alignment = LEFT if col in (3, 4, 5, 7, 8) else CTR
    oc.cell(row, 3, f'=IFERROR(INDEX({FOR_B},MATCH($B{row},{FOR_A},0)),"")')
    oc.cell(row, 3).font = F9CALC; oc.cell(row, 3).fill = FILL_CALC
    for col in (1, 10): oc.cell(row, col).number_format = DATA
dv(oc, f'={FOR_A}', f'B{OFIRST}:B{OLAST}')
dv(oc, f'=Listas!$F$2:$F${1+len(OCORRENCIAS)}', f'E{OFIRST}:E{OLAST}')
dv(oc, f'=Listas!$G$2:$G${1+len(GRAVIDADE)}', f'F{OFIRST}:F{OLAST}')
dv(oc, '=Listas!$H$2:$H$3', f'I{OFIRST}:I{OLAST}')
oc.conditional_formatting.add(f'F{OFIRST}:F{OLAST}',
    CellIsRule(operator='equal', formula=['"Alta"'],
               font=Font(name=F, size=9, bold=True, color=RED), fill=fill('FDF0EF')))
oc.freeze_panes = f'A{OFIRST}'
oc.auto_filter.ref = f'A{OFIRST-1}:J{OLAST}'
print_cfg(oc, f'A1:J{OLAST}', retrato=False)

# ══════════════════════════════════════════════ ABA · Dashboard
db = wb.create_sheet('Dashboard', 0)
db.sheet_view.showGridLines = False
NC = 12
for i, w in enumerate([3, 30, 15, 15, 4, 26, 15, 15, 4, 26, 14, 14], start=1):
    db.column_dimensions[get_column_letter(i)].width = w
FR = lambda c: f"Fornecedores!${c}${FFIRST}:${c}${FLAST}"
r = faixa_marca(db, NC, 'PAINEL DE FORNECEDORES',
                'Tudo calculado a partir das abas Fornecedores, Compras e Ocorrências · período controlado na aba Listas')
r += 1

def kpi(row, col, rotulo, formula, fmt, cor=NAVY, larg=3):
    db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + larg - 1)
    c = db.cell(row, col, rotulo.upper())
    c.font = font(8, True, MUTED); c.fill = fill(WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = Border(top=side(LINE), left=side(GOLD, 'medium'), right=side(LINE))
    db.row_dimensions[row].height = 17
    db.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + larg - 1)
    v = db.cell(row + 1, col, formula)
    v.font = Font(name=F, size=18, bold=True, color=cor); v.fill = fill(WHITE)
    v.number_format = fmt
    v.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    v.border = Border(bottom=side(LINE), left=side(GOLD, 'medium'), right=side(LINE))
    db.row_dimensions[row + 1].height = 30

r = titulo_secao(db, r, NC, 'Posição do período')
r += 1
kpi(r, 2, 'Fornecedores cadastrados', f'=COUNTIF({FR("A")},"?*")', '#,##0')
kpi(r, 6, 'Compras no período', f'=SUM({FR("Q")})', MOEDA0, NAVY)
kpi(r, 10, 'Nº de compras', f'=SUM({FR("S")})', '#,##0', BLUE)
r += 3
kpi(r, 2, 'Ocorrências registradas', f'=SUM({FR("T")})', '#,##0', RED)
kpi(r, 6, 'Índice de problemas geral',
    f'=IF(SUM({FR("S")})=0,"",SUM({FR("T")})/SUM({FR("S")}))', PCT1, RED)
kpi(r, 10, 'Maior concentração', f'=IF(SUM({FR("Q")})=0,"",MAX({FR("R")}))', PCT1, AMBER)
r += 3
kpi(r, 2, 'Fornecedores ativos', f'=COUNTIF({FR("M")},"Ativo")+COUNTIF({FR("M")},"Preferencial")', '#,##0', OK)
kpi(r, 6, 'Score médio da base',
    f'=IF(COUNT({FR("W")})=0,"",ROUND(AVERAGE({FR("W")}),2))', N2, OK)
kpi(r, 10, 'Fornecedores restritos ou bloqueados',
    f'=COUNTIF({FR("M")},"Restrito")+COUNTIF({FR("M")},"Bloqueado")', '#,##0', RED)
r += 3

def bloco(row, col, titulo, chaves, formulas, cabs, larg_key=None):
    db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    c = db.cell(row, col, '  ' + titulo.upper())
    c.font = font(8.5, True, WHITE); c.fill = fill(NAVY2); c.alignment = LEFT
    db.row_dimensions[row].height = 19
    for i, h in enumerate(cabs):
        cc = db.cell(row + 1, col + i, h)
        cc.font = font(8, True, NAVY); cc.fill = fill(GOLDBG)
        cc.alignment = CTR if i else LEFT; cc.border = Border(bottom=side(GOLD))
    for k, chave in enumerate(chaves):
        rr = row + 2 + k
        a = db.cell(rr, col, chave); a.font = font(9); a.alignment = LEFT
        for j, f_ in enumerate(formulas, start=1):
            cc = db.cell(rr, col + j, f_[0].format(r=rr, c=get_column_letter(col)))
            cc.font = font(9, c=NAVY2 if j == 1 else NAVY); cc.alignment = CTR if j == 1 else RIGHT
            cc.number_format = f_[1]; cc.border = BOTTOM2
        a.border = BOTTOM2
        db.row_dimensions[rr].height = 17
    return row + 2, row + 1 + len(chaves)

r = titulo_secao(db, r, NC, 'Onde o dinheiro está indo')
r += 1
b1 = bloco(r, 2, 'Por segmento', SEGMENTOS,
           [(f'=COUNTIF({FR("C")},${{c}}{{r}})', '#,##0'),
            (f'=SUMIF({FR("C")},${{c}}{{r}},{FR("Q")})', MOEDA0)],
           ('', 'Fornec.', 'Compras'))
b2 = bloco(r, 6, 'Por situação', SITUACOES,
           [(f'=COUNTIF({FR("M")},${{c}}{{r}})', '#,##0'),
            (f'=SUMIF({FR("M")},${{c}}{{r}},{FR("Q")})', MOEDA0)],
           ('', 'Fornec.', 'Compras'))
b3 = bloco(r, 10, 'Ocorrências por tipo', OCORRENCIAS,
           [(f'=COUNTIF(Ocorrências!$E$7:$E$156,${{c}}{{r}})', '#,##0'),
            (f'=COUNTIFS(Ocorrências!$E$7:$E$156,${{c}}{{r}},Ocorrências!$F$7:$F$156,"Alta")', '#,##0')],
           ('', 'Total', 'Graves'))
r = max(b1[1], b2[1], b3[1]) + 1
for col, (ini, fim) in ((2, b1), (6, b2), (10, b3)):
    a = db.cell(r, col, 'Total'); a.font = font(9, True, WHITE); a.fill = fill(NAVY)
    a.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for j, fmt in ((1, '#,##0'), (2, MOEDA0 if col < 10 else '#,##0')):
        L = get_column_letter(col + j)
        cc = db.cell(r, col + j, f'=SUM({L}{ini}:{L}{fim})')
        cc.font = font(9, True, GOLDS); cc.fill = fill(NAVY)
        cc.alignment = CTR if j == 1 else RIGHT; cc.number_format = fmt
    db.row_dimensions[r].height = 19
r += 2

r = titulo_secao(db, r, NC, 'Como ler este painel')
r += 1
for texto in [
    'Nada aqui é digitado — tudo vem das abas Fornecedores, Compras e Ocorrências.',
    'O período de análise é definido pelas duas datas na aba Listas. Mudou lá, muda tudo aqui.',
    '"Maior concentração" é o percentual de compra do fornecedor mais forte. Acima de 40%, a dependência já é um risco de negociação e de abastecimento.',
    'O índice de problemas é ocorrências dividido por compras — é medida, não impressão. Um fornecedor barato com 30% de problemas custa caro.',
    'Preço, atendimento e qualidade continuam sendo julgamento seu; o score apenas combina esse julgamento com o histórico real.',
]:
    db.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
    c = db.cell(r, 2, '•  ' + texto); c.font = font(9, c='41505D'); c.alignment = LEFT
    db.row_dimensions[r].height = 16
    r += 1
print_cfg(db, f'A1:{get_column_letter(NC)}{r}', retrato=False)

# ══════════════════════════════════════════════ ABA · Ficha do Fornecedor
fi = wb.create_sheet('Ficha do Fornecedor')
fi.sheet_view.showGridLines = False
NCF = 6
for i, w in enumerate([20, 26, 16, 16, 14, 14], start=1):
    fi.column_dimensions[get_column_letter(i)].width = w
r = faixa_marca(fi, NCF, 'FICHA DO FORNECEDOR',
                'Escolha o código no campo abaixo — todo o restante é preenchido sozinho')
r += 1
r = titulo_secao(fi, r, NCF, 'Fornecedor')
fi.cell(r, 1, 'Código').font = font(9, True, MUTED)
sel = fi.cell(r, 2, 'FOR-001')
sel.font = font(13, True, NAVY); sel.fill = FILL_IN; sel.alignment = CTR
sel.border = Border(bottom=side(GOLD, 'medium'))
dv(fi, f'={FOR_A}', f'B{r}:B{r}')
SEL = f'$B${r}'
fi.merge_cells(start_row=r, start_column=3, end_row=r, end_column=NCF)
c = fi.cell(r, 3, f'=IFERROR(INDEX({FOR_B},MATCH({SEL},{FOR_A},0)),"— código não encontrado —")')
c.font = font(15, True, NAVY); c.alignment = LEFT
fi.row_dimensions[r].height = 30
r += 2

def campo(row, rot, col_src, fmt=None, larg=None):
    fi.cell(row, 1, rot).font = font(9, True, MUTED)
    fi.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCF)
    v = fi.cell(row, 2, f'=IFERROR(INDEX(Fornecedores!${col_src}${FFIRST}:${col_src}${FLAST},'
                        f'MATCH({SEL},{FOR_A},0)),"")')
    v.font = font(10.5, True, NAVY); v.alignment = LEFT
    v.border = Border(bottom=side(LINE2))
    if fmt: v.number_format = fmt
    fi.row_dimensions[row].height = 21
    return row + 1

r = titulo_secao(fi, r, NCF, 'Identificação')
for rot, col in [('Segmento', 'C'), ('Insumos que fornece', 'D'), ('CNPJ', 'E'), ('Situação', 'M')]:
    r = campo(r, rot, col)
r += 1
r = titulo_secao(fi, r, NCF, 'Contato')
for rot, col in [('Vendedor', 'F'), ('Contato / WhatsApp', 'G'), ('E-mail', 'H')]:
    r = campo(r, rot, col)
r += 1
r = titulo_secao(fi, r, NCF, 'Pagamento')
for rot, col in [('Chave PIX', 'I'), ('Tipo da chave', 'J'),
                 ('Métodos aceitos', 'K'), ('Condição padrão', 'L')]:
    r = campo(r, rot, col)
r += 1

r = titulo_secao(fi, r, NCF, 'Avaliação e histórico', 'notas de 1 a 5 · demais números medidos no período')
IND = [('Preço', 'N', N2), ('Atendimento', 'O', N2), ('Qualidade', 'P', N2),
       ('Score geral', 'W', N2), ('Compras no período', 'Q', MOEDA0), ('% de compra', 'R', PCT1),
       ('Nº de compras', 'S', '#,##0'), ('Ocorrências', 'T', '#,##0'),
       ('Índice de problemas', 'U', PCT1), ('Entregas no prazo', 'V', PCT),
       ('Última compra', 'X', DATA)]
for i in range(0, len(IND), 2):
    for j, (rot, col, fmt) in enumerate(IND[i:i + 2]):
        c0 = 1 + j * 3
        cc = fi.cell(r, c0, rot); cc.font = font(9, True, MUTED); cc.alignment = LEFT
        fi.merge_cells(start_row=r, start_column=c0 + 1, end_row=r, end_column=c0 + 2)
        v = fi.cell(r, c0 + 1, f'=IFERROR(INDEX(Fornecedores!${col}${FFIRST}:${col}${FLAST},'
                               f'MATCH({SEL},{FOR_A},0)),"")')
        v.font = font(12, True, NAVY2); v.alignment = LEFT; v.number_format = fmt
        v.border = Border(bottom=side(LINE2))
    fi.row_dimensions[r].height = 23
    r += 1
r += 1

r = titulo_secao(fi, r, NCF, 'Pontos de atenção')
fi.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=NCF)
c = fi.cell(r, 1, f'=IFERROR(INDEX(Fornecedores!$Y${FFIRST}:$Y${FLAST},MATCH({SEL},{FOR_A},0)),"")')
c.font = font(10, c='33414F'); c.alignment = LEFTW
c.border = Border(left=side(LINE), right=side(LINE), top=side(LINE), bottom=side(LINE))
c.fill = fill('FDFBF6')
for k in range(4): fi.row_dimensions[r + k].height = 20
r += 5
fi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCF)
c = fi.cell(r, 1, 'Ficha gerada a partir do cadastro de fornecedores · '
                  'os indicadores refletem o período definido na aba Listas.')
c.font = font(8, False, MUTED, True); c.alignment = LEFT
print_cfg(fi, f'A1:F{r}', retrato=True)

# ── ordem final
ordem = ['Dashboard', 'Fornecedores', 'Compras', 'Ocorrências', 'Ficha do Fornecedor', 'Listas']
wb._sheets = [wb[t] for t in ordem]
wb.active = 0
out = 'Valvic_Gestao_Fornecedores.xlsx'
wb.save(out)
print(f'gerado · {len(wb.worksheets)} abas · {out}')
