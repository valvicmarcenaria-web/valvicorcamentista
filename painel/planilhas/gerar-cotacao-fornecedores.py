#!/usr/bin/env python3
"""Gera a planilha de Cotação Comparativa de Fornecedores da Valvic.

Abas: Instruções · Cotação · Pedido de Cotação (A4) · Exemplo · Mapa de Cotações · Listas

Uso:  python3 gerar-cotacao-fornecedores.py
Saída: Valvic_Cotacao_Fornecedores.xlsx
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.formatting.rule import FormulaRule

# ─────────────────────────────────────────────── paleta Valvic
NAVY, NAVY2 = '0E2038', '16314F'
GOLD, GOLDS, GOLDBG = 'C2A05A', 'D8BD80', 'F6EDD6'
INK, MUTED = '1B2733', '6C7785'
LINE, LINE2 = 'E8E3D8', 'DFDACD'
OK, BLUE, RED, AMBER = '2F7D4F', '2F5D8C', 'B0413F', 'B57A16'
OKBG, REDBG, AMBBG, BLUEBG = 'E2F0E7', 'FBE7E6', 'FBF1DC', 'E7EEF6'
INPUT, WHITE, CALC = 'FFF9E3', 'FFFFFF', 'F4F6F8'

F = 'Arial'
def font(sz=10, b=False, c=INK, i=False):
    return Font(name=F, size=sz, bold=b, color=c, italic=i)
def fill(c): return PatternFill('solid', fgColor=c)
def side(c=LINE, st='thin'): return Side(style=st, color=c)
GRID    = Border(bottom=side(LINE), left=side(LINE), right=side(LINE))
BOTTOM2 = Border(bottom=side(LINE2))
CTR   = Alignment(horizontal='center', vertical='center')
CTRW  = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT  = Alignment(horizontal='left', vertical='center')
LEFTI = Alignment(horizontal='left', vertical='center', indent=1)
LEFTW = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFTIW = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
RIGHT = Alignment(horizontal='right', vertical='center', indent=1)
MOEDA, MOEDA0, PCT1, DATA, QTD = 'R$ #,##0.00', 'R$ #,##0', '0.0%', 'DD/MM/YYYY', '#,##0.##'
DIAS = '0 "dias"'

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════ listas
UNIDADES = ['un', 'pç', 'cj', 'jg', 'cx', 'pct', 'm', 'm²', 'm³', 'ml', 'kg', 'L',
            'chapa', 'barra', 'rolo', 'par', 'serviço', 'verba']
CONDICOES = ['À vista', '7 dias', '14 dias', '21 dias', '28 dias', '30 dias', '30/60',
             '30/60/90', '2x sem juros', '3x sem juros', 'Parcelado no cartão',
             'Entrada + saldo', 'A combinar']
FORMAS = ['PIX à vista', 'Boleto à vista', 'Boleto a prazo', 'Cartão de crédito',
          'Transferência', 'Dinheiro', 'A combinar']
SIT_COT = ['Em cotação', 'Aguardando fornecedor', 'Em análise', 'Fechada',
           'Cancelada', 'Suspensa']
PRIORID = ['Normal', 'Urgente', 'Programada']

# ══════════════════════════════════════════════ helpers
def faixa_marca(ws, ncols, titulo, sub, linha=1):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ncols)
    c = ws.cell(linha, 1, 'VALVIC MARCENARIA')
    c.font = Font(name=F, size=13, bold=True, color=WHITE); c.alignment = LEFTI
    ws.row_dimensions[linha].height = 26
    ws.merge_cells(start_row=linha+1, start_column=1, end_row=linha+1, end_column=ncols)
    c = ws.cell(linha+1, 1, 'Vargas Decor Ltda   ·   CNPJ 17.269.304/0001-51   ·   Belo Horizonte / MG')
    c.font = Font(name=F, size=8, color='9FB0C4'); c.alignment = LEFTI
    ws.row_dimensions[linha+1].height = 15
    ws.merge_cells(start_row=linha+2, start_column=1, end_row=linha+2, end_column=ncols)
    c = ws.cell(linha+2, 1, titulo)
    c.font = Font(name=F, size=15, bold=True, color=WHITE); c.alignment = LEFTI
    ws.row_dimensions[linha+2].height = 28
    ws.merge_cells(start_row=linha+3, start_column=1, end_row=linha+3, end_column=ncols)
    c = ws.cell(linha+3, 1, sub)
    c.font = Font(name=F, size=8.5, color='7A5B17', italic=True); c.alignment = LEFTI
    ws.row_dimensions[linha+3].height = 17
    for lr in range(linha, linha+4):
        bg = NAVY if lr < linha+2 else (NAVY2 if lr == linha+2 else GOLDBG)
        for cc in range(1, ncols+1):
            ws.cell(lr, cc).fill = fill(bg)
    return linha + 4

def titulo_secao(ws, row, ncols, texto, nota=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, ('  ' + texto.upper()) + (f'          {nota}' if nota else ''))
    c.font = Font(name=F, size=9, bold=True, color=NAVY); c.alignment = LEFT
    for cc in range(1, ncols+1):
        cel = ws.cell(row, cc)
        cel.fill = fill(GOLDBG); cel.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 20
    return row + 1

def cab_tabela(ws, row, headers, widths=None, alt=32):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(name=F, size=8.5, bold=True, color=WHITE); c.fill = fill(NAVY2)
        c.alignment = CTRW
        c.border = Border(left=side(NAVY2), right=side(NAVY2), bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = alt
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1

def print_cfg(ws, area, retrato=True, margens=(0.4, 0.3, 0.4, 0.3)):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait' if retrato else 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.print_area = area
    l, r, t, b = margens
    ws.page_margins.left, ws.page_margins.right = l, r
    ws.page_margins.top, ws.page_margins.bottom = t, b
    ws.sheet_view.showGridLines = False

def dv(ws, formula, rng):
    d = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(d); d.add(rng)

def col_lista(ws, col, titulo, valores, larg=26):
    ws.column_dimensions[col].width = larg
    c = ws[f'{col}1']; c.value = titulo
    c.font = font(9, True, WHITE); c.fill = fill(NAVY2); c.alignment = CTR
    for i, v in enumerate(valores, start=2):
        cc = ws[f'{col}{i}']; cc.value = v; cc.font = font(10); cc.border = BOTTOM2

def bloco(ws, row, col, span, valor=None, *, f=None, bg=None, al=None, nf=None, bd=True):
    """Escreve numa faixa mesclada de `span` colunas a partir de `col`."""
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row, col, valor)
    if f:  c.font = f
    if al: c.alignment = al
    if nf: c.number_format = nf
    for k in range(span):
        cel = ws.cell(row, col+k)
        if bd: cel.border = GRID
        if bg: cel.fill = fill(bg)
    return c


# ══════════════════════════════════════════════════════════════════════════
#  ABA · Cotação  (modelo — duplicável)
# ══════════════════════════════════════════════════════════════════════════
NCOL = 20
NIT  = 30
R_ID, R_IDV = 6, 7
R_FORN, R_VEND, R_BAR = 10, 11, 12      # nome · vendedor · barra fixa de totais
R_HEAD = 13
R_IT0  = 14
R_ITF  = R_IT0 + NIT - 1                # 43
R_APT  = R_ITF + 2                      # 45
R_A0   = R_APT + 1                      # 46
AP = {k: R_A0 + i for i, k in enumerate([
    'itens', 'situacao', 'subsem', 'imposto', 'subtotal', 'frete', 'totbruto',
    'pdesc', 'avista', 'condprazo', 'pacresc', 'aprazo', 'custoprazo',
    'entrega', 'validade', 'subcompleto', 'completo'])}
R_VERT = AP['completo'] + 2
R_V0   = R_VERT + 1

# (col do unitário, letra unit, letra %imp, letra total)
TRIOS = [(5, 'E', 'F', 'G'), (8, 'H', 'I', 'J'), (11, 'K', 'L', 'M'), (14, 'N', 'O', 'P')]
C_MUNIT, C_MTOT, C_MFORN, C_OBS = 17, 18, 19, 20     # Q R S T
NOMES_PADRAO = ['Fornecedor 1', 'Fornecedor 2', 'Fornecedor 3', 'Fornecedor 4']

HDR_COT = ['#', 'ITEM / ESPECIFICAÇÃO', 'UN.', 'QTD'] + \
          ['UNIT. (R$)', '% IMP.', 'TOTAL (R$)'] * 4 + \
          ['MENOR UNIT.\nC/ IMPOSTO', 'MENOR TOTAL', 'MELHOR FORNECEDOR',
           'OBSERVAÇÃO DO ITEM']
W_COT = [5, 32, 6.5, 8] + [10.5, 7, 12] * 4 + [12, 12, 17, 20]

F_LAB  = Font(name=F, size=8, bold=True, color=MUTED)
F_VAL  = Font(name=F, size=10.5, bold=True, color=NAVY2)
F_FORN = Font(name=F, size=11, bold=True, color=NAVY)
F_ROT  = Font(name=F, size=9, color=INK)
F_NUM  = Font(name=F, size=9.5, color=NAVY2)
F_NUMI = Font(name=F, size=9.5, bold=True, color='7A5B17')
F_DEST = Font(name=F, size=11.5, bold=True, color=WHITE)


def montar_cotacao(ws, dados=None):
    ws.sheet_view.showGridLines = False
    r = faixa_marca(ws, NCOL, 'COTAÇÃO COMPARATIVA DE FORNECEDORES',
                    'Fundo creme = você preenche   ·   fundo cinza = calculado automaticamente   '
                    '·   deixe o preço em branco quando o fornecedor NÃO tiver o item')
    assert r == R_ID - 1, f'faixa terminou em {r}, esperado {R_ID-1}'
    ws.row_dimensions[r].height = 6

    # ── identificação
    ident = [(1, 2, 'COTAÇÃO Nº'), (3, 5, 'DEMANDA / PROJETO'), (8, 4, 'SOLICITANTE'),
             (12, 3, 'DATA DA COTAÇÃO'), (15, 6, 'PRAZO PARA RESPOSTA')]
    for col, span, rot in ident:
        bloco(ws, R_ID, col, span, rot, f=F_LAB, bg=WHITE, al=LEFTI, bd=False)
        bloco(ws, R_IDV, col, span, None, f=F_VAL, bg=INPUT, al=LEFTI)
    ws.row_dimensions[R_ID].height = 14
    ws.row_dimensions[R_IDV].height = 22
    ws[f'L{R_IDV}'].number_format = DATA
    ws[f'O{R_IDV}'].number_format = DATA
    ws.row_dimensions[R_ID + 2].height = 6

    # ── barra fixa: fornecedor · vendedor · situação/imposto/total
    titulo_secao(ws, R_FORN - 1, NCOL, 'Comparativo de propostas',
                 'esta faixa fica congelada no topo — o total de cada fornecedor '
                 'acompanha você enquanto digita os preços')
    bloco(ws, R_FORN, 1, 4, 'FORNECEDOR  ▸', f=Font(name=F, size=9.5, bold=True, color=NAVY),
          bg=CALC, al=RIGHT)
    bloco(ws, R_VEND, 1, 4, 'Vendedor · contato  ▸', f=Font(name=F, size=8.5, color=MUTED),
          bg=CALC, al=RIGHT)
    bloco(ws, R_BAR, 1, 4, 'ITENS  ·  % IMPOSTO PADRÃO  ·  TOTAL DO PEDIDO  ▸',
          f=Font(name=F, size=9, bold=True, color=WHITE), bg=NAVY2, al=RIGHT)
    for i, (c0, u, p, t) in enumerate(TRIOS):
        bloco(ws, R_FORN, c0, 3, NOMES_PADRAO[i], f=F_FORN, bg=INPUT, al=CTRW)
        bloco(ws, R_VEND, c0, 3, None, f=Font(name=F, size=8.5, color=INK), bg=INPUT, al=CTRW)
        # situação curta
        c = ws.cell(R_BAR, c0)
        c.value = (f'=IF(COUNTA($B${R_IT0}:$B${R_ITF})=0,"",'
                   f'COUNT({u}${R_IT0}:{u}${R_ITF})&"/"&COUNTA($B${R_IT0}:$B${R_ITF}))')
        c.font = Font(name=F, size=10, bold=True, color=WHITE); c.fill = fill(NAVY2)
        c.alignment = CTR; c.border = GRID
        # % de imposto padrão do fornecedor
        c = ws.cell(R_BAR, c0 + 1)
        c.number_format = PCT1; c.font = Font(name=F, size=9.5, bold=True, color='7A5B17')
        c.fill = fill(INPUT); c.alignment = CTR; c.border = GRID
        # total do pedido
        c = ws.cell(R_BAR, c0 + 2)
        c.value = f'=IF({u}{AP["totbruto"]}="","",{u}{AP["totbruto"]})'
        c.number_format = MOEDA; c.font = Font(name=F, size=11, bold=True, color=GOLDS)
        c.fill = fill(NAVY); c.alignment = CTR; c.border = GRID
    bloco(ws, R_FORN, C_MUNIT, 4, 'MELHOR PREÇO POR ITEM',
          f=Font(name=F, size=9.5, bold=True, color=WHITE), bg=NAVY2, al=CTR)
    bloco(ws, R_VEND, C_MUNIT, 4, 'compara pelo TOTAL do item, já com o imposto de cada um',
          f=Font(name=F, size=8, color=MUTED, i=True), bg=CALC, al=CTRW)
    bloco(ws, R_BAR, C_MUNIT, 4,
          'total do pedido = itens + imposto + frete   ·   desconto à vista e acréscimo a prazo ficam na apuração',
          f=Font(name=F, size=8, color=MUTED, i=True), bg=CALC, al=CTRW)
    ws.row_dimensions[R_FORN].height = 26
    ws.row_dimensions[R_VEND].height = 20
    ws.row_dimensions[R_BAR].height = 24

    # ── tabela de itens
    cab_tabela(ws, R_HEAD, HDR_COT, W_COT, alt=34)
    for r_ in range(R_IT0, R_ITF + 1):
        ws.row_dimensions[r_].height = 18
        c = ws.cell(r_, 1, r_ - R_IT0 + 1); c.font = Font(name=F, size=8.5, color=MUTED)
        c.alignment = CTR; c.border = GRID; c.fill = fill(CALC)
        for col, al, nf, fo in ((2, LEFTI, None, font(9.5)),
                                (3, CTR, None, font(9.5)),
                                (4, CTR, QTD, font(9.5, True, NAVY2))):
            c = ws.cell(r_, col); c.alignment = al; c.fill = fill(INPUT)
            c.border = GRID; c.font = fo
            if nf: c.number_format = nf
        for (c0, u, p, t) in TRIOS:
            c = ws.cell(r_, c0); c.number_format = MOEDA; c.fill = fill(INPUT)
            c.border = GRID; c.font = F_NUM; c.alignment = RIGHT
            c = ws.cell(r_, c0 + 1); c.number_format = PCT1; c.fill = fill(INPUT)
            c.border = GRID; c.font = Font(name=F, size=8.5, color='7A5B17')
            c.alignment = CTR
            c = ws.cell(r_, c0 + 2)
            c.value = (f'=IF(OR($D{r_}="",{u}{r_}=""),"",ROUND($D{r_}*{u}{r_}*'
                       f'(1+IF({p}{r_}="",IF({p}${R_BAR}="",0,{p}${R_BAR}),{p}{r_})),2))')
            c.number_format = MOEDA; c.fill = fill(CALC); c.border = GRID
            c.font = F_NUM; c.alignment = RIGHT
        TOTS = ','.join(f'{t}{r_}' for _, _, _, t in TRIOS)
        # menor total do item
        c = ws.cell(r_, C_MTOT)
        c.value = f'=IF(COUNT({TOTS})=0,"",MIN({TOTS}))'
        c.number_format = MOEDA; c.fill = fill(CALC); c.border = GRID
        c.font = Font(name=F, size=9.5, bold=True, color=OK); c.alignment = RIGHT
        # menor unitário já com imposto (derivado do menor total)
        c = ws.cell(r_, C_MUNIT)
        rt = get_column_letter(C_MTOT)
        c.value = (f'=IF(OR({rt}{r_}="",$D{r_}="",$D{r_}=0),"",'
                   f'ROUND({rt}{r_}/$D{r_},2))')
        c.number_format = MOEDA; c.fill = fill(CALC); c.border = GRID
        c.font = Font(name=F, size=9.5, color=OK); c.alignment = RIGHT
        # fornecedor vencedor da linha
        c = ws.cell(r_, C_MFORN)
        cond = f'${rt}{r_}'
        c.value = (f'=IF({cond}="","",IF({TRIOS[0][3]}{r_}={cond},${TRIOS[0][1]}${R_FORN},'
                   f'IF({TRIOS[1][3]}{r_}={cond},${TRIOS[1][1]}${R_FORN},'
                   f'IF({TRIOS[2][3]}{r_}={cond},${TRIOS[2][1]}${R_FORN},'
                   f'${TRIOS[3][1]}${R_FORN}))))')
        c.fill = fill(CALC); c.border = GRID
        c.font = Font(name=F, size=8.5, bold=True, color=NAVY2); c.alignment = CTRW
        c = ws.cell(r_, C_OBS); c.fill = fill(INPUT); c.border = GRID
        c.font = font(8.5); c.alignment = LEFTI

    dv(ws, "=Listas!$A$2:$A$19", f'C{R_IT0}:C{R_ITF}')

    # ── apuração por fornecedor
    ws.row_dimensions[R_ITF + 1].height = 8
    titulo_secao(ws, R_APT, NCOL, 'Apuração por fornecedor',
                 'as linhas creme são de preenchimento · as demais se calculam sozinhas')

    LINHAS = [
        ('itens',      'Itens cotados',                 'calc', None,  'Quantos dos itens da lista este fornecedor conseguiu cotar.'),
        ('situacao',   'Situação do pedido',            'calc', None,  'COMPLETO = tem tudo. PARCIAL = vai faltar item e você precisará de um segundo fornecedor.'),
        ('subsem',     'Subtotal sem imposto',          'calc', MOEDA, 'Soma de quantidade × preço unitário, do jeito que o fornecedor mandou.'),
        ('imposto',    'Imposto sobre os itens (R$)',   'calc', MOEDA, 'Vem do % de imposto da barra fixa, ou do % digitado na linha do item quando houver.'),
        ('subtotal',   'Subtotal com imposto',          'calc', MOEDA, ''),
        ('frete',      'Frete estimado (R$)',           'in',   MOEDA, 'Coloque 0 quando o frete estiver embutido ou for por conta do fornecedor.'),
        ('totbruto',   'TOTAL DO PEDIDO',               'calc', MOEDA, 'É este número que aparece na barra fixa lá em cima.'),
        ('pdesc',      '% desconto à vista',            'in',   PCT1,  'Desconto que o fornecedor dá para pagamento à vista.'),
        ('avista',     'TOTAL À VISTA',                 'dest', MOEDA, ''),
        ('condprazo',  'Condição a prazo',              'in',   None,  'Ex.: 30/60/90, 28 dias, 3x sem juros.'),
        ('pacresc',    '% acréscimo a prazo',           'in',   PCT1,  'Juros ou acréscimo cobrado no parcelado. Se não há acréscimo, deixe zerado.'),
        ('aprazo',     'TOTAL A PRAZO',                 'dest', MOEDA, ''),
        ('custoprazo', 'Custo do prazo (R$)',           'calc', MOEDA, 'Quanto o parcelamento custa. Compare com o que aquele fôlego de caixa vale.'),
        ('entrega',    'Prazo de entrega',              'in',   DIAS,  'Em dias corridos a partir do pedido.'),
        ('validade',   'Validade da proposta',          'in',   DATA,  'Até quando o fornecedor garante este preço.'),
        ('subcompleto', 'Subtotal se atender 100%',     'calc', MOEDA, 'Auxiliar do veredito — em branco quando o fornecedor não tem o pedido inteiro.'),
        ('completo',   'Total à vista se atender 100%', 'calc', MOEDA, 'Auxiliar do veredito — é daqui que sai o "melhor entre os completos".'),
    ]
    for chave, rotulo, tipo, nf, nota in LINHAS:
        rr = AP[chave]
        dest = tipo == 'dest'
        ws.row_dimensions[rr].height = 22 if dest else 17
        bloco(ws, rr, 1, 4, rotulo,
              f=(Font(name=F, size=9.5, bold=True, color=WHITE) if dest else
                 (Font(name=F, size=9, bold=True, color='7A5B17') if tipo == 'in' else F_ROT)),
              bg=(NAVY2 if dest else (GOLDBG if tipo == 'in' else WHITE)), al=RIGHT)
        for (c0, u, p, t) in TRIOS:
            bg = NAVY if dest else (INPUT if tipo == 'in' else CALC)
            fo = F_DEST if dest else (F_NUMI if tipo == 'in' else F_NUM)
            bloco(ws, rr, c0, 3, None, f=fo, bg=bg, al=CTR, nf=nf)
        bloco(ws, rr, C_MUNIT, 4, nota,
              f=Font(name=F, size=8, color=(GOLDS if dest else MUTED), i=True),
              bg=(NAVY2 if dest else WHITE), al=LEFTIW)

    NB = f'COUNTA($B${R_IT0}:$B${R_ITF})'
    for (c0, u, p, t) in TRIOS:
        NC = f'COUNT({u}${R_IT0}:{u}${R_ITF})'
        A = lambda k: f'{u}{AP[k]}'
        ws.cell(AP['itens'], c0).value = f'=IF({NB}=0,"",{NC}&" de "&{NB})'
        ws.cell(AP['situacao'], c0).value = (
            f'=IF({NB}=0,"",IF({NC}=0,"NÃO COTOU",IF({NC}={NB},"COMPLETO",'
            f'"PARCIAL — faltam "&({NB}-{NC})&" item(ns)")))')
        ws.cell(AP['subsem'], c0).value = (
            f'=IF({NC}=0,"",IFERROR(ROUND(SUMPRODUCT($D${R_IT0}:$D${R_ITF},'
            f'{u}${R_IT0}:{u}${R_ITF}),2),""))')
        ws.cell(AP['subtotal'], c0).value = (
            f'=IF({NC}=0,"",ROUND(SUM({t}${R_IT0}:{t}${R_ITF}),2))')
        ws.cell(AP['imposto'], c0).value = (
            f'=IF(OR({A("subtotal")}="",{A("subsem")}=""),"",'
            f'ROUND({A("subtotal")}-{A("subsem")},2))')
        ws.cell(AP['totbruto'], c0).value = (
            f'=IF({A("subtotal")}="","",ROUND({A("subtotal")}+'
            f'IF({A("frete")}="",0,{A("frete")}),2))')
        ws.cell(AP['avista'], c0).value = (
            f'=IF({A("totbruto")}="","",ROUND({A("totbruto")}*'
            f'(1-IF({A("pdesc")}="",0,{A("pdesc")})),2))')
        ws.cell(AP['aprazo'], c0).value = (
            f'=IF({A("totbruto")}="","",ROUND({A("totbruto")}*'
            f'(1+IF({A("pacresc")}="",0,{A("pacresc")})),2))')
        ws.cell(AP['custoprazo'], c0).value = (
            f'=IF(OR({A("avista")}="",{A("aprazo")}=""),"",'
            f'ROUND({A("aprazo")}-{A("avista")},2))')
        ws.cell(AP['subcompleto'], c0).value = (
            f'=IF({A("situacao")}="COMPLETO",{A("subtotal")},"")')
        ws.cell(AP['completo'], c0).value = (
            f'=IF({A("situacao")}="COMPLETO",{A("avista")},"")')
        ws.cell(AP['itens'], c0).alignment = CTR
        ws.cell(AP['situacao'], c0).font = Font(name=F, size=9, bold=True, color=NAVY2)
        dv(ws, "=Listas!$B$2:$B$14",
           f'{u}{AP["condprazo"]}:{t}{AP["condprazo"]}')

    # ── veredito
    ws.row_dimensions[R_VERT - 1].height = 8
    titulo_secao(ws, R_VERT, NCOL, 'Veredito',
                 'leitura pronta para decidir — confira sempre a situação do pedido')
    ref = lambda k: [f'${u}${AP[k]}' for _, u, _, _ in TRIOS]
    NM = [f'${u}${R_FORN}' for _, u, _, _ in TRIOS]

    def escolhe(base, alvo, destino):
        return (f'IF({base[0]}={alvo},{destino[0]},IF({base[1]}={alvo},{destino[1]},'
                f'IF({base[2]}={alvo},{destino[2]},{destino[3]})))')

    cond_txt = [f'IF({c}="","condição a prazo não informada",{c})' for c in ref('condprazo')]
    VER = [('Melhor preço À VISTA', ref('avista'), ref('situacao'), OKBG, OK),
           ('Melhor à vista ENTRE OS COMPLETOS', ref('completo'), None, BLUEBG, BLUE),
           ('Melhor preço A PRAZO', ref('aprazo'), cond_txt, AMBBG, AMBER)]
    rr = R_V0
    for rotulo, base, extra, bg, cor in VER:
        ws.row_dimensions[rr].height = 22
        bloco(ws, rr, 1, 4, rotulo, f=Font(name=F, size=9, bold=True, color=cor),
              bg=bg, al=RIGHT)
        mn = f'MIN({",".join(base)})'
        cond = f'COUNT({",".join(base)})=0'
        c = bloco(ws, rr, 5, 6, None, f=Font(name=F, size=10, bold=True, color=NAVY),
                  bg=bg, al=CTR)
        c.value = f'=IF({cond},"—",{escolhe(base, mn, NM)})'
        c = bloco(ws, rr, 11, 6, None, f=Font(name=F, size=12, bold=True, color=cor),
                  bg=bg, al=CTR, nf=MOEDA)
        c.value = f'=IF({cond},"",{mn})'
        c = bloco(ws, rr, C_MUNIT, 4, None, f=Font(name=F, size=8.5, color=INK),
                  bg=bg, al=CTRW)
        c.value = (f'=IF({cond},"—",{escolhe(base, mn, extra)})' if extra is not None else
                   f'=IF({cond},"Nenhum fornecedor tem o pedido completo — '
                   f'você vai precisar dividir a compra","Atende 100% dos itens do pedido")')
        rr += 1

    # compra fracionada
    RT = get_column_letter(C_MTOT)
    ws.row_dimensions[rr].height = 22
    bloco(ws, rr, 1, 4, 'COMPRA FRACIONADA (item a item)',
          f=Font(name=F, size=9, bold=True, color=NAVY2), bg=CALC, al=RIGHT)
    bloco(ws, rr, 5, 6, 'melhor preço de cada linha',
          f=Font(name=F, size=9, color=MUTED, i=True), bg=CALC, al=CTR)
    c = bloco(ws, rr, 11, 6, None, f=Font(name=F, size=12, bold=True, color=NAVY2),
              bg=CALC, al=CTR, nf=MOEDA)
    c.value = (f'=IF(SUM(${RT}${R_IT0}:${RT}${R_ITF})=0,"",'
               f'ROUND(SUM(${RT}${R_IT0}:${RT}${R_ITF}),2))')
    SUBC = ','.join(ref('subcompleto'))
    SEMCOT = f'(COUNTA($B${R_IT0}:$B${R_ITF})-COUNT(${RT}${R_IT0}:${RT}${R_ITF}))'
    c = bloco(ws, rr, C_MUNIT, 4, None, f=Font(name=F, size=8.5, color=INK), bg=CALC, al=CTRW)
    DIF = f'ROUND(MIN({SUBC})-$K{rr},2)'
    c.value = (
        f'=IF($K{rr}="","",'
        f'IF({SEMCOT}>0,"Atenção: "&{SEMCOT}&" item(ns) fora da comparação — sem quantidade '
        f'ou sem nenhum preço. ","")&'
        f'IF(COUNT({SUBC})=0,"Nenhum fornecedor tem o pedido completo para servir de comparação.",'
        f'IF({DIF}>0,'
        f'"Economia de "&TEXT(MIN({SUBC})-$K{rr},"R$ #,##0.00")&'
        f'" sobre o subtotal do melhor fornecedor completo — mas some frete de cada um.",'
        f'IF({DIF}=0,'
        f'"Não compensa dividir: o melhor fornecedor completo já tem o menor preço em todos os itens.",'
        f'"Sem ganho: dividir a compra sai "&TEXT($K{rr}-MIN({SUBC}),"R$ #,##0.00")&'
        f'" mais caro que o melhor fornecedor completo."))))')
    rr += 1

    # ── nota de rodapé
    ws.row_dimensions[rr].height = 8
    rr += 1
    ws.row_dimensions[rr].height = 44
    bloco(ws, rr, 1, NCOL,
          '  Como ler: a coluna % IMP. é o imposto que vem POR FORA do preço (o clássico "preços s/ IPI = 6,5%"). '
          'Deixe a linha em branco para usar o % padrão do fornecedor na barra fixa, ou digite um % diferente só naquele item. '
          'A comparação entre fornecedores é feita pelo TOTAL do item, já com o imposto de cada um — nunca pelo preço de tabela. '
          'Subtotal e compra fracionada são sem frete; a decisão final é pelo TOTAL À VISTA ou TOTAL A PRAZO.',
          f=Font(name=F, size=8.5, color='41505D'), bg=GOLDBG, al=LEFTIW)

    # ── formatação condicional
    for (c0, u, p, t) in TRIOS:
        ws.conditional_formatting.add(f'{u}{R_IT0}:{u}{R_ITF}', FormulaRule(
            formula=[f'AND($B{R_IT0}<>"",{u}{R_IT0}="")'], fill=fill(REDBG), stopIfTrue=True))
        ws.conditional_formatting.add(f'{t}{R_IT0}:{t}{R_ITF}', FormulaRule(
            formula=[f'AND({t}{R_IT0}<>"",${RT}{R_IT0}<>"",{t}{R_IT0}=${RT}{R_IT0})'],
            fill=fill(OKBG), font=Font(bold=True, color=OK)))
        rs = f'{u}{AP["situacao"]}:{t}{AP["situacao"]}'
        ws.conditional_formatting.add(rs, FormulaRule(
            formula=[f'${u}${AP["situacao"]}="COMPLETO"'], fill=fill(OKBG),
            font=Font(bold=True, color=OK), stopIfTrue=True))
        ws.conditional_formatting.add(rs, FormulaRule(
            formula=[f'LEFT(${u}${AP["situacao"]},7)="PARCIAL"'], fill=fill(AMBBG),
            font=Font(bold=True, color=AMBER), stopIfTrue=True))
        ws.conditional_formatting.add(rs, FormulaRule(
            formula=[f'${u}${AP["situacao"]}="NÃO COTOU"'], fill=fill(REDBG),
            font=Font(bold=True, color=RED)))
        # situação curta na barra fixa
        ws.conditional_formatting.add(f'{u}{R_BAR}', FormulaRule(
            formula=[f'AND({u}${R_BAR}<>"",${u}${AP["situacao"]}="COMPLETO")'],
            fill=fill(OK), font=Font(bold=True, color=WHITE), stopIfTrue=True))
        ws.conditional_formatting.add(f'{u}{R_BAR}', FormulaRule(
            formula=[f'AND({u}${R_BAR}<>"",LEFT(${u}${AP["situacao"]},7)="PARCIAL")'],
            fill=fill(AMBER), font=Font(bold=True, color=WHITE)))
        # menor total à vista / a prazo / total do pedido
        for chave, rng in (('avista', f'{u}{AP["avista"]}:{t}{AP["avista"]}'),
                           ('aprazo', f'{u}{AP["aprazo"]}:{t}{AP["aprazo"]}'),
                           ('totbruto', f'{t}{R_BAR}')):
            alvo = ','.join(ref(chave))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND(${u}${AP[chave]}<>"",${u}${AP[chave]}=MIN({alvo}))'],
                fill=fill(GOLD), font=Font(bold=True, color=NAVY)))

    ws.freeze_panes = f'E{R_IT0}'
    print_cfg(ws, f'A1:{get_column_letter(NCOL)}{rr}', retrato=False)

    # ── dados de exemplo
    if dados:
        for ref_, val in (('A', dados['num']), ('C', dados['demanda']),
                          ('H', dados['solicitante']), ('L', dados['data']),
                          ('O', dados['prazo'])):
            ws[f'{ref_}{R_IDV}'] = val
        for i, (nome, vend, pimp) in enumerate(dados['fornecedores']):
            c0 = TRIOS[i][0]
            ws.cell(R_FORN, c0).value = nome
            ws.cell(R_VEND, c0).value = vend
            if pimp is not None:
                ws.cell(R_BAR, c0 + 1).value = pimp
        for j, it in enumerate(dados['itens']):
            r_ = R_IT0 + j
            ws.cell(r_, 2).value, ws.cell(r_, 3).value, ws.cell(r_, 4).value = it[0], it[1], it[2]
            for i, v in enumerate(it[3]):
                if v is not None:
                    ws.cell(r_, TRIOS[i][0]).value = v
            if it[4]:
                ws.cell(r_, C_OBS).value = it[4]
            for i, pct in (it[5] or {}).items():
                ws.cell(r_, TRIOS[i][0] + 1).value = pct
        for i, apu in enumerate(dados['apuracao']):
            c0 = TRIOS[i][0]
            for chave, val in apu.items():
                ws.cell(AP[chave], c0).value = val


cot = wb.active
cot.title = 'Cotação'
montar_cotacao(cot)

# ══════════════════════════════════════════════════════════════════════════
#  ABA · Pedido de Cotação (A4)
# ══════════════════════════════════════════════════════════════════════════
pc = wb.create_sheet('Pedido de Cotação')
pc.sheet_view.showGridLines = False
NC_PC = 8
W_PC = [5, 42, 7, 8, 14, 8, 15, 22]
r = faixa_marca(pc, NC_PC, 'PEDIDO DE COTAÇÃO',
                'Formulário para enviar ao fornecedor · imprima em A4 ou envie em PDF')
for i, w in enumerate(W_PC, start=1):
    pc.column_dimensions[get_column_letter(i)].width = w
pc.row_dimensions[r].height = 6
r += 1
R_ORIG = r
bloco(pc, r, 1, 3, 'Puxar os itens da aba:', f=Font(name=F, size=8.5, bold=True, color=MUTED),
      bg=WHITE, al=RIGHT, bd=False)
bloco(pc, r, 4, 2, 'Cotação', f=Font(name=F, size=10, bold=True, color=NAVY2),
      bg=INPUT, al=CTR)
bloco(pc, r, 6, 3, 'digite aqui o nome exato da aba de cotação',
      f=Font(name=F, size=8, color=MUTED, i=True), bg=WHITE, al=LEFTI, bd=False)
pc.row_dimensions[r].height = 20
r += 2

SRC = f'"\'"&$D${R_ORIG}&"\'!"'
def puxa(rf): return f'=IFERROR(INDIRECT({SRC}&"{rf}"),"")'

r = titulo_secao(pc, r, NC_PC, 'Identificação')
for rot, rf in (('Cotação nº', f'A{R_IDV}'), ('Demanda / projeto', f'C{R_IDV}'),
                ('Solicitante', f'H{R_IDV}'), ('Data', f'L{R_IDV}'),
                ('Responder até', f'O{R_IDV}')):
    bloco(pc, r, 1, 3, rot + '  ▸', f=Font(name=F, size=9, color=MUTED), bg=WHITE, al=RIGHT)
    c = bloco(pc, r, 4, 5, puxa(rf), f=Font(name=F, size=10, bold=True, color=NAVY2),
              bg=CALC, al=LEFTI)
    if rot in ('Data', 'Responder até'):
        c.number_format = DATA
    pc.row_dimensions[r].height = 18
    r += 1

pc.row_dimensions[r].height = 8
r += 1
r = titulo_secao(pc, r, NC_PC, 'Dados do fornecedor', 'preenchimento do fornecedor')
for rot in ('Empresa', 'CNPJ', 'Vendedor', 'Telefone / WhatsApp', 'E-mail'):
    bloco(pc, r, 1, 3, rot + '  ▸', f=Font(name=F, size=9, color=MUTED), bg=WHITE, al=RIGHT)
    bloco(pc, r, 4, 5, None, f=font(10), bg=INPUT, al=LEFTI)
    pc.row_dimensions[r].height = 18
    r += 1

pc.row_dimensions[r].height = 8
r += 1
r = titulo_secao(pc, r, NC_PC, 'Itens solicitados',
                 'informe o preço unitário e, se houver, o % de imposto que vem por fora')
r = cab_tabela(pc, r, ['#', 'ITEM / ESPECIFICAÇÃO', 'UN.', 'QTD', 'VALOR UNIT. (R$)',
                       '% IMP.', 'VALOR TOTAL (R$)', 'MARCA / OBSERVAÇÃO'], None, 30)
R_PC0 = r
for j in range(NIT):
    rr = R_PC0 + j
    src = R_IT0 + j
    pc.row_dimensions[rr].height = 17
    c = pc.cell(rr, 1, puxa(f'A{src}')); c.font = Font(name=F, size=8.5, color=MUTED)
    c.alignment = CTR; c.fill = fill(CALC); c.border = GRID
    for col, rf, al in ((2, f'B{src}', LEFTI), (3, f'C{src}', CTR), (4, f'D{src}', CTR)):
        c = pc.cell(rr, col, puxa(rf)); c.font = font(9.5); c.alignment = al
        c.fill = fill(CALC); c.border = GRID
        if col == 4: c.number_format = QTD
    c = pc.cell(rr, 5); c.number_format = MOEDA; c.fill = fill(INPUT)
    c.border = GRID; c.font = font(9.5); c.alignment = RIGHT
    c = pc.cell(rr, 6); c.number_format = PCT1; c.fill = fill(INPUT)
    c.border = GRID; c.font = font(8.5, c='7A5B17'); c.alignment = CTR
    c = pc.cell(rr, 7, f'=IF(OR($D{rr}="",$E{rr}=""),"",'
                       f'ROUND($D{rr}*$E{rr}*(1+IF($F{rr}="",0,$F{rr})),2))')
    c.number_format = MOEDA; c.fill = fill(CALC); c.border = GRID
    c.font = font(9.5, c=NAVY2); c.alignment = RIGHT
    c = pc.cell(rr, 8); c.fill = fill(INPUT); c.border = GRID
    c.font = font(8.5); c.alignment = LEFTI
R_PCF = R_PC0 + NIT - 1
r = R_PCF + 1
pc.row_dimensions[r].height = 22
bloco(pc, r, 1, 6, 'SUBTOTAL DOS ITENS (com imposto)',
      f=Font(name=F, size=10, bold=True, color=NAVY), bg=GOLDBG, al=RIGHT)
bloco(pc, r, 7, 1, f'=IF(COUNT($E${R_PC0}:$E${R_PCF})=0,"",'
                   f'ROUND(SUM($G${R_PC0}:$G${R_PCF}),2))',
      f=Font(name=F, size=11, bold=True, color=NAVY), bg=GOLDBG, al=RIGHT, nf=MOEDA)
bloco(pc, r, 8, 1, None, bg=GOLDBG)
r += 2

r = titulo_secao(pc, r, NC_PC, 'Condições comerciais', 'preenchimento do fornecedor')
COND_PC = [('Frete até a Valvic (R$)', MOEDA),
           ('Preço À VISTA — total (R$)', MOEDA),
           ('Condição a prazo oferecida', None),
           ('Preço A PRAZO — total (R$)', MOEDA),
           ('Prazo de entrega (dias corridos)', DIAS),
           ('Validade desta proposta', DATA)]
R_COND0 = r
for rot, nf in COND_PC:
    bloco(pc, r, 1, 4, rot + '  ▸', f=Font(name=F, size=9.5, color=INK), bg=WHITE, al=RIGHT)
    bloco(pc, r, 5, 4, None, f=Font(name=F, size=10, bold=True, color=NAVY2), bg=INPUT,
          al=LEFTI, nf=nf)
    pc.row_dimensions[r].height = 19
    r += 1
dv(pc, "=Listas!$B$2:$B$14", f'E{R_COND0 + 2}:H{R_COND0 + 2}')

pc.row_dimensions[r].height = 8
r += 1
pc.row_dimensions[r].height = 46
bloco(pc, r, 1, NC_PC,
      '  Observações: (1) informe item por item — quando não trabalhar com algum item, deixe o valor em branco e escreva "não temos" na observação. '
      '(2) Se o seu preço for SEM impostos (ex.: "preços s/ IPI = 6,5%"), informe o percentual na coluna % IMP. — assim comparamos todo mundo pelo mesmo critério. '
      '(3) A Valvic compara as propostas pelo valor total entregue, considerando imposto, frete e prazo. (4) Dúvidas: fale com o solicitante indicado acima.',
      f=Font(name=F, size=8.5, color='41505D'), bg=GOLDBG, al=LEFTIW)
r += 2
pc.row_dimensions[r].height = 34
bloco(pc, r, 1, 4, 'Assinatura / carimbo do fornecedor', f=Font(name=F, size=8.5, color=MUTED),
      bg=WHITE, al=CTR)
bloco(pc, r, 5, 4, 'Data', f=Font(name=F, size=8.5, color=MUTED), bg=WHITE, al=CTR)
print_cfg(pc, f'A1:H{r}', retrato=True)

# ══════════════════════════════════════════════════════════════════════════
#  ABA · Exemplo
# ══════════════════════════════════════════════════════════════════════════
EX = {
    'num': 'COT-2026-014',
    'demanda': 'Ferragens — obra Cristiane (closet + cozinha)',
    'solicitante': 'Karla · Administrativo',
    'data': datetime.date(2026, 8, 14),
    'prazo': datetime.date(2026, 8, 18),
    # (nome, vendedor, % imposto padrão)
    'fornecedores': [
        ('Bigfer', 'Marcos · (31) 9xxxx-xxxx', None),
        ('JR Ferragens', 'Andréia · (31) 9xxxx-xxxx', None),
        ('MGV Distribuidora', 'Rafael · (31) 9xxxx-xxxx — "preços s/ IPI = 6,5%"', 0.065),
        ('Ferragens Ipê', 'Sandro · (31) 9xxxx-xxxx', None),
    ],
    # (descrição, un, qtd, [4 unitários], observação, {índice: % imposto da linha})
    'itens': [
        ('Corrediça telescópica 45 cm — soft close', 'par', 24,
         [38.90, 36.50, 41.20, 37.80], '', None),
        ('Dobradiça caneco 35 mm curva — soft close', 'un', 96,
         [7.40, 7.90, 6.95, 7.55], '', None),
        ('Puxador perfil alumínio preto 3 m', 'barra', 12,
         [64.00, 61.50, None, 63.20], 'MGV não trabalha com este perfil', None),
        ('Pistão a gás 100 N', 'un', 8, [22.50, 21.80, 24.00, 22.10], '', None),
        ('Corrediça oculta 50 cm — push', 'par', 6,
         [128.00, 132.00, 125.50, None], '', None),
        ('Suporte prateleira invisível 12 cm', 'un', 40,
         [4.20, 4.55, 3.98, 4.30], 'MGV: item isento de IPI', {2: 0.0}),
        ('Cabideiro extensível 60 cm — alumínio', 'un', 4,
         [96.00, 89.90, 94.50, 92.00], '', None),
        ('Fecho magnético embutir', 'un', 30, [3.10, 3.40, 2.95, 3.25], '', None),
    ],
    'apuracao': [
        {'frete': 180.00, 'pdesc': 0.05, 'condprazo': '30/60', 'pacresc': 0.0,
         'entrega': 7,  'validade': datetime.date(2026, 8, 25)},
        {'frete': 0.00,   'pdesc': 0.03, 'condprazo': '28 dias', 'pacresc': 0.0,
         'entrega': 10, 'validade': datetime.date(2026, 8, 22)},
        {'frete': 240.00, 'pdesc': 0.07, 'condprazo': '30/60/90', 'pacresc': 0.04,
         'entrega': 5,  'validade': datetime.date(2026, 8, 20)},
        {'frete': 120.00, 'pdesc': 0.04, 'condprazo': '3x sem juros', 'pacresc': 0.0,
         'entrega': 12, 'validade': datetime.date(2026, 8, 28)},
    ],
}
ex = wb.create_sheet('Exemplo')
montar_cotacao(ex, EX)

# ══════════════════════════════════════════════════════════════════════════
#  ABA · Mapa de Cotações
# ══════════════════════════════════════════════════════════════════════════
mp = wb.create_sheet('Mapa de Cotações')
mp.sheet_view.showGridLines = False
HDR_MP = ['Nº da cotação', 'Data', 'Demanda / projeto', 'Solicitante', 'Prioridade',
          'Nº de fornecedores', 'Fornecedor escolhido', 'Forma de pagamento',
          'Menor proposta (R$)', 'Maior proposta (R$)', 'Valor fechado (R$)',
          'Economia (R$)', 'Economia (%)', 'Situação', 'Observações']
W_MP = [15, 11, 38, 18, 12, 12, 24, 20, 15, 15, 15, 13, 11, 20, 46]
r = faixa_marca(mp, len(HDR_MP), 'MAPA DE COTAÇÕES',
                'O caderno de bordo das compras · uma linha por cotação · preencha ao fechar a compra')

# explicação
mp.row_dimensions[r].height = 6
r += 1
r = titulo_secao(mp, r, len(HDR_MP), 'Para que serve esta aba')
EXPL = [
    'Cada aba de cotação morre quando a compra é fechada. Esta aba é o que sobra: o histórico de todas elas, uma linha por cotação.',
    'Preencha UMA linha ao fechar cada compra, copiando os números do rodapé da aba de cotação: o menor total à vista, o maior, e quanto você realmente fechou.',
    'Economia (R$) e Economia (%) saem sozinhas — é o que você deixou de gastar por ter cotado em vez de comprar no primeiro fornecedor que respondeu.',
    'Serve para três coisas: mostrar ao Walton que cotar dá dinheiro, descobrir com quem você mais fecha (poder de barganha na próxima negociação) e achar a cotação antiga quando o mesmo item voltar a ser comprado.',
]
for t in EXPL:
    bloco(mp, r, 1, len(HDR_MP), '   ' + t, f=Font(name=F, size=9, color='2A3744'),
          bg=WHITE, al=LEFTIW, bd=False)
    mp.row_dimensions[r].height = 26 if len(t) > 130 else 18
    r += 1
mp.row_dimensions[r].height = 8
r += 1

KPI = [('Cotações registradas', '=COUNTA($A$%d:$A$%d)', '0'),
       ('Em aberto', '=COUNTIF($N$%d:$N$%d,"Em cotação")+COUNTIF($N$%d:$N$%d,"Aguardando fornecedor")+COUNTIF($N$%d:$N$%d,"Em análise")', '0'),
       ('Valor fechado no total', '=IF(COUNT($K$%d:$K$%d)=0,"",SUM($K$%d:$K$%d))', MOEDA0),
       ('Economia acumulada', '=IF(COUNT($L$%d:$L$%d)=0,"",SUM($L$%d:$L$%d))', MOEDA0),
       ('Economia média', '=IF(COUNT($M$%d:$M$%d)=0,"",AVERAGE($M$%d:$M$%d))', PCT1)]
MP0, MPF = r + 4, r + 3 + 80
for i, (rot, fx, nf) in enumerate(KPI):
    c0 = 1 + i * 3
    bloco(mp, r, c0, 3, rot, f=Font(name=F, size=8, bold=True, color=MUTED), bg=WHITE,
          al=CTR, bd=False)
    n = fx.count('%d') // 2
    bloco(mp, r + 1, c0, 3, fx % ((MP0, MPF) * n),
          f=Font(name=F, size=15, bold=True, color=NAVY), bg=GOLDBG, al=CTR, nf=nf)
mp.row_dimensions[r].height = 14
mp.row_dimensions[r + 1].height = 30
mp.row_dimensions[r + 2].height = 8
r = cab_tabela(mp, r + 3, HDR_MP, W_MP)
assert r == MP0, f'MP0 esperado {MP0}, obtido {r}'

for rr in range(MP0, MPF + 1):
    mp.row_dimensions[rr].height = 17
    for col in range(1, len(HDR_MP) + 1):
        c = mp.cell(rr, col); c.border = GRID; c.font = font(9.5)
        c.alignment = LEFTI; c.fill = fill(INPUT)
    for col, nf, al in ((2, DATA, CTR), (5, None, CTR), (6, '0', CTR), (9, MOEDA, RIGHT),
                        (10, MOEDA, RIGHT), (11, MOEDA, RIGHT), (14, None, CTR)):
        c = mp.cell(rr, col); c.alignment = al
        if nf: c.number_format = nf
    c = mp.cell(rr, 12, f'=IF(OR($J{rr}="",$K{rr}=""),"",ROUND($J{rr}-$K{rr},2))')
    c.number_format = MOEDA; c.fill = fill(CALC); c.alignment = RIGHT
    c.font = Font(name=F, size=9.5, bold=True, color=OK); c.border = GRID
    c = mp.cell(rr, 13, f'=IF(OR($L{rr}="",$J{rr}="",$J{rr}=0),"",$L{rr}/$J{rr})')
    c.number_format = PCT1; c.fill = fill(CALC); c.alignment = CTR
    c.font = Font(name=F, size=9.5, color=NAVY2); c.border = GRID

# exemplos semeados
SEED_MP = [
    ('COT-2026-014', datetime.date(2026, 8, 14), 'Ferragens — obra Cristiane (closet + cozinha)',
     'Karla', 'Normal', 4, 'JR Ferragens', 'Boleto a prazo', 3191.71, 3975.75, 3862.93,
     'Em análise', 'É a cotação da aba "Exemplo". Ipê foi o menor (R$ 3.191,71) mas não tinha '
     'a corrediça oculta; fechado com a JR por atender 100% em uma entrega só.'),
    ('COT-2026-013', datetime.date(2026, 8, 6), 'MDF branco TX 18 mm — reposição de estoque',
     'Karla', 'Programada', 3, 'MADEGEM', 'Boleto a prazo', 18420.00, 20150.00, 18420.00,
     'Fechada', 'MADEGEM entregou em 3 dias. Concorrente 2 tinha frete grátis mas 12 dias de prazo.'),
    ('COT-2026-012', datetime.date(2026, 7, 29), 'Puxadores KG398J — obra Maria (Vale dos Cristais)',
     'Jonathan', 'Urgente', 4, 'Bigfer', 'PIX à vista', 6180.00, 7940.00, 5871.00,
     'Fechada', 'Desconto extra de 5% negociado no PIX à vista, fora da cotação. '
     'Economia real ficou acima da tabela.'),
    ('COT-2026-011', datetime.date(2026, 7, 22), 'Fita de borda 22 mm — 6 cores',
     'Deivson', 'Normal', 3, 'MADEGEM', 'Boleto a prazo', 2340.00, 2610.00, 2340.00,
     'Fechada', 'Duas cores só a MADEGEM tinha. Compra única para não abrir dois fretes.'),
    ('COT-2026-010', datetime.date(2026, 7, 15), 'Iluminação LED perfil embutido — obra Rejane',
     'Jonathan', 'Normal', 4, '—', 'A combinar', 4180.00, 5900.00, None,
     'Cancelada', 'Cliente mudou o projeto e tirou a iluminação do escopo. Cotação arquivada.'),
    ('COT-2026-015', datetime.date(2026, 8, 13), 'Serviço de laqueação — 12 portas (terceirizado)',
     'Deivson', 'Urgente', 3, None, None, None, None, None,
     'Aguardando fornecedor', 'Dois responderam, falta o terceiro. Prazo de resposta até 18/08.'),
]
for j, linha in enumerate(SEED_MP):
    rr = MP0 + j
    for col, val in zip((1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 15), linha):
        if val is not None:
            mp.cell(rr, col).value = val

dv(mp, "=Listas!$D$2:$D$4", f'E{MP0}:E{MPF}')
dv(mp, "=Listas!$C$2:$C$8", f'H{MP0}:H{MPF}')
dv(mp, "=Listas!$E$2:$E$7", f'N{MP0}:N{MPF}')
mp.conditional_formatting.add(f'N{MP0}:N{MPF}', FormulaRule(
    formula=[f'$N{MP0}="Fechada"'], fill=fill(OKBG), font=Font(bold=True, color=OK)))
mp.conditional_formatting.add(f'N{MP0}:N{MPF}', FormulaRule(
    formula=[f'OR($N{MP0}="Cancelada",$N{MP0}="Suspensa")'], fill=fill(REDBG),
    font=Font(color=RED)))
mp.conditional_formatting.add(f'A{MP0}:O{MPF}', FormulaRule(
    formula=[f'AND($E{MP0}="Urgente",$N{MP0}<>"Fechada")'], fill=fill('FDF3F2')))
mp.freeze_panes = f'C{MP0}'
print_cfg(mp, f'A1:O{MPF}', retrato=False)

# ══════════════════════════════════════════════════════════════════════════
#  ABA · Listas
# ══════════════════════════════════════════════════════════════════════════
ls = wb.create_sheet('Listas')
ls.sheet_view.showGridLines = False
col_lista(ls, 'A', 'Unidade', UNIDADES, 14)
col_lista(ls, 'B', 'Condição a prazo', CONDICOES, 22)
col_lista(ls, 'C', 'Forma de pagamento', FORMAS, 22)
col_lista(ls, 'D', 'Prioridade', PRIORID, 16)
col_lista(ls, 'E', 'Situação da cotação', SIT_COT, 24)
ls['G1'] = 'Estas listas alimentam os menus suspensos'
ls['G1'].font = font(9, True, WHITE); ls['G1'].fill = fill(NAVY2); ls['G1'].alignment = CTR
ls.column_dimensions['G'].width = 70
for i, txt in enumerate([
    'Pode acrescentar itens no fim de cada coluna — o menu suspenso já cobre linhas de sobra.',
    'Não apague as colunas nem mude a ordem: as validações apontam para posições fixas.',
], start=2):
    ls.cell(i, 7, txt).font = font(9, c='41505D')

# ══════════════════════════════════════════════════════════════════════════
#  ABA · Instruções
# ══════════════════════════════════════════════════════════════════════════
ins = wb.create_sheet('Instruções', 0)
ins.sheet_view.showGridLines = False
NC_IN = 6
r = faixa_marca(ins, NC_IN, 'COMO USAR ESTA PLANILHA',
                'Cotação comparativa de fornecedores · demandas pontuais')
for i, w in enumerate([4, 34, 30, 30, 30, 12], start=1):
    ins.column_dimensions[get_column_letter(i)].width = w
r += 1

def par(row, texto, f=None, h=None, bg=None):
    ins.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC_IN)
    c = ins.cell(row, 1, '   ' + texto)
    c.font = f or font(10, c='2A3744')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if bg:
        for k in range(1, NC_IN + 1):
            ins.cell(row, k).fill = fill(bg)
    ins.row_dimensions[row].height = h or 18
    return row + 1

BLOCOS = [
    ('O QUE ESTA PLANILHA RESOLVE', [
        'Você tem uma demanda pontual — ferragens de uma obra, um lote de chapas, um serviço terceirizado — e precisa comparar até quatro fornecedores sem se perder em conversa de WhatsApp.',
        'Ela faz quatro coisas que a conta de cabeça não faz: mostra quem tem o pedido COMPLETO, põe todo mundo no mesmo critério de imposto, mostra o custo real depois do frete, e separa o preço À VISTA do preço A PRAZO.',
    ]),
    ('O PASSO A PASSO', [
        '1 · Clique com o botão direito na aba "Cotação" → Mover ou copiar → marque "Criar uma cópia". Renomeie a cópia (ex.: COT-2026-015 Chapas).',
        '2 · Na cópia, preencha a identificação (número, demanda, solicitante, data, prazo de resposta).',
        '3 · Liste os itens: descrição, unidade e quantidade. Uma linha por item, com especificação clara — medida, cor, acabamento, marca de referência.',
        '4 · Escreva o nome dos quatro fornecedores na barra fixa e o vendedor com o contato logo abaixo.',
        '5 · Se quiser mandar o pedido formatado, vá na aba "Pedido de Cotação", escreva o nome da sua aba no campo do topo e salve em PDF. Ela puxa os itens sozinha.',
        '6 · Conforme as respostas chegam, digite só o VALOR UNITÁRIO de cada fornecedor. O total sai sozinho e a barra fixa no topo vai somando o pedido.',
        '7 · Preencha por fornecedor: frete, % de desconto à vista, condição e % de acréscimo a prazo, prazo de entrega e validade.',
        '8 · Leia o VEREDITO no rodapé e registre o resultado na aba "Mapa de Cotações".',
    ]),
    ('A REGRA MAIS IMPORTANTE', [
        'Se o fornecedor NÃO tem um item, deixe o preço em branco. Não escreva zero, não escreva traço.',
        'A célula fica avermelhada e o fornecedor passa a contar como PARCIAL — é assim que a planilha sabe quem entrega o pedido inteiro e quem vai te obrigar a abrir uma segunda compra.',
    ]),
    ('IMPOSTO QUE VEM POR FORA (a coluna % IMP.)', [
        'Fornecedor que manda "PREÇOS S/ IPI = 6,5%" está cotando mais barato do que vai cobrar. Comparar o preço dele com o de quem já cotou com imposto é comparar coisa diferente.',
        'Solução: escreva 6,5% no campo % IMPOSTO PADRÃO daquele fornecedor, na barra fixa. Todos os itens dele passam a ser calculados com o acréscimo.',
        'Item com alíquota diferente (isento, ST, importado): digite o % na coluna % IMP. daquela linha. O que está na linha manda; o padrão da barra fixa só vale quando a linha está vazia.',
        'A comparação entre fornecedores é sempre feita pelo TOTAL do item, já com o imposto de cada um — nunca pelo preço de tabela.',
    ]),
    ('A BARRA FIXA DO TOPO', [
        'As linhas até o cabeçalho da tabela ficam congeladas: por mais que você desça na lista de itens, os quatro fornecedores continuam à vista.',
        'Na barra você vê, de cada um: quantos itens ele cotou (ex.: 7/8 — verde quando está completo, âmbar quando falta), o % de imposto padrão, e o TOTAL DO PEDIDO já com imposto e frete, atualizando enquanto você digita.',
        'O menor total do pedido fica com fundo dourado. Desconto à vista e acréscimo a prazo entram só na apuração, mais abaixo.',
    ]),
    ('COMO LER O RESULTADO', [
        'Situação do pedido — COMPLETO (verde) atende tudo; PARCIAL (âmbar) mostra quantos itens faltam; NÃO COTOU (vermelho) não respondeu.',
        'Total à vista e Total a prazo — já contêm imposto, frete, desconto e acréscimo. O menor de cada linha fica com fundo dourado.',
        'Custo do prazo — quanto o parcelamento custa em reais. Compare com o alívio que aquele prazo dá no caixa antes de decidir.',
        'Melhor à vista entre os completos — este é o número que costuma valer a decisão: o mais barato que entrega tudo de uma vez.',
        'Compra fracionada — soma o melhor preço de cada item. É sem frete e implica várias entregas: só vale quando a diferença for grande de verdade.',
        'Empate: quando dois fornecedores dão exatamente o mesmo preço, a planilha aponta o que estiver mais à esquerda. Nesse caso o desempate é seu — prazo de entrega, histórico, relacionamento.',
        'Item sem quantidade, ou que ninguém cotou, fica fora da conta e a planilha avisa na linha da compra fracionada. Se aparecer esse aviso, volte na lista antes de decidir.',
    ]),
    ('O MAPA DE COTAÇÕES — o caderno de bordo', [
        'Cada aba de cotação morre quando a compra é fechada. O Mapa é o que sobra: uma linha por cotação, para sempre.',
        'Você preenche três números vindos do rodapé da cotação — a MENOR proposta recebida, a MAIOR, e quanto realmente FECHOU. A economia sai sozinha: Economia (R$) = maior proposta − valor fechado, e Economia (%) é essa diferença sobre a maior proposta. Ou seja: quanto você deixou de gastar por ter cotado em vez de aceitar a primeira resposta.',
        'Exemplo 1 — a cotação da aba "Exemplo" (COT-2026-014): menor R$ 3.191,71 (Ipê, que não tinha um item) · maior R$ 3.975,75 (Bigfer) · fechado R$ 3.862,93 com a JR, que atendia 100% → economia R$ 112,82 (2,8%). Cotação apertada: os quatro estavam próximos e o ganho veio de escolher quem entregava tudo.',
        'Exemplo 2 — negociação fora da tabela (COT-2026-012, puxadores): menor R$ 6.180,00 · maior R$ 7.940,00 · fechado R$ 5.871,00 depois de arrancar mais 5% no PIX à vista → economia R$ 2.069,00 (26,1%). O Mapa registra o que você conseguiu, não o que a planilha calculou.',
        'Exemplo 3 — cotação que não virou compra (COT-2026-010): menor e maior preenchidos, valor fechado em branco, situação "Cancelada". A economia fica vazia e ela não entra nos totais. Cotação cancelada continua no histórico: da próxima vez que o item voltar, o preço de julho já está ali.',
        'A aba já vem com seis linhas de exemplo — três fechadas, uma em análise, uma cancelada e uma aguardando fornecedor. Apague-as quando começar as suas, ou deixe até pegar o jeito.',
        'Os cinco números do topo se atualizam sozinhos: cotações registradas, quantas estão em aberto, o total efetivamente comprado, a economia acumulada e a economia média.',
        'Serve para três coisas: provar ao Walton que cotar dá dinheiro, descobrir com quem você mais fecha (é o seu poder de barganha na próxima negociação) e reencontrar a cotação antiga quando o mesmo item voltar a ser comprado.',
    ]),
    ('CUIDADOS', [
        'Preço não é a única variável: prazo de entrega atrasado na obra custa mais caro que qualquer desconto.',
        'Confira a validade da proposta antes de emitir o pedido — preço vencido volta diferente.',
        'Compare sempre a mesma especificação. Corrediça soft close de marcas diferentes não é o mesmo item.',
        'Compra fracionada gera um frete por fornecedor. A planilha avisa, mas a conta do frete extra é sua.',
    ]),
]
for titulo, linhas in BLOCOS:
    r = titulo_secao(ins, r, NC_IN, titulo)
    ins.row_dimensions[r].height = 4
    r += 1
    for t in linhas:
        r = par(r, t, h=30 if len(t) > 115 else (18 if len(t) < 80 else 24))
    ins.row_dimensions[r].height = 8
    r += 1

r = par(r, 'Abas desta planilha:   Instruções  ·  Cotação (modelo, duplique)  ·  Pedido de Cotação (formulário A4)  ·  '
           'Exemplo (preenchido)  ·  Mapa de Cotações (registro)  ·  Listas (menus)',
        f=Font(name=F, size=9, bold=True, color='7A5B17'), h=30, bg=GOLDBG)
print_cfg(ins, f'A1:F{r - 1}', retrato=True)

wb.active = 0
SAIDA = '/home/user/valvicorcamentista/painel/planilhas/Valvic_Cotacao_Fornecedores.xlsx'
wb.save(SAIDA)
print('OK →', SAIDA)
print(f'itens {R_IT0}-{R_ITF} | barra fixa {R_BAR} | apuração {R_A0}-{AP["completo"]} '
      f'| veredito {R_V0} | colunas {NCOL}')
