#!/usr/bin/env python3
"""Testes da planilha de Custo por Projeto.

Calcula as fórmulas de verdade e confere a ficha inteira contra um modelo
independente escrito em Python. Para o Painel Geral, resolve o INDIRECT à mão
— o motor não avalia nome de aba dinâmico, mas o endereço apontado por cada
coluna pode e deve ser conferido.

Uso:  python3 testar-custo-projeto.py
"""
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
import formulas

ARQ = 'Valvic_Custo_por_Projeto.xlsx'
EX = 'Exemplo P-2026-041'
FM = 'Ficha Modelo'
PG = 'Painel Geral'

# mapa de linhas da ficha (o TESTE 0 confirma que ainda vale)
R_RES0, R_RESF, R_CUSTO_TOT, R_MC, R_ALERTA = 16, 21, 22, 23, 24
R_VENDA = 28
R_IMP, R_MAQ, R_TRX, R_CVEND, R_PROJ, R_RTP = 32, 33, 34, 35, 36, 37
R_CV_SUB, R_LIQ = 38, 39
R_AMB0, R_AMBF, R_AMB_TOT = 43, 54, 55
R_COORD, R_PRODC, R_MONTC, R_CO_SUB = 59, 60, 61, 62
R_CL0, R_CLF, R_CL_TOT = 66, 77, 78
OC = {'Material': (82, 88, 89), 'Serviços terceirizados': (90, 93, 94),
      'Logística': (95, 99, 100)}
R_RB0, R_RBF, R_RB_SUB = 104, 115, 116
R_LAN0, R_LANF, R_LAN_TOT = 120, 179, 180

# Tolerância de 1 centavo em valores que nascem de ROUND() sobre um produto em
# ponto flutuante — o motor de teste e o Excel podem divergir no último centavo.
CENTAVO = 0.011

falhas, testes = [], 0
def ck(nome, obtido, esperado, tol=0.005):
    global testes
    testes += 1
    if isinstance(esperado, (int, float)) and isinstance(obtido, (int, float)):
        ok = abs(obtido - esperado) <= tol
    else:
        ok = str(obtido).strip() == str(esperado).strip()
    if not ok:
        falhas.append(f'{nome}\n     obtido   = {obtido!r}\n     esperado = {esperado!r}')


def xr(x, n=2):
    q = Decimal(1).scaleb(-n)
    return float(Decimal(f'{float(x):.15g}').quantize(q, rounding=ROUND_HALF_UP))


print('Calculando a planilha com o motor de fórmulas...')
xl = formulas.ExcelModel().loads(ARQ).finish()
sol = xl.calculate()
pat = re.compile(r"^'\[" + re.escape(ARQ) + r"\]([^']+)'!([A-Z]+[0-9]+)$", re.I)
VAL = {}
for k, v in sol.items():
    m = pat.match(str(k))
    if not m:
        continue
    try:
        val = v.value[0, 0]
    except Exception:
        val = getattr(v, 'value', v)
    if hasattr(val, 'item'):
        try: val = val.item()
        except Exception: pass
    VAL[(m.group(1).upper(), m.group(2).upper())] = val

def C(aba, ref):
    return VAL.get((aba.upper(), ref.upper()), '<ausente>')

print(f'{len(VAL)} células resolvidas\n' + '─' * 76)
wbv = openpyxl.load_workbook(ARQ)
ws = wbv[EX]

print('TESTE 0 · o mapa de linhas da ficha é o que o teste assume')
ANCORAS = [(R_VENDA, 1, 'Valor de venda do projeto'), (R_LIQ, 1, '(=) RECEITA LÍQUIDA'),
           (R_CV_SUB, 1, '(=) SUBTOTAL DOS CUSTOS DE VENDA'),
           (R_CO_SUB, 1, '(=) SUBTOTAL DAS COMISSÕES'),
           (R_CUSTO_TOT, 1, '(=) CUSTO TOTAL DO PROJETO'),
           (R_MC, 1, '(=) MARGEM DE CONTRIBUIÇÃO'),
           (89, 1, '(=) SUBTOTAL · MATERIAL'),
           (94, 1, '(=) SUBTOTAL · SERVIÇOS TERCEIRIZADOS'),
           (100, 1, '(=) SUBTOTAL · LOGÍSTICA'),
           (R_CL_TOT, 1, 'TOTAL'), (R_LAN_TOT, 1, 'TOTAL LANÇADO NO LIVRO')]
for r, c, txt in ANCORAS:
    ck(f'  âncora linha {r}', wbv[FM].cell(r, c).value, txt)
ck('  o Painel só lê até a linha 24', R_ALERTA, 24)
print(f'  {testes} verificações')

# ═════════════ modelo independente
V = 90000.0
imp, maq, cvend = xr(0.075 * V), xr(0.02 * V), xr(0.03 * V)
trx_o, trx_r, proj = 120.0, 148.0, 1500.0
rt_o = xr(0.05 * (V - imp - maq - trx_o))
rt_r = xr(0.05 * (V - imp - maq - trx_r))
cv_o = xr(imp + maq + trx_o + cvend + proj + rt_o)
cv_r = xr(imp + maq + trx_r + cvend + proj + rt_r)
liq_o, liq_r = xr(V - cv_o), xr(V - cv_r)
AMB = [('Cozinha', 30000, 'Jackson', 0.03, 'Samuel', 0.02),
       ('Suíte', 20000, 'Samuel', 0.03, 'Cezar', 0.02),
       ('Lavanderia', 10000, 'Joelson', 0.03, 'Samuel', 0.02),
       ('Sala', 30000, 'Deivson', 0.03, 'Jackson', 0.02)]
prod, mont, por_pessoa = {}, {}, {}
tot_prod = tot_mont = 0.0
for nome, val, pq, pp, mq, mp in AMB:
    p = val / V
    vp, vm = xr(p * liq_r * pp), xr(p * liq_r * mp)
    prod[nome], mont[nome] = vp, vm
    tot_prod += vp; tot_mont += vm
    por_pessoa[pq] = por_pessoa.get(pq, 0) + vp
    por_pessoa[mq] = por_pessoa.get(mq, 0) + vm
tot_prod, tot_mont = xr(tot_prod), xr(tot_mont)
coord_o, coord_r = xr(0.01 * liq_o), xr(0.01 * liq_r)
com_o = xr(coord_o + xr(0.03 * liq_o) + xr(0.02 * liq_o))
com_r = xr(coord_r + tot_prod + tot_mont)
por_pessoa['Deivson'] = por_pessoa.get('Deivson', 0) + coord_r

# categorias, a partir do livro de lançamentos lido do arquivo
lanc = []
for r in range(R_LAN0, R_LANF + 1):
    val = ws.cell(r, 5).value
    if isinstance(val, (int, float)):
        lanc.append((ws.cell(r, 2).value, float(val), ws.cell(r, 9).value))
real_cat, acomprar_cat, apagar_cat = {}, {}, {}
for cat, val, st in lanc:
    if st in ('Comprado (a pagar)', 'Pago'):
        real_cat[cat] = real_cat.get(cat, 0) + val
    if st == 'A comprar':
        acomprar_cat[cat] = acomprar_cat.get(cat, 0) + val
    if st == 'Comprado (a pagar)':
        apagar_cat[cat] = apagar_cat.get(cat, 0) + val
grupo_real, grupo_orc = {}, {}
for grupo, (i0, i1, isub) in OC.items():
    grupo_real[grupo] = xr(sum(real_cat.get(ws.cell(r, 1).value, 0)
                               for r in range(i0, i1 + 1)))
    grupo_orc[grupo] = xr(sum(ws.cell(r, 3).value or 0 for r in range(i0, i1 + 1)))
rb_o = 1500.0
rb_r = xr(sum(ws.cell(r, 5).value or 0 for r in range(R_RB0, R_RBF + 1)))
custo_o = xr(cv_o + com_o + sum(grupo_orc.values()) + rb_o)
custo_r = xr(cv_r + com_r + sum(grupo_real.values()) + rb_r)
mc_o, mc_r = xr(V - custo_o), xr(V - custo_r)

n0 = testes
print('\nTESTE 1 · cascata de custos de venda e receita líquida')
for ref, esp in ((f'C{R_IMP}', imp), (f'D{R_IMP}', imp), (f'C{R_MAQ}', maq),
                 (f'D{R_MAQ}', maq), (f'C{R_TRX}', trx_o), (f'D{R_TRX}', trx_r),
                 (f'C{R_CVEND}', cvend), (f'D{R_CVEND}', cvend), (f'C{R_PROJ}', proj),
                 (f'D{R_PROJ}', proj), (f'C{R_RTP}', rt_o), (f'D{R_RTP}', rt_r),
                 (f'C{R_CV_SUB}', cv_o), (f'D{R_CV_SUB}', cv_r),
                 (f'C{R_LIQ}', liq_o), (f'D{R_LIQ}', liq_r)):
    ck(f'  {ref}', C(EX, ref), esp)
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 2 · ambientes — comissão de produção e de montagem')
for i, (nome, val, pq, pp, mq, mp) in enumerate(AMB):
    r = R_AMB0 + i
    ck(f'  {nome} · % do total', C(EX, f'B{r}'), val / V, tol=1e-9)
    ck(f'  {nome} · produção ({pq})', C(EX, f'F{r}'), prod[nome])
    ck(f'  {nome} · montagem ({mq})', C(EX, f'I{r}'), mont[nome])
ck('  soma dos ambientes', C(EX, f'C{R_AMB_TOT}'), V)
ck('  total produção', C(EX, f'F{R_AMB_TOT}'), tot_prod)
ck('  total montagem', C(EX, f'I{R_AMB_TOT}'), tot_mont)
ck('  conferência dos ambientes', str(C(EX, f'D{R_AMB_TOT}'))[:2], 'OK')
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 3 · comissões por colaborador (só quem recebe)')
nomes = [ws.cell(r, 1).value for r in range(R_CL0, R_CLF + 1) if ws.cell(r, 1).value]
ck('  a lista tem 8 nomes e 4 linhas livres', len(nomes), 8)
for proibido in ('Bruna', 'Hugo', 'Karla', 'Filipe'):
    ck(f'  {proibido} não está na lista de comissões', proibido in nomes, False)
for r in range(R_CL0, R_CLF + 1):
    nome = ws.cell(r, 1).value
    if nome:
        ck(f'  {nome} · total', C(EX, f'H{r}'), xr(por_pessoa.get(nome, 0)))
ck('  soma por pessoa', C(EX, f'H{R_CL_TOT}'), com_r, tol=CENTAVO)
ck('  soma por pessoa = subtotal de comissões', C(EX, f'H{R_CL_TOT}'),
   C(EX, f'D{R_CO_SUB}'), tol=1e-9)
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 4 · livro de lançamentos alimenta as categorias')
ck('  lançamentos lidos', len(lanc), 21)
ck('  total do livro', C(EX, f'E{R_LAN_TOT}'), xr(sum(v for _, v, _ in lanc)))
for grupo, (i0, i1, isub) in OC.items():
    for r in range(i0, i1 + 1):
        cat = ws.cell(r, 1).value
        ck(f'  {cat[:28]} · realizado', C(EX, f'D{r}'), xr(real_cat.get(cat, 0)))
        ck(f'  {cat[:28]} · a comprar', C(EX, f'G{r}'), xr(acomprar_cat.get(cat, 0)))
        ck(f'  {cat[:28]} · a pagar', C(EX, f'H{r}'), xr(apagar_cat.get(cat, 0)))
    ck(f'  SUBTOTAL {grupo} · realizado', C(EX, f'D{isub}'), grupo_real[grupo])
    ck(f'  SUBTOTAL {grupo} · orçado', C(EX, f'C{isub}'), grupo_orc[grupo])
soma_cat = xr(sum(grupo_real.values()))
ck('  soma dos 3 grupos = total comprado no livro', soma_cat,
   xr(sum(v for _, v, st in lanc if st in ('Comprado (a pagar)', 'Pago'))))
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 5 · resumo no topo, custo total e margem')
esperados = [(cv_o, cv_r), (com_o, com_r), (grupo_orc['Material'], grupo_real['Material']),
             (grupo_orc['Serviços terceirizados'], grupo_real['Serviços terceirizados']),
             (grupo_orc['Logística'], grupo_real['Logística']), (rb_o, rb_r)]
for i, (co, cr_) in enumerate(esperados):
    r = R_RES0 + i
    ck(f'  resumo linha {r} orçado', C(EX, f'C{r}'), co, tol=CENTAVO)
    ck(f'  resumo linha {r} realizado', C(EX, f'D{r}'), cr_, tol=CENTAVO)
ck('  custo total orçado', C(EX, f'C{R_CUSTO_TOT}'), custo_o, tol=CENTAVO)
ck('  custo total realizado', C(EX, f'D{R_CUSTO_TOT}'), custo_r, tol=CENTAVO)
ck('  MC orçada', C(EX, f'C{R_MC}'), mc_o, tol=CENTAVO)
ck('  MC realizada', C(EX, f'D{R_MC}'), mc_r, tol=CENTAVO)
ck('  custo total = soma das 6 categorias', C(EX, f'D{R_CUSTO_TOT}'),
   xr(sum(C(EX, f'D{R_RES0+i}') for i in range(6))), tol=1e-9)
ck('  MC = venda - custo total', C(EX, f'D{R_MC}'),
   xr(C(EX, 'A13') - C(EX, f'D{R_CUSTO_TOT}')), tol=1e-9)
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 6 · faixa de resultado, alerta de compras e entrega')
ck('  KPI venda', C(EX, 'A13'), V)
ck('  KPI custo total', C(EX, 'C13'), custo_r, tol=CENTAVO)
ck('  KPI MC R$', C(EX, 'E13'), mc_r, tol=CENTAVO)
ck('  KPI MC %', C(EX, 'G13'), mc_r / V, tol=CENTAVO / V)
ck('  KPI MC % orçada', C(EX, 'H13'), mc_o / V, tol=CENTAVO / V)
ck('  KPI desvio de custo', C(EX, 'I13'), xr(custo_r - custo_o), tol=CENTAVO)
ac = xr(sum(v for _, v, st in lanc if st == 'A comprar'))
ap = xr(sum(v for _, v, st in lanc if st == 'Comprado (a pagar)'))
ck(f'  ainda a comprar (numérico, lido pelo Painel)', C(EX, f'I{R_ALERTA}'), ac)
# o separador de milhar do TEXT() segue a máquina; comparamos só os dígitos
alerta = re.sub(r'[.,\s]', '', str(C(EX, f'A{R_ALERTA}')))
ck('  o alerta cita o valor já comprado a pagar', f'{ap:.0f}' in alerta, True)
ck('  o alerta cita o custo projetado', 'custoprojetado' in alerta.lower(), True)
ck('  o alerta cita a MC projetada', 'MCprojetada' in str(C(EX, f'A{R_ALERTA}')).replace(' ', ''), True)
ck('  dias de atraso', C(EX, 'H9'), 7)
ck('  situação da entrega', C(EX, 'J9'), 'Atrasado')
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 7 · a ficha em branco não pode gerar lixo')
for ref in ('A13', 'C13', 'E13', 'G13', 'I13', f'D{R_CV_SUB}', f'D{R_LIQ}',
            f'F{R_AMB0}', f'I{R_AMB0}', f'H{R_CL0}', f'D{R_RES0}', f'D{R_CUSTO_TOT}',
            f'D{R_MC}', 'H9', 'J9', f'D{89}', f'G{89}'):
    v = C(FM, ref)
    ck(f'  {FM}!{ref} vazia', '' if v in ('', 0, 0.0) else v, '')
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 8 · o Painel aponta para os endereços certos da ficha')
pg = wbv[PG]
P0 = next(r for r in range(1, pg.max_row + 1) if pg.cell(r, 2).value == EX)
ESPERADO = {3: 'A7', 4: 'A9', 5: 'F7', 6: 'H7', 7: 'J7', 8: 'H9', 9: 'J9',
            10: 'A13', 11: 'C13', 12: f'I{R_ALERTA}', 13: 'E13', 14: 'G13',
            15: 'H13', 17: 'I13'}
for col, ref in ESPERADO.items():
    f = pg.cell(P0, col).value or ''
    m = re.search(r"&\"'!([A-Z]+[0-9]+)\"\)", f)
    ck(f'  Painel col {openpyxl.utils.get_column_letter(col)}',
       m.group(1) if m else f, ref)
    ck(f'  Painel col {openpyxl.utils.get_column_letter(col)} usa IFERROR+INDIRECT',
       f.startswith('=IFERROR(INDIRECT('), True)
for k in range(6):
    for j, letra in enumerate(('C', 'D')):
        f = pg.cell(P0, 34 + k * 2 + j).value or ''
        m = re.search(r"&\"'!([A-Z]+[0-9]+)\"\)", f)
        ck(f'  Painel auxiliar categoria {k+1}{letra}',
           m.group(1) if m else f, f'{letra}{R_RES0 + k}')
ck('  todos os endereços lidos estão dentro do limite seguro (linha 24)',
   max(int(re.search(r'\d+', v).group()) for v in ESPERADO.values()) <= 24, True)
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 9 · modelo e exemplo têm exatamente a mesma estrutura')
fm = wbv[FM]
dif = sum(1 for r in range(1, 185) for c in range(1, 11)
          if ((isinstance(fm.cell(r, c).value, str) and fm.cell(r, c).value.startswith('='))
              or (isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.startswith('=')))
          and fm.cell(r, c).value != ws.cell(r, c).value)
ck('  fórmulas idênticas', dif, 0)
ck('  mesmo nº de mesclagens', len(fm.merged_cells.ranges), len(ws.merged_cells.ranges))
ck('  12 linhas de ambiente', R_AMBF - R_AMB0 + 1, 12)
ck('  12 linhas de colaborador', R_CLF - R_CL0 + 1, 12)
ck('  12 linhas de retrabalho', R_RBF - R_RB0 + 1, 12)
ck('  60 linhas de lançamento', R_LANF - R_LAN0 + 1, 60)
print(f'  {testes - n0} verificações')

n0 = testes
print('\nTESTE 10 · menus suspensos (Categoria, Forma de pagamento, Status)')
NOMES_ESPERADOS = {'CATEGORIA_COMPRA': 'Listas!$D$2:$D$17',
                   'FORMA_PAGAMENTO': 'Listas!$E$2:$E$9',
                   'STATUS_COMPRA': 'Listas!$F$2:$F$4',
                   'EQUIPE_COMISSAO': 'Listas!$A$2:$A$9',
                   'VENDEDOR': 'Listas!$B$2:$B$6',
                   'CAUSA_RETRABALHO': 'Listas!$C$2:$C$10'}
for nome, alvo in NOMES_ESPERADOS.items():
    d = wbv.defined_names.get(nome)
    ck(f'  nome definido {nome}', d.attr_text if d else '<ausente>', alvo)

VAL_ESPERADAS = {'B120:B179': '=CATEGORIA_COMPRA', 'G120:G179': '=FORMA_PAGAMENTO',
                 'I120:I179': '=STATUS_COMPRA', 'D43:D54': '=EQUIPE_COMISSAO',
                 'G43:G54': '=EQUIPE_COMISSAO', 'A66:A77': '=EQUIPE_COMISSAO',
                 'B104:B115': '=CAUSA_RETRABALHO', 'D7:E7': '=VENDEDOR',
                 'F9:G9': '=EQUIPE_COMISSAO'}
for aba in (FM, EX):
    achadas = {str(d.sqref): d.formula1 for d in wbv[aba].data_validations.dataValidation}
    ck(f'  {aba} · nº de menus', len(achadas), len(VAL_ESPERADAS))
    for rng, formula in VAL_ESPERADAS.items():
        ck(f'  {aba}!{rng} aponta para', achadas.get(rng, '<ausente>'), formula)
    for d in wbv[aba].data_validations.dataValidation:
        ck(f'  {aba}!{d.sqref} usa nome definido, não intervalo de aba',
           d.formula1.lstrip("=").split("!")[0] in NOMES_ESPERADOS, True)
        ck(f'  {aba}!{d.sqref} mostra a setinha na célula', d.showDropDown, False)

# as categorias da ficha têm de existir exatamente na lista da aba Listas
lista = [wbv['Listas'].cell(r, 4).value for r in range(2, 18)]
ck('  a lista tem 16 categorias', len([x for x in lista if x]), 16)
for grupo, (i0, i1, isub) in OC.items():
    for r in range(i0, i1 + 1):
        cat = wbv[FM].cell(r, 1).value
        ck(f'  categoria "{cat}" existe no menu', cat in lista, True)
for cat in lista:
    ck(f'  "{cat}" cabe no menu (sem vírgula, até 26 caracteres)',
       ',' not in cat and len(cat) <= 26, True)
print(f'  {testes - n0} verificações')

print('\n' + '═' * 76)
print(f'{testes} verificações · {len(falhas)} falha(s)')
if falhas:
    print('═' * 76)
    for f in falhas[:25]:
        print('  ✗ ' + f)
    sys.exit(1)
print('TODOS OS TESTES PASSARAM')
