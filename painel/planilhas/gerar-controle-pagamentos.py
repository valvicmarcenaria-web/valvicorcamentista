#!/usr/bin/env python3
"""Gera o Controle de Pagamentos de Projetos da Valvic.

Mantém o layout que a equipe já usa — uma linha por projeto, os pagamentos à
direita, uma aba por ano — e reconstrói tudo limpo: datas corrigidas, saldo e
situação calculados, totais de mês que conferem, painel e notas de migração.

Uso:  python3 gerar-controle-pagamentos.py
Saída: Valvic_Controle_Pagamentos.xlsx
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (Paragraph, ParagraphProperties, CharacterProperties)

import importlib.util as _il, pathlib as _pl
_spec = _il.spec_from_file_location(
    'extrair_pagamentos', _pl.Path(__file__).with_name('extrair-pagamentos.py'))
_mod = _il.module_from_spec(_spec); _spec.loader.exec_module(_mod)
ler = _mod.ler
consolidar = _mod.consolidar
OC = _mod.OC
MESES = _mod.MESES

# ─────────────────────────────────────────────── paleta Valvic
NAVY, NAVY2 = '0E2038', '16314F'
GOLD, GOLDS, GOLDBG = 'C2A05A', 'D8BD80', 'F6EDD6'
INK, MUTED = '1B2733', '6C7785'
LINE, LINE2 = 'E8E3D8', 'DFDACD'
OK, BLUE, RED, AMBER = '2F7D4F', '2F5D8C', 'B0413F', 'B57A16'
OKBG, REDBG, AMBBG, BLUEBG = 'E2F0E7', 'FBE7E6', 'FBF1DC', 'E7EEF6'
INPUT, WHITE, CALC = 'FFF9E3', 'FFFFFF', 'F4F6F8'

F = 'Arial'
def font(sz=10, b=False, c=INK, i=False):
    return Font(name=F, size=sz, bold=b, color=c, italic=i)
def fill(c): return PatternFill('solid', fgColor=c)
def side(c=LINE, st='thin'): return Side(style=st, color=c)
GRID = Border(bottom=side(LINE), left=side(LINE), right=side(LINE))
BOTTOM2 = Border(bottom=side(LINE2))
CTR = Alignment(horizontal='center', vertical='center')
CTRW = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
LEFTI = Alignment(horizontal='left', vertical='center', indent=1)
LEFTIW = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
RIGHT = Alignment(horizontal='right', vertical='center', indent=1)
MOEDA, MOEDA0, PCT0, PCT1, DATA = 'R$ #,##0.00', 'R$ #,##0', '0%', '0.0%', 'DD/MM/YYYY'

ANOS = ['2026', '2025', '2024', '2023']
ANO_ATUAL = '2026'
META = {'2026': 3000000, '2025': 2000000, '2024': 0, '2023': 0}
ANTERIOR = 'Anterior (2022)'
FORMAS = ['pix', 'cartão', 'dinheiro', 'espécie', 'boleto', 'transferência', 'ted',
          'cheque', 'permuta', 'material', 'depósito', 'pix + cartão']

# ── layout da aba de ano
C_CLI, C_PROJ, C_VAL, C_FORMA, C_REC, C_SALDO, C_PCT, C_SIT = 1, 2, 3, 4, 5, 6, 7, 8
BLOCOS = [9, 12, 15, 18, 21]              # I, L, O, R, U
C_AUX = 24                                 # X · auxiliar oculta: saldo por projeto
C_AUX2 = 25                                # Y · auxiliar oculta: contrato por projeto
NCOL = 25
R_PAINEL, R_HEAD, R_IT0 = 6, 9, 10
LP = {c: get_column_letter(c) for c in range(1, NCOL + 1)}

wb = openpyxl.Workbook()
wb.remove(wb.active)                       # descarta a aba vazia padrão


# ══════════════════════════════════════════════ helpers
def faixa_marca(ws, ncols, titulo, sub, linha=1):
    for i, (txt, fo, bg, alt) in enumerate([
            ('VALVIC MARCENARIA', Font(name=F, size=13, bold=True, color=WHITE), NAVY, 26),
            ('Vargas Decor Ltda   ·   CNPJ 17.269.304/0001-51   ·   Belo Horizonte / MG',
             Font(name=F, size=8, color='9FB0C4'), NAVY, 15),
            (titulo, Font(name=F, size=15, bold=True, color=WHITE), NAVY2, 28),
            (sub, Font(name=F, size=8.5, color='7A5B17', italic=True), GOLDBG, 17)]):
        r = linha + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1, txt); c.font = fo; c.alignment = LEFTI
        ws.row_dimensions[r].height = alt
        for cc in range(1, ncols + 1):
            ws.cell(r, cc).fill = fill(bg)
    return linha + 4


def titulo_secao(ws, row, ncols, texto, nota=''):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, ('  ' + texto.upper()) + (f'          {nota}' if nota else ''))
    c.font = Font(name=F, size=9, bold=True, color=NAVY); c.alignment = LEFT
    for cc in range(1, ncols + 1):
        cel = ws.cell(row, cc)
        cel.fill = fill(GOLDBG); cel.border = Border(bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = 20
    return row + 1


def cab_tabela(ws, row, headers, widths=None, alt=30):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(name=F, size=8.5, bold=True, color=WHITE); c.fill = fill(NAVY2)
        c.alignment = CTRW
        c.border = Border(left=side(NAVY2), right=side(NAVY2), bottom=side(GOLD, 'medium'))
    ws.row_dimensions[row].height = alt
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1


def print_cfg(ws, area, retrato=False, margens=(0.4, 0.3, 0.4, 0.3)):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait' if retrato else 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.print_area = area
    l, r, t, b = margens
    ws.page_margins.left, ws.page_margins.right = l, r
    ws.page_margins.top, ws.page_margins.bottom = t, b
    ws.sheet_view.showGridLines = False


def dv(ws, formula, rng):
    d = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(d); d.add(rng)


def bloco(ws, row, col, span, valor=None, *, f=None, bg=None, al=None, nf=None, bd=True):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row, col, valor)
    if f: c.font = f
    if al: c.alignment = al
    if nf: c.number_format = nf
    for k in range(span):
        cel = ws.cell(row, col + k)
        if bd: cel.border = GRID
        if bg: cel.fill = fill(bg)
    return c


# ══════════════════════════════════════════════ abas de ano
HDR = ['Cliente', 'Projeto', 'Investimento', 'Forma de pagamento', 'Recebido',
       'Saldo a receber', '% recebido', 'Situação']
for i in range(1, 6):
    HDR += [f'{i}º Pagt. (R$)', 'Data', 'Descrição']
HDR += ['aux', 'aux2']
W = [22, 28, 13, 26, 13, 14, 10, 16] + [13, 11, 15] * 5 + [3, 3]

F_ROT = Font(name=F, size=9, color=INK)
F_NUM = Font(name=F, size=9.5, color=NAVY2)
F_ENT = Font(name=F, size=9.5, color=INK)
F_DIV = Font(name=F, size=10, bold=True, color=WHITE)

dados = ler()
mapa_anos = {}


def montar_ano(aba, blocos_dados, meta, sobras):
    ws = wb.create_sheet(aba)
    ws.sheet_view.showGridLines = False
    r = faixa_marca(ws, NCOL, f'CONTROLE DE PAGAMENTOS · {aba}',
                    'Uma linha por projeto · fundo creme = você preenche · '
                    'fundo cinza = calculado · Recebido, Saldo, % e Situação saem sozinhos')
    assert r == R_PAINEL - 1
    ws.row_dimensions[r].height = 6

    # ── painel do ano (rótulos em 6, valores em 7)
    PAINEL = [(1, 3, 'VENDIDO NO ANO', MOEDA0), (4, 2, 'RECEBIDO', MOEDA0),
              (6, 2, 'A RECEBER', MOEDA0), (8, 1, '% RECEBIDO', PCT0),
              (9, 3, 'META DO ANO', MOEDA0), (12, 3, '% DA META', PCT1),
              (15, 3, 'FALTA PARA A META', MOEDA0), (18, 6, 'CONFERÊNCIA', None)]
    for col, span, rot, nf in PAINEL:
        bloco(ws, R_PAINEL, col, span, rot, f=Font(name=F, size=8, bold=True, color=MUTED),
              bg=WHITE, al=CTR, bd=False)
        bloco(ws, R_PAINEL + 1, col, span, None,
              f=Font(name=F, size=13, bold=True, color=NAVY), bg=GOLDBG, al=CTR, nf=nf)
    ws.row_dimensions[R_PAINEL].height = 14
    ws.row_dimensions[R_PAINEL + 1].height = 30
    ws.row_dimensions[R_PAINEL + 2].height = 8

    # ── cabeçalho e linhas
    cab_tabela(ws, R_HEAD, HDR, W)
    r = R_IT0
    divisores, primeira_mes, ultima_mes = [], None, None
    for b in blocos_dados:
        nome = b['mes']
        d = r
        divisores.append((nome, d))
        if nome != ANTERIOR:
            primeira_mes = d if primeira_mes is None else primeira_mes
        ws.row_dimensions[d].height = 20
        i0 = d + 1
        i1 = d + max(1, len(b['projetos']) + sobras)
        # linha divisora
        bloco(ws, d, C_CLI, 2, nome, f=F_DIV, bg=NAVY2, al=LEFTI)
        for col, val, nf, fo in (
                (C_VAL, f'=SUM({LP[C_VAL]}{i0}:{LP[C_VAL]}{i1})', MOEDA0, F_DIV),
                (C_REC, f'=SUM({LP[C_REC]}{i0}:{LP[C_REC]}{i1})', MOEDA0, F_DIV),
                (C_SALDO, f'=SUM({LP[C_SALDO]}{i0}:{LP[C_SALDO]}{i1})', MOEDA0, F_DIV),
                (C_PCT, f'=IF(OR({LP[C_VAL]}{d}="",{LP[C_VAL]}{d}=0),"",'
                        f'{LP[C_REC]}{d}/{LP[C_VAL]}{d})', PCT0, F_DIV),
                (C_SIT, f'=COUNT({LP[C_VAL]}{i0}:{LP[C_VAL]}{i1})&" projeto(s)"', None,
                 Font(name=F, size=9, color=GOLDS))):
            c = ws.cell(d, col, val); c.font = fo; c.fill = fill(NAVY2)
            c.alignment = CTR; c.border = GRID
            if nf: c.number_format = nf
        c = ws.cell(d, C_FORMA); c.fill = fill(NAVY2); c.border = GRID
        for col in range(BLOCOS[0], NCOL + 1):
            cc = ws.cell(d, col); cc.fill = fill(NAVY2); cc.border = GRID

        # linhas de projeto
        for k in range(i1 - i0 + 1):
            rr = i0 + k
            p = b['projetos'][k] if k < len(b['projetos']) else None
            ws.row_dimensions[rr].height = 17
            for col, larg in ((C_CLI, None), (C_PROJ, None), (C_FORMA, None)):
                c = ws.cell(rr, col); c.font = F_ENT; c.fill = fill(INPUT)
                c.alignment = LEFTI; c.border = GRID
            c = ws.cell(rr, C_VAL); c.font = Font(name=F, size=9.5, bold=True, color=NAVY)
            c.fill = fill(INPUT); c.alignment = RIGHT; c.border = GRID
            c.number_format = MOEDA
            REC = (f'=IF($C{rr}="","",ROUND(SUM(' +
                   ','.join(f'{LP[c0]}{rr}' for c0 in BLOCOS) + '),2))')
            for col, val, nf, fo, al in (
                    (C_REC, REC, MOEDA, F_NUM, RIGHT),
                    (C_SALDO, f'=IF($C{rr}="","",ROUND($C{rr}-$E{rr},2))', MOEDA,
                     Font(name=F, size=9.5, bold=True, color=NAVY2), RIGHT),
                    (C_PCT, f'=IF(OR($C{rr}="",$C{rr}=0),"",$E{rr}/$C{rr})', PCT0, F_NUM, CTR),
                    (C_SIT, f'=IF($C{rr}="","",IF($E{rr}=0,"A receber",'
                            f'IF(ROUND($F{rr},2)=0,"Quitado",'
                            f'IF($F{rr}<0,"Recebido a mais","Parcial"))))', None,
                     Font(name=F, size=8.5, bold=True, color=NAVY2), CTR)):
                c = ws.cell(rr, col, val); c.font = fo; c.fill = fill(CALC)
                c.alignment = al; c.border = GRID
                if nf: c.number_format = nf
            for bi, c0 in enumerate(BLOCOS):
                pg = p['pagamentos'][bi] if (p and bi < len(p['pagamentos'])) else None
                c = ws.cell(rr, c0, pg['valor'] if pg else None)
                c.number_format = MOEDA; c.font = F_NUM; c.fill = fill(INPUT)
                c.alignment = RIGHT; c.border = GRID
                c = ws.cell(rr, c0 + 1, pg['data'] if pg else None)
                c.number_format = DATA; c.font = font(9); c.fill = fill(INPUT)
                c.alignment = CTR; c.border = GRID
                c = ws.cell(rr, c0 + 2, (pg['desc'] or None) if pg else None)
                c.font = font(8.5); c.fill = fill(INPUT); c.alignment = LEFTI
                c.border = GRID
            # auxiliar: repete o saldo só nas linhas de projeto, para o ranking
            c = ws.cell(rr, C_AUX, f'=IF($C{rr}="","",$F{rr})')
            c.number_format = MOEDA; c.font = font(8, c=MUTED)
            c = ws.cell(rr, C_AUX2, f'=IF($C{rr}="","",$C{rr})')
            c.number_format = MOEDA; c.font = font(8, c=MUTED)
            if p:
                ws.cell(rr, C_CLI).value = p['cliente'] or None
                ws.cell(rr, C_PROJ).value = p['projeto'] or None
                ws.cell(rr, C_VAL).value = p['valor']
                ws.cell(rr, C_FORMA).value = p['forma'] or None
        r = i1 + 1
        if nome != ANTERIOR:
            ultima_mes = i1
        ws.row_dimensions[r].height = 6
        r += 1
    fim = r - 1

    # ── fórmulas do painel do ano
    ini = primeira_mes if primeira_mes else R_IT0
    RG = lambda col: f'${LP[col]}${ini}:${LP[col]}${ultima_mes}'
    NAODIV = f'(COUNTIF(Listas!$A$2:$A$15,$A${ini}:$A${ultima_mes})=0)'
    # Vendido: SUMPRODUCT sobre as linhas que não são divisoras (coluna C é numérica).
    # Recebido: SUM direto das 5 colunas de pagamento — elas não têm texto, e as linhas
    # divisoras ficam vazias nelas, então não há risco de contar duas vezes.
    PAGS = ','.join(f'${LP[c0]}${ini}:${LP[c0]}${ultima_mes}' for c0 in BLOCOS)
    ws.cell(R_PAINEL + 1, 1).value = f'=ROUND(SUMPRODUCT({NAODIV}*{RG(C_VAL)}),2)'
    ws.cell(R_PAINEL + 1, 4).value = f'=ROUND(SUM({PAGS}),2)'
    ws.cell(R_PAINEL + 1, 6).value = (f'=ROUND($A${R_PAINEL+1}-$D${R_PAINEL+1},2)')
    ws.cell(R_PAINEL + 1, 8).value = f'=IF($A${R_PAINEL+1}=0,"",$D${R_PAINEL+1}/$A${R_PAINEL+1})'
    ws.cell(R_PAINEL + 1, 9).value = meta or None
    ws.cell(R_PAINEL + 1, 9).fill = fill(INPUT)
    ws.cell(R_PAINEL + 1, 12).value = (f'=IF(OR($I${R_PAINEL+1}="",$I${R_PAINEL+1}=0),"",'
                                       f'$A${R_PAINEL+1}/$I${R_PAINEL+1})')
    ws.cell(R_PAINEL + 1, 15).value = (f'=IF(OR($I${R_PAINEL+1}="",$I${R_PAINEL+1}=0),"",'
                                       f'MAX(0,$I${R_PAINEL+1}-$A${R_PAINEL+1}))')
    somadiv = '+'.join(f'${LP[C_VAL]}${d}' for nome, d in divisores if nome != ANTERIOR)
    ws.cell(R_PAINEL + 1, 18).value = (
        f'=IF(ROUND(({somadiv})-$A${R_PAINEL+1},2)=0,"OK · os meses somam igual às linhas",'
        f'"DIVERGE em "&TEXT(({somadiv})-$A${R_PAINEL+1},"R$ #,##0.00")&'
        f' " — confira se alguma linha ficou fora de um bloco de mês")')
    ws.cell(R_PAINEL + 1, 18).font = Font(name=F, size=9, bold=True, color=NAVY)
    ws.cell(R_PAINEL + 1, 18).alignment = CTRW

    # ── validação e formatação condicional
    dv(ws, '=Listas!$B$2:$B$13', f'{LP[C_FORMA]}{R_IT0}:{LP[C_FORMA]}{fim}')
    for c0 in BLOCOS:
        dv(ws, '=Listas!$B$2:$B$13', f'{LP[c0+2]}{R_IT0}:{LP[c0+2]}{fim}')
    sit = f'{LP[C_SIT]}{R_IT0}:{LP[C_SIT]}{fim}'
    for txt, bg, cor in (('Quitado', OKBG, OK), ('Parcial', AMBBG, AMBER),
                         ('A receber', REDBG, RED), ('Recebido a mais', BLUEBG, BLUE)):
        ws.conditional_formatting.add(sit, FormulaRule(
            formula=[f'${LP[C_SIT]}{R_IT0}="{txt}"'], fill=fill(bg),
            font=Font(bold=True, color=cor), stopIfTrue=True))
    ws.conditional_formatting.add(f'{LP[C_SALDO]}{R_IT0}:{LP[C_SALDO]}{fim}', FormulaRule(
        formula=[f'AND(${LP[C_SALDO]}{R_IT0}<>"",${LP[C_SALDO]}{R_IT0}>0)'],
        font=Font(bold=True, color=RED)))
    ws.conditional_formatting.add(f'{LP[C_SALDO]}{R_IT0}:{LP[C_SALDO]}{fim}', FormulaRule(
        formula=[f'AND(${LP[C_SALDO]}{R_IT0}<>"",${LP[C_SALDO]}{R_IT0}<0)'],
        fill=fill(BLUEBG), font=Font(bold=True, color=BLUE)))
    ws.conditional_formatting.add(f'{LP[C_CLI]}{R_IT0}:{LP[C_FORMA]}{fim}', FormulaRule(
        formula=[f'AND($C{R_IT0}<>"",$H{R_IT0}="A receber")'], fill=fill('FDF3F2')))

    ws.column_dimensions[LP[C_AUX]].hidden = True
    ws.column_dimensions[LP[C_AUX2]].hidden = True
    ws.freeze_panes = f'{LP[BLOCOS[0]]}{R_IT0}'
    print_cfg(ws, f'A1:{LP[C_SIT]}{fim}')
    mapa_anos[aba] = dict(ini=ini, fim=ultima_mes, divisores=divisores, fim_geral=fim)
    return ws


for aba in ANOS:
    blocos = dados[aba]
    por_mes = {b['mes']: b for b in blocos}
    ordenados = []
    if 'Anterior' in por_mes:
        b = por_mes.pop('Anterior'); b['mes'] = ANTERIOR
        ordenados.append(b)
    for m in MESES:
        ordenados.append(por_mes.get(m, {'mes': m, 'projetos': []}))
    for b in ordenados:
        for p in b['projetos']:
            consolidar(p, 5)
    montar_ano(aba, ordenados, META[aba], sobras=3 if aba == ANO_ATUAL else 2)


# ══════════════════════════════════════════════ dados de apoio ao painel
DIVISOR = {a: dict(mapa_anos[a]['divisores']) for a in ANOS}
ANOS_CHEIOS = ['2023', '2024', '2025']        # anos fechados, base da sazonalidade

def rgn(aba, col, ini=None, fim=None):
    m = mapa_anos[aba]
    return (f"'{aba}'!${col}${ini or m['ini']}:${col}${fim or m['fim_geral']}")

def fluxo_mes(ano, mes):
    """Recebimentos com data de pagamento dentro do mês, somando todas as abas.

    Aqui a faixa começa em R_IT0, e não no primeiro bloco de mês: o dinheiro que
    entrou por contratos de 2022 (bloco "Anterior" da aba 2023) é caixa de verdade
    e precisa aparecer no fluxo, mesmo ficando fora do total de vendas do ano.
    """
    partes = []
    for aba in ANOS:
        for c0 in BLOCOS:
            v, d = LP[c0], LP[c0 + 1]
            rv, rd = rgn(aba, v, R_IT0), rgn(aba, d, R_IT0)
            partes.append(
                f'SUMIFS({rv},{rd},">="&DATE({ano},{mes},1),'
                f'{rd},"<"&DATE({ano},{mes + 1},1))')
    return '=ROUND(' + '+'.join(partes) + ',2)'

def qtd_projetos(aba):
    return f'=COUNT({rgn(aba, LP[C_AUX])})'


# ══════════════════════════════════════════════ ABA · Painel
pa = wb.create_sheet('Painel', 0)
pa.sheet_view.showGridLines = False
NC_PA = 20
r = faixa_marca(pa, NC_PA, 'PAINEL DE PAGAMENTOS',
                'Tudo aqui é calculado a partir das abas de ano — não digite nada nesta aba')
for i, w in enumerate([20, 26, 13, 13, 13, 11, 11] + [11] * 13, start=1):
    pa.column_dimensions[get_column_letter(i)].width = w
pa.row_dimensions[r].height = 6
r += 1

A = mapa_anos[ANO_ATUAL]
KPI = [('VENDIDO EM ' + ANO_ATUAL, f"='{ANO_ATUAL}'!$A$7", MOEDA0),
       ('RECEBIDO', f"='{ANO_ATUAL}'!$D$7", MOEDA0),
       ('A RECEBER', f"='{ANO_ATUAL}'!$F$7", MOEDA0),
       ('% RECEBIDO', f"='{ANO_ATUAL}'!$H$7", PCT0),
       ('PROJETOS', qtd_projetos(ANO_ATUAL), '0'),
       ('TICKET MÉDIO', f"=IF({qtd_projetos(ANO_ATUAL)[1:]}=0,\"\","
                        f"'{ANO_ATUAL}'!$A$7/{qtd_projetos(ANO_ATUAL)[1:]})", MOEDA0)]
R_KPI = r
for i, (rot, fx, nf) in enumerate(KPI):
    c0 = 1 + i * 3
    bloco(pa, r, c0, 3, rot, f=Font(name=F, size=8, bold=True, color=MUTED), bg=WHITE,
          al=CTR, bd=False)
    bloco(pa, r + 1, c0, 3, fx, f=Font(name=F, size=15, bold=True, color=NAVY),
          bg=GOLDBG, al=CTR, nf=nf)
pa.row_dimensions[r].height = 14
pa.row_dimensions[r + 1].height = 32
pa.row_dimensions[r + 2].height = 8
r += 3

# ── ano a ano
r = titulo_secao(pa, r, NC_PA, 'Ano a ano', 'vendido, recebido e o que ficou em aberto')
R_ANO_H = r
r = cab_tabela(pa, r, ['Ano', 'Projetos', 'Vendido', 'Recebido', 'A receber',
                       '% recebido', 'Ticket médio'] + [''] * 13, None, 26)
R_ANO0 = r
for aba in ANOS:
    pa.row_dimensions[r].height = 19
    vals = [(1, aba, None, LEFTI), (2, qtd_projetos(aba), '0', CTR),
            (3, f"='{aba}'!$A$7", MOEDA0, RIGHT), (4, f"='{aba}'!$D$7", MOEDA0, RIGHT),
            (5, f"='{aba}'!$F$7", MOEDA0, RIGHT), (6, f"='{aba}'!$H$7", PCT0, CTR),
            (7, f"=IF({qtd_projetos(aba)[1:]}=0,\"\",'{aba}'!$A$7/{qtd_projetos(aba)[1:]})",
             MOEDA0, RIGHT)]
    for col, v, nf, al in vals:
        c = pa.cell(r, col, v)
        c.font = Font(name=F, size=10, bold=(col == 1), color=NAVY if col == 1 else NAVY2)
        c.fill = fill(CALC); c.alignment = al; c.border = GRID
        if nf: c.number_format = nf
    r += 1
R_ANOF = r - 1
pa.row_dimensions[r].height = 8
r += 1

# ── mês a mês do ano corrente
r = titulo_secao(pa, r, NC_PA, f'{ANO_ATUAL} mês a mês',
                 'vendas pela data do contrato · acumulado comparado à meta')
R_MES_H = r
r = cab_tabela(pa, r, ['Mês', 'Projetos', 'Vendido', 'Recebido', 'A receber',
                       '% recebido', 'Acum. vendido', 'Meta acum.'] + [''] * 12, None, 26)
R_MES0 = r
for k, mes in enumerate(MESES, start=1):
    d = DIVISOR[ANO_ATUAL][mes]
    pa.row_dimensions[r].height = 18
    ac = (f'=IF(DATE({ANO_ATUAL},{k},1)>TODAY(),"",'
          f'IF(SUM($C${R_MES0}:$C{r})=0,"",SUM($C${R_MES0}:$C{r})))')
    for col, v, nf, al in ((1, mes, None, LEFTI),
                           (2, f"=COUNT('{ANO_ATUAL}'!$X${d + 1}:$X${d + 9})", '0', CTR),
                           (3, f"='{ANO_ATUAL}'!$C${d}", MOEDA0, RIGHT),
                           (4, f"='{ANO_ATUAL}'!$E${d}", MOEDA0, RIGHT),
                           (5, f"='{ANO_ATUAL}'!$F${d}", MOEDA0, RIGHT),
                           (6, f"='{ANO_ATUAL}'!$G${d}", PCT0, CTR),
                           (7, ac, MOEDA0, RIGHT),
                           (8, f"=ROUND('{ANO_ATUAL}'!$I$7/12*{k},0)", MOEDA0, RIGHT)):
        c = pa.cell(r, col, v)
        c.font = Font(name=F, size=9.5, bold=(col == 1),
                      color=NAVY if col == 1 else (MUTED if col == 8 else NAVY2))
        c.fill = fill(CALC); c.alignment = al; c.border = GRID
        if nf: c.number_format = nf
    r += 1
R_MESF = r - 1
pa.conditional_formatting.add(f'G{R_MES0}:G{R_MESF}', FormulaRule(
    formula=[f'AND($G{R_MES0}<>"",$G{R_MES0}>=$H{R_MES0})'],
    fill=fill(OKBG), font=Font(bold=True, color=OK)))
pa.conditional_formatting.add(f'G{R_MES0}:G{R_MESF}', FormulaRule(
    formula=[f'AND($G{R_MES0}<>"",$G{R_MES0}<$H{R_MES0})'], font=Font(color=RED)))
pa.row_dimensions[r].height = 8
r += 1

# ── maiores saldos
r = titulo_secao(pa, r, NC_PA, f'Maiores saldos em aberto de {ANO_ATUAL}',
                 'ordenado pelo saldo · empate exato mostra o primeiro da lista')
r = cab_tabela(pa, r, ['#', 'Cliente', 'Projeto', 'Investimento', 'Recebido',
                       'Saldo a receber', '% recebido', '% do total em aberto'] + [''] * 12,
               None, 26)
R_TOP0 = r
RS = f"'{ANO_ATUAL}'!$X${A['ini']}:$X${A['fim']}"
for k in range(1, 13):
    pa.row_dimensions[r].height = 18
    L = f'LARGE({RS},{k})'
    MT = f'MATCH({L},{RS},0)'
    def idx(col):
        return (f'=IF(COUNT({RS})<{k},"",IF({L}<=0,"",'
                f"INDEX('{ANO_ATUAL}'!${col}${A['ini']}:${col}${A['fim']},{MT})))")
    for col, v, nf, al in ((1, k, None, CTR), (2, idx('A'), None, LEFTI),
                           (3, idx('B'), None, LEFTI), (4, idx('C'), MOEDA0, RIGHT),
                           (5, idx('E'), MOEDA0, RIGHT),
                           (6, f'=IF(COUNT({RS})<{k},"",IF({L}<=0,"",{L}))', MOEDA0, RIGHT),
                           (7, idx('G'), PCT0, CTR),
                           (8, f"=IF($F{r}=\"\",\"\",$F{r}/'{ANO_ATUAL}'!$F$7)", PCT1, CTR)):
        c = pa.cell(r, col, v)
        c.font = Font(name=F, size=9.5, bold=(col == 6),
                      color=RED if col == 6 else (MUTED if col in (1, 8) else NAVY2))
        c.fill = fill(CALC); c.alignment = al; c.border = GRID
        if nf: c.number_format = nf
    r += 1
R_TOPF = r - 1
pa.row_dimensions[r].height = 8
r += 1

# ── leitura automática
r = titulo_secao(pa, r, NC_PA, 'Leitura do quadro', 'frases montadas a partir dos números acima')
LEITURAS = [
    (f'="Concentração da carteira em aberto: os 3 maiores saldos somam "&'
     f'TEXT(SUM($F${R_TOP0}:$F${R_TOP0 + 2}),"R$ #,##0")&" de "&'
     f'TEXT(\'{ANO_ATUAL}\'!$F$7,"R$ #,##0")&" — "&'
     f'TEXT(SUM($F${R_TOP0}:$F${R_TOP0 + 2})/\'{ANO_ATUAL}\'!$F$7,"0%")&" do total a receber."'),
    (f'="Ritmo do ano até "&INDEX($A${R_MES0}:$A${R_MESF},MONTH(TODAY()))&": vendido "&'
     f'TEXT(\'{ANO_ATUAL}\'!$A$7,"R$ #,##0")&" contra meta acumulada de "&'
     f'TEXT(INDEX($H${R_MES0}:$H${R_MESF},MONTH(TODAY())),"R$ #,##0")&" — "&'
     f'IF(\'{ANO_ATUAL}\'!$A$7>=INDEX($H${R_MES0}:$H${R_MESF},MONTH(TODAY())),'
     f'"acima do ritmo necessário.","faltam "&'
     f'TEXT(INDEX($H${R_MES0}:$H${R_MESF},MONTH(TODAY()))-\'{ANO_ATUAL}\'!$A$7,"R$ #,##0")&'
     f'" para o ritmo da meta.")'),
    (f'="Projetos sem nenhum recebimento em {ANO_ATUAL}: "&'
     f'COUNTIF(\'{ANO_ATUAL}\'!$H${A["ini"]}:$H${A["fim"]},"A receber")&'
     f'" de "&{qtd_projetos(ANO_ATUAL)[1:]}&" — somam "&'
     f'TEXT(SUMIF(\'{ANO_ATUAL}\'!$H${A["ini"]}:$H${A["fim"]},"A receber",'
     f'\'{ANO_ATUAL}\'!$C${A["ini"]}:$C${A["fim"]}),"R$ #,##0")&" de contrato."'),
    ('="Veja a aba Análise & Sazonalidade para o fluxo de caixa mês a mês e o padrão '
     'sazonal de vendas."'),
]
for txt in LEITURAS:
    pa.row_dimensions[r].height = 22
    bloco(pa, r, 1, NC_PA, txt, f=Font(name=F, size=9.5, color='2A3744'), bg=WHITE,
          al=LEFTIW, bd=False)
    r += 1
R_FIM_PA = r

# ── gráficos
def estilo_eixo(ch, titulo=None):
    ch.style = None
    ch.y_axis.majorGridlines = ChartLines()
    ch.y_axis.numFmt = 'R$ #,##0'
    ch.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(
        defRPr=CharacterProperties(sz=800, solidFill=MUTED)), endParaRPr=CharacterProperties(sz=800))])
    ch.x_axis.txPr = ch.y_axis.txPr
    ch.legend.position = 'b'
    ch.legend.overlay = False
    if titulo:
        ch.title = titulo

def pinta(serie, cor):
    serie.graphicalProperties.solidFill = cor
    serie.graphicalProperties.line.solidFill = cor

g1 = BarChart(); g1.type = 'col'; g1.grouping = 'clustered'
g1.add_data(Reference(pa, min_col=3, max_col=4, min_row=R_ANO_H, max_row=R_ANOF),
            titles_from_data=True)
g1.set_categories(Reference(pa, min_col=1, min_row=R_ANO0, max_row=R_ANOF))
estilo_eixo(g1, 'Vendido × recebido por ano')
pinta(g1.series[0], NAVY2); pinta(g1.series[1], GOLD)
g1.width, g1.height = 15, 8
G_COL, G_ALT = 'I', 17                    # 8 cm ≈ 17 linhas
pa.add_chart(g1, f'{G_COL}{R_KPI + 3}')

g2 = BarChart(); g2.type = 'col'; g2.grouping = 'clustered'
g2.add_data(Reference(pa, min_col=3, max_col=4, min_row=R_MES_H, max_row=R_MESF),
            titles_from_data=True)
g2.set_categories(Reference(pa, min_col=1, min_row=R_MES0, max_row=R_MESF))
estilo_eixo(g2, f'{ANO_ATUAL} · vendas e recebimentos mês a mês')
pinta(g2.series[0], NAVY2); pinta(g2.series[1], GOLD)
g2.width, g2.height = 15, 8
g2l = LineChart()
g2l.add_data(Reference(pa, min_col=7, max_col=8, min_row=R_MES_H, max_row=R_MESF),
             titles_from_data=True)
g2l.y_axis.axId = 200
g2l.y_axis.numFmt = 'R$ #,##0'
for s, cor in zip(g2l.series, (OK, RED)):
    s.graphicalProperties.line.solidFill = cor
    s.graphicalProperties.line.width = 22000
    s.smooth = False
    s.marker.symbol = 'none'
g2l.series[1].graphicalProperties.line.dashStyle = 'dash'
g2 += g2l
pa.add_chart(g2, f'{G_COL}{R_KPI + 3 + G_ALT}')

print_cfg(pa, f'A1:H{R_FIM_PA}')


# ══════════════════════════════════════════════ ABA · Análise & Sazonalidade
an = wb.create_sheet('Análise & Sazonalidade', 1)
an.sheet_view.showGridLines = False
NC_AN = 22
r = faixa_marca(an, NC_AN, 'ANÁLISE & SAZONALIDADE',
                'Fluxo de caixa pela data do pagamento · padrão sazonal de vendas · '
                'concentração da carteira')
for i, w in enumerate([13, 14, 14, 14, 14, 14, 11, 12, 15] + [11] * 13, start=1):
    an.column_dimensions[get_column_letter(i)].width = w
an.row_dimensions[r].height = 6
r += 1

# ── 1 · fluxo de caixa realizado
r = titulo_secao(an, r, NC_AN, 'Fluxo de caixa realizado',
                 'quanto entrou em cada mês, pela DATA DO PAGAMENTO — não pela data do contrato')
R_FX_H = r
r = cab_tabela(an, r, ['Mês'] + ANOS[::-1] + ['Média 23–25', 'Peso no ano'] + [''] * 15,
               None, 26)
R_FX0 = r
R_FX_TOT_PH = R_FX0 + len(MESES) - 1      # última linha de mês
for k, mes in enumerate(MESES, start=1):
    an.row_dimensions[r].height = 18
    c = an.cell(r, 1, mes); c.font = Font(name=F, size=9.5, bold=True, color=NAVY)
    c.fill = fill(CALC); c.alignment = LEFTI; c.border = GRID
    for j, ano in enumerate(ANOS[::-1], start=2):
        c = an.cell(r, j, fluxo_mes(int(ano), k))
        c.number_format = MOEDA0; c.font = Font(name=F, size=9.5, color=NAVY2)
        c.fill = fill(CALC); c.alignment = RIGHT; c.border = GRID
    c = an.cell(r, 6, f'=ROUND(AVERAGE($B{r}:$D{r}),0)')
    c.number_format = MOEDA0; c.font = Font(name=F, size=9.5, bold=True, color=NAVY)
    c.fill = fill(GOLDBG); c.alignment = RIGHT; c.border = GRID
    c = an.cell(r, 7, f'=IF(SUM($F${R_FX0}:$F${R_FX_TOT_PH})=0,"",'
                      f'$F{r}/SUM($F${R_FX0}:$F${R_FX_TOT_PH}))')
    c.number_format = PCT1; c.font = Font(name=F, size=9.5, color=NAVY2)
    c.fill = fill(CALC); c.alignment = CTR; c.border = GRID
    r += 1
R_FXF = r - 1
an.row_dimensions[r].height = 20
c = an.cell(r, 1, 'TOTAL'); c.font = Font(name=F, size=10, bold=True, color=WHITE)
c.fill = fill(NAVY2); c.alignment = LEFTI; c.border = GRID
for j in range(2, 8):
    col = get_column_letter(j)
    c = an.cell(r, j, f'=ROUND(SUM({col}${R_FX0}:{col}${R_FXF}),0)' if j < 7 else None)
    c.number_format = MOEDA0; c.font = Font(name=F, size=10, bold=True, color=GOLDS)
    c.fill = fill(NAVY2); c.alignment = RIGHT; c.border = GRID
R_FX_TOT = r
an.conditional_formatting.add(f'G{R_FX0}:G{R_FXF}', FormulaRule(
    formula=[f'AND($G{R_FX0}<>"",$G{R_FX0}>=0.11)'], fill=fill(OKBG),
    font=Font(bold=True, color=OK)))
an.conditional_formatting.add(f'G{R_FX0}:G{R_FXF}', FormulaRule(
    formula=[f'AND($G{R_FX0}<>"",$G{R_FX0}<=0.05)'], fill=fill(REDBG),
    font=Font(bold=True, color=RED)))
r += 2

# ── 2 · sazonalidade de vendas
r = titulo_secao(an, r, NC_AN, 'Padrão sazonal de vendas',
                 'vendas pela data do contrato · índice = média do mês ÷ média geral')
R_SZ_H = r
r = cab_tabela(an, r, ['Mês'] + ANOS[::-1] + ['Média 23–25', 'Índice', 'Dispersão',
                                              'Classificação'] + [''] * 13, None, 26)
R_SZ0 = r
R_SZ_TOT_PH = R_SZ0 + len(MESES)          # linha do TOTAL, conhecida de antemão
for mes in MESES:
    an.row_dimensions[r].height = 18
    c = an.cell(r, 1, mes); c.font = Font(name=F, size=9.5, bold=True, color=NAVY)
    c.fill = fill(CALC); c.alignment = LEFTI; c.border = GRID
    for j, ano in enumerate(ANOS[::-1], start=2):
        d = DIVISOR[ano][mes]
        c = an.cell(r, j, f"='{ano}'!$C${d}")
        c.number_format = MOEDA0; c.font = Font(name=F, size=9.5, color=NAVY2)
        c.fill = fill(CALC); c.alignment = RIGHT; c.border = GRID
    for col, v, nf, bg, fo in (
            (6, f'=ROUND(AVERAGE($B{r}:$D{r}),0)', MOEDA0, GOLDBG,
             Font(name=F, size=9.5, bold=True, color=NAVY)),
            (7, f'=IF($F${R_SZ_TOT_PH}=0,"",$F{r}/($F${R_SZ_TOT_PH}/12))', '0.00', CALC,
             Font(name=F, size=10, bold=True, color=NAVY2)),
            (8, f'=IF(OR($F{r}=0,COUNT($B{r}:$D{r})<2),"",'
                f'ROUND(STDEV($B{r}:$D{r})/$F{r},2))', '0.00', CALC,
             Font(name=F, size=9, color=MUTED)),
            (9, f'=IF($G{r}="","",IF($H{r}>0.7,"Irregular",'
                f'IF($G{r}>=1.2,"Forte",IF($G{r}<=0.8,"Fraco","Normal"))))', None, CALC,
             Font(name=F, size=9, bold=True, color=NAVY2))):
        c = an.cell(r, col, v); c.number_format = nf or 'General'
        c.font = fo; c.fill = fill(bg); c.alignment = CTR if col > 6 else RIGHT
        c.border = GRID
    r += 1
R_SZF = r - 1
an.row_dimensions[r].height = 20
c = an.cell(r, 1, 'TOTAL / MÉDIA'); c.font = Font(name=F, size=10, bold=True, color=WHITE)
c.fill = fill(NAVY2); c.alignment = LEFTI; c.border = GRID
for j in range(2, 10):
    col = get_column_letter(j)
    v = f'=ROUND(SUM({col}${R_SZ0}:{col}${R_SZF}),0)' if j <= 6 else None
    c = an.cell(r, j, v); c.number_format = MOEDA0
    c.font = Font(name=F, size=10, bold=True, color=GOLDS)
    c.fill = fill(NAVY2); c.alignment = RIGHT; c.border = GRID
R_SZ_TOT = r
for txt, bg, cor in (('Forte', OKBG, OK), ('Fraco', REDBG, RED),
                     ('Irregular', AMBBG, AMBER)):
    an.conditional_formatting.add(f'I{R_SZ0}:I{R_SZF}', FormulaRule(
        formula=[f'$I{R_SZ0}="{txt}"'], fill=fill(bg), font=Font(bold=True, color=cor),
        stopIfTrue=True))
r += 2

# ── 3 · concentração e ticket
r = titulo_secao(an, r, NC_AN, 'Concentração da carteira',
                 'o quanto o ano depende de poucos projetos')
R_CC_H = r
r = cab_tabela(an, r, ['Ano', 'Projetos', 'Vendido', 'Ticket médio', 'Maior projeto',
                       '% do maior', 'Top 3', '% do top 3', 'Leitura'] + [''] * 13,
               None, 26)
R_CC0 = r
for ano in ANOS:
    m = mapa_anos[ano]
    Y = rgn(ano, LP[C_AUX2])
    an.row_dimensions[r].height = 19
    n = f'COUNT({Y})'
    for col, v, nf, al in (
            (1, ano, None, LEFTI), (2, f'={n}', '0', CTR),
            (3, f"='{ano}'!$A$7", MOEDA0, RIGHT),
            (4, f'=IF({n}=0,"",ROUND(\'{ano}\'!$A$7/{n},0))', MOEDA0, RIGHT),
            (5, f'=IF({n}=0,"",MAX({Y}))', MOEDA0, RIGHT),
            (6, f"=IF(OR({n}=0,'{ano}'!$A$7=0),\"\",MAX({Y})/'{ano}'!$A$7)", PCT1, CTR),
            (7, f'=IF({n}<3,"",ROUND(LARGE({Y},1)+LARGE({Y},2)+LARGE({Y},3),0))',
             MOEDA0, RIGHT),
            (8, f"=IF(OR($G{r}=\"\",'{ano}'!$A$7=0),\"\",$G{r}/'{ano}'!$A$7)", PCT1, CTR),
            (9, f'=IF($H{r}="","",IF($H{r}>=0.45,"Muito concentrado",'
                f'IF($H{r}>=0.3,"Concentrado","Distribuído")))', None, CTR)):
        c = an.cell(r, col, v); c.font = Font(name=F, size=9.5, bold=(col in (1, 9)),
                                              color=NAVY if col == 1 else NAVY2)
        c.fill = fill(CALC); c.alignment = al; c.border = GRID
        if nf: c.number_format = nf
    r += 1
R_CCF = r - 1
for txt, bg, cor in (('Muito concentrado', REDBG, RED), ('Concentrado', AMBBG, AMBER),
                     ('Distribuído', OKBG, OK)):
    an.conditional_formatting.add(f'I{R_CC0}:I{R_CCF}', FormulaRule(
        formula=[f'$I{R_CC0}="{txt}"'], fill=fill(bg), font=Font(bold=True, color=cor),
        stopIfTrue=True))
r += 2

# ── 4 · leitura e ressalvas
r = titulo_secao(an, r, NC_AN, 'Leitura do padrão')
MX_IX = f'MAX($G${R_SZ0}:$G${R_SZF})'
MN_IX = f'MIN($G${R_SZ0}:$G${R_SZF})'
MN_DP = f'MIN($H${R_SZ0}:$H${R_SZF})'
MX_PS = f'MAX($G${R_FX0}:$G${R_FXF})'
LEIT = [
    (f'="Maior índice de vendas: "&INDEX($A${R_SZ0}:$A${R_SZF},'
     f'MATCH({MX_IX},$G${R_SZ0}:$G${R_SZF},0))&" ("&TEXT({MX_IX},"0.00")&", classificado como "&'
     f'INDEX($I${R_SZ0}:$I${R_SZF},MATCH({MX_IX},$G${R_SZ0}:$G${R_SZF},0))&'
     f'"). Menor índice: "&INDEX($A${R_SZ0}:$A${R_SZF},'
     f'MATCH({MN_IX},$G${R_SZ0}:$G${R_SZF},0))&" ("&TEXT({MN_IX},"0.00")&")."'),
    (f'="O mês de padrão MAIS CONFIÁVEL é "&INDEX($A${R_SZ0}:$A${R_SZF},'
     f'MATCH({MN_DP},$H${R_SZ0}:$H${R_SZF},0))&" — dispersão de apenas "&'
     f'TEXT({MN_DP},"0.00")&" entre os três anos, com índice "&'
     f'TEXT(INDEX($G${R_SZ0}:$G${R_SZF},MATCH({MN_DP},$H${R_SZ0}:$H${R_SZF},0)),"0.00")&'
     f'". Esse é o único tipo de mês em que o índice vale como referência de planejamento."'),
    (f'="Dos 12 meses, "&COUNTIF($I${R_SZ0}:$I${R_SZF},"Irregular")&'
     f'" estão marcados como Irregular e "&COUNTIF($I${R_SZ0}:$I${R_SZF},"Normal")&'
     f'" como Normal. Índice alto com dispersão alta não é sazonalidade: é um projeto '
     f'grande que caiu naquele mês em um dos anos."'),
    (f'="Caixa: o mês que historicamente mais traz dinheiro é "&'
     f'INDEX($A${R_FX0}:$A${R_FXF},MATCH({MX_PS},$G${R_FX0}:$G${R_FXF},0))&'
     f'", com "&TEXT({MX_PS},"0%")&" do caixa do ano. Repare que o mês de maior VENDA e o '
     f'de maior CAIXA não coincidem — é o descasamento entre vender e receber."'),
]
for txt in LEIT:
    an.row_dimensions[r].height = 30
    bloco(an, r, 1, NC_AN, txt, f=Font(name=F, size=9.5, color='2A3744'), bg=WHITE,
          al=LEFTIW, bd=False)
    r += 1
an.row_dimensions[r].height = 8
r += 1
an.row_dimensions[r].height = 58
bloco(an, r, 1, NC_AN,
      '  Ressalvas honestas sobre a sazonalidade: (1) a base tem três anos fechados — 2023, 2024 e 2025 — e três '
      'pontos por mês é pouco para afirmar padrão; trate o índice como sinal, não como previsão. '
      '(2) A operação começou em março de 2023, então janeiro e fevereiro daquele ano puxam a média para baixo. '
      '(3) Setembro de 2025 sozinho (R$ 572 mil, dois mega-projetos) distorce o mês inteiro — é por isso que a '
      'coluna Dispersão existe: quando ela passa de 0,70, a planilha marca o mês como Irregular. '
      '(4) 2026 aparece nas tabelas mas fica FORA do cálculo da média, porque o ano não fechou.',
      f=Font(name=F, size=8.5, color='41505D'), bg=GOLDBG, al=LEFTIW)
R_FIM_AN = r

# ── gráficos da análise
g3 = LineChart()
g3.add_data(Reference(an, min_col=2, max_col=5, min_row=R_FX_H, max_row=R_FXF),
            titles_from_data=True)
g3.set_categories(Reference(an, min_col=1, min_row=R_FX0, max_row=R_FXF))
estilo_eixo(g3, 'Fluxo de caixa · entradas por mês')
for sr, cor in zip(g3.series, ('9FB0C4', BLUE, GOLD, NAVY2)):
    sr.graphicalProperties.line.solidFill = cor
    sr.graphicalProperties.line.width = 20000
    sr.smooth = False
    sr.marker.symbol = 'circle'
    sr.marker.size = 5
g3.width, g3.height = 17, 9
GA_COL, GA_ALT = 'K', 19                  # 9 cm ≈ 19 linhas
an.add_chart(g3, f'{GA_COL}{R_FX_H}')

g4 = BarChart(); g4.type = 'col'
g4.add_data(Reference(an, min_col=7, min_row=R_SZ_H, max_row=R_SZF),
            titles_from_data=True)
g4.set_categories(Reference(an, min_col=1, min_row=R_SZ0, max_row=R_SZF))
estilo_eixo(g4, 'Índice sazonal de vendas  ·  1,00 = mês médio')
g4.y_axis.numFmt = '0.00'
pinta(g4.series[0], GOLD)
g4.width, g4.height = 17, 9
an.add_chart(g4, f'{GA_COL}{R_FX_H + GA_ALT}')

g5 = BarChart(); g5.type = 'col'; g5.grouping = 'clustered'
g5.add_data(Reference(an, min_col=2, max_col=5, min_row=R_SZ_H, max_row=R_SZF),
            titles_from_data=True)
g5.set_categories(Reference(an, min_col=1, min_row=R_SZ0, max_row=R_SZF))
estilo_eixo(g5, 'Vendas mês a mês, ano a ano')
for sr, cor in zip(g5.series, ('9FB0C4', BLUE, GOLD, NAVY2)):
    pinta(sr, cor)
g5.width, g5.height = 17, 9
an.add_chart(g5, f'{GA_COL}{R_FX_H + 2 * GA_ALT}')

print_cfg(an, f'A1:I{R_FIM_AN}')


# ══════════════════════════════════════════════ ABA · Aporte Walton
orig = openpyxl.load_workbook('/root/.claude/uploads/2544489f-df71-5f40-87c6-89025901a0cf/'
                              '1e190286-Controle_pagamentos_de_projetos.xlsx')
wa = wb.create_sheet('Aporte Walton')
wa.sheet_view.showGridLines = False
r = faixa_marca(wa, 4, 'GESTÃO DO APORTE DO WALTON',
                'Entradas, saídas para caixa e investimentos · saldo calculado')
for i, w in enumerate([16, 40, 4, 18], start=1):
    wa.column_dimensions[get_column_letter(i)].width = w
wa.row_dimensions[r].height = 6
r += 1
ows = orig['walton']
SECOES = [('Entradas', 4, 8), ('Saídas', 11, 23), ('Investimentos', 27, 37)]
refs = {}
for titulo, a, b in SECOES:
    r = titulo_secao(wa, r, 4, titulo)
    i0 = r
    for rr in range(a, b + 1):
        v, d = ows.cell(rr, 1).value, ows.cell(rr, 2).value
        if not isinstance(v, (int, float)) and not d:
            continue
        wa.row_dimensions[r].height = 17
        c = wa.cell(r, 1, v if isinstance(v, (int, float)) else None)
        c.number_format = MOEDA; c.font = Font(name=F, size=10, color=NAVY2)
        c.alignment = RIGHT; c.fill = fill(INPUT); c.border = GRID
        c = wa.cell(r, 2, d); c.font = font(9.5); c.alignment = LEFTI
        c.fill = fill(INPUT); c.border = GRID
        r += 1
    for _ in range(3):                     # linhas de sobra
        wa.row_dimensions[r].height = 17
        for col in (1, 2):
            c = wa.cell(r, col); c.fill = fill(INPUT); c.border = GRID
            if col == 1: c.number_format = MOEDA; c.alignment = RIGHT
        r += 1
    wa.row_dimensions[r].height = 20
    c = wa.cell(r, 1, f'=ROUND(SUM(A{i0}:A{r-1}),2)')
    c.number_format = MOEDA; c.font = Font(name=F, size=11, bold=True, color=WHITE)
    c.fill = fill(NAVY2); c.alignment = RIGHT; c.border = GRID
    c = wa.cell(r, 2, f'Total · {titulo.lower()}')
    c.font = Font(name=F, size=9.5, bold=True, color=GOLDS); c.fill = fill(NAVY2)
    c.alignment = LEFTI; c.border = GRID
    refs[titulo] = r
    r += 2
wa.row_dimensions[r].height = 28
c = wa.cell(r, 1, f'=ROUND(A{refs["Entradas"]}-A{refs["Saídas"]}-A{refs["Investimentos"]},2)')
c.number_format = MOEDA; c.font = Font(name=F, size=14, bold=True, color=GOLDS)
c.fill = fill(NAVY); c.alignment = RIGHT; c.border = GRID
c = wa.cell(r, 2, 'SALDO DO APORTE  ·  entradas − saídas − investimentos')
c.font = Font(name=F, size=10, bold=True, color=WHITE); c.fill = fill(NAVY)
c.alignment = LEFTI; c.border = GRID
for col in (3, 4):
    wa.cell(r, col).fill = fill(NAVY)
print_cfg(wa, f'A1:D{r}', retrato=True)

# ══════════════════════════════════════════════ ABA · Crédito Samuel
cs = wb.create_sheet('Crédito Samuel')
cs.sheet_view.showGridLines = False
r = faixa_marca(cs, 3, 'CONTROLE DE CRÉDITO · SAMUEL', 'Lançamentos e saldo acumulado')
for i, w in enumerate([14, 16, 40], start=1):
    cs.column_dimensions[get_column_letter(i)].width = w
cs.row_dimensions[r].height = 6
r += 1
ocs = orig['credito samuel']
lin = []
for rr in range(4, ocs.max_row + 1):
    d, v = ocs.cell(rr, 1).value, ocs.cell(rr, 2).value
    if isinstance(v, (int, float)):
        if isinstance(d, datetime.datetime) and d.day <= 12:
            d = datetime.date(d.year, d.day, d.month)
        elif isinstance(d, datetime.datetime):
            d = d.date()
        lin.append((d, v))
r = cab_tabela(cs, r, ['Data', 'Valor (R$)', 'Observação'], None, 26)
i0 = r
for d, v in lin + [(None, None)] * 12:
    cs.row_dimensions[r].height = 17
    c = cs.cell(r, 1, d); c.number_format = DATA; c.alignment = CTR
    c.fill = fill(INPUT); c.border = GRID; c.font = font(9.5)
    c = cs.cell(r, 2, v); c.number_format = MOEDA; c.alignment = RIGHT
    c.fill = fill(INPUT); c.border = GRID; c.font = Font(name=F, size=9.5, color=NAVY2)
    c = cs.cell(r, 3); c.fill = fill(INPUT); c.border = GRID
    c.font = font(9); c.alignment = LEFTI
    r += 1
cs.row_dimensions[r].height = 24
c = cs.cell(r, 1, 'TOTAL ACUMULADO'); c.font = Font(name=F, size=10, bold=True, color=WHITE)
c.fill = fill(NAVY2); c.alignment = LEFTI; c.border = GRID
c = cs.cell(r, 2, f'=ROUND(SUM(B{i0}:B{r-1}),2)')
c.number_format = MOEDA; c.font = Font(name=F, size=12, bold=True, color=GOLDS)
c.fill = fill(NAVY2); c.alignment = RIGHT; c.border = GRID
cs.cell(r, 3).fill = fill(NAVY2); cs.cell(r, 3).border = GRID
cs.freeze_panes = f'A{i0}'
print_cfg(cs, f'A1:C{r}', retrato=True)

# ══════════════════════════════════════════════ ABA · Listas
ls = wb.create_sheet('Listas')
ls.sheet_view.showGridLines = False
ls.column_dimensions['A'].width = 22
ls.column_dimensions['B'].width = 20
ls['A1'] = 'Linhas divisoras'; ls['B1'] = 'Forma de pagamento'
for col in ('A1', 'B1'):
    ls[col].font = font(9, True, WHITE); ls[col].fill = fill(NAVY2); ls[col].alignment = CTR
for i, m in enumerate(MESES + [ANTERIOR], start=2):
    c = ls.cell(i, 1, m); c.font = font(10); c.border = BOTTOM2
for i, f_ in enumerate(FORMAS, start=2):
    c = ls.cell(i, 2, f_); c.font = font(10); c.border = BOTTOM2
ls['D1'] = 'Para que servem'
ls['D1'].font = font(9, True, WHITE); ls['D1'].fill = fill(NAVY2); ls['D1'].alignment = CTR
ls.column_dimensions['D'].width = 86
for i, t in enumerate([
    'Coluna A — os rótulos das linhas divisoras. É por esta lista que as abas de ano '
    'sabem quais linhas são de mês e quais são de projeto ao somar o total.',
    'NÃO renomeie os meses nem mude a ordem: o total do ano deixa de fechar.',
    'Coluna B — alimenta os menus de Forma de pagamento e de Descrição. Pode acrescentar '
    'itens no fim da coluna.',
], start=2):
    c = ls.cell(i, 4, t); c.font = font(9, c='41505D')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ls.row_dimensions[i].height = 30

# ══════════════════════════════════════════════ ABA · Notas de migração
nt = wb.create_sheet('Notas de migração')
nt.sheet_view.showGridLines = False
NC_NT = 6
r = faixa_marca(nt, NC_NT, 'NOTAS DE MIGRAÇÃO',
                'O que foi corrigido ao reconstruir a planilha · leia antes de usar')
for i, w in enumerate([5, 30, 26, 26, 26, 12], start=1):
    nt.column_dimensions[get_column_letter(i)].width = w
r += 1

def par(row, texto, f=None, h=None, bg=None):
    nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC_NT)
    c = nt.cell(row, 1, '   ' + texto)
    c.font = f or font(10, c='2A3744')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if bg:
        for k in range(1, NC_NT + 1):
            nt.cell(row, k).fill = fill(bg)
    nt.row_dimensions[row].height = h or 18
    return row + 1

tipos = OC.por_tipo()
n_inv = len(tipos.get('data invertida', []))
n_txt = len(tipos.get('texto virou data', []))
n_fora = len(tipos.get('texto fora de lugar', []))

BLOCOS_NT = [
    ('O PROBLEMA DAS DATAS — a correção mais importante', [
        f'{n_inv} datas estavam com o dia e o mês trocados. Um pagamento de 5 de fevereiro estava gravado como 2 de maio.',
        'A causa: em algum momento o arquivo foi editado por uma ferramenta configurada em inglês, que lê "05/02" como 5 de fevereiro invertido — mês 05, dia 02.',
        'A prova é que a separação é perfeita nas quatro abas: TODAS as 67 células que viraram data tinham dia e mês menores ou iguais a 12 (as únicas que a ferramenta conseguia inverter), e TODAS as 113 que ficaram como texto tinham o primeiro número maior que 12 (as que ela não conseguia). Nenhuma exceção nos dois sentidos.',
        f'Foram corrigidas as {n_inv} datas invertidas e convertidas em data de verdade as {n_txt} que estavam como texto — agora dá para ordenar, filtrar e somar por período.',
        'Uma única não pôde ser recuperada: "23.0" na linha 17 de 2023 (Cristiane). Ficou em branco, com o texto original na Descrição.',
    ]),
    ('O QUE MAIS FOI ARRUMADO', [
        f'{n_fora} células da coluna Data continham na verdade a forma de pagamento ("pix", "material", "permuta"). O texto foi movido para a Descrição.',
        'A coluna que se chamava "Status atual" era, na verdade, o saldo a receber. Passou a se chamar Saldo a receber, e ganhou ao lado o % recebido e uma Situação (Quitado, Parcial, A receber, Recebido a mais) com cor.',
        'A fórmula do saldo era =SUM(C10-H10-L10-...), que devolvia o valor do contrato mesmo em linha vazia. Agora fica em branco quando não há projeto.',
        'O total do ano era uma lista fixa de células (=SUM(C9,C13,C20...)) que passava a somar errado assim que alguém inserisse uma linha. Agora é calculado sobre todas as linhas de projeto, e existe uma CONFERÊNCIA no topo que avisa se alguma linha ficou fora de um bloco de mês.',
        'A "Meta mensal" era a meta anual dividida por 7 — um número que mudava sozinho. Saiu. No lugar entraram % da meta e quanto falta para a meta.',
        'Colunas vazias de separação foram removidas: a planilha foi de 33 para 23 colunas, sem perder nada.',
        'Cada bloco de mês ganhou linhas de sobra em branco, para lançar projeto novo sem precisar inserir linha.',
    ]),
    ('DIVERGÊNCIAS QUE VOCÊ PRECISA RESOLVER', [
        'Monte Negro (2024, casa completa, contrato de R$ 100.000): existem 7 pagamentos lançados em linhas soltas somando R$ 97.800, mais R$ 22.200 no segundo bloco — total de R$ 120.000 contra um contrato de R$ 100.000.',
        'A fórmula original somava só as 5 primeiras linhas (=SUM(H33:H37) = R$ 77.800), o que fechava exatamente os R$ 100.000. Ou seja: dois pagamentos de R$ 10.000 foram lançados depois e nunca entraram na conta.',
        'Trouxe os sete pagamentos, porque apagar dinheiro recebido é pior do que mostrar a inconsistência. O projeto aparece como "Recebido a mais" e o ano de 2024 fecha com saldo negativo de R$ 20.000. Decida: o contrato subiu, ou esses lançamentos são de outro projeto?',
        'Os 12 projetos que estavam no topo da aba 2023, antes de Janeiro, são contratos de 2022 ainda em cobrança. Ficaram num bloco chamado "Anterior (2022)" e continuam FORA do total de 2023, como já era no original.',
    ]),
    ('O QUE NÃO MUDOU', [
        'Um projeto por linha, os pagamentos à direita, uma aba por ano, cinco pagamentos por projeto — igual ao que vocês já usam.',
        'Nenhum valor de contrato ou de pagamento foi alterado. O que mudou foi data, posição de texto e fórmula.',
        'Cliente, Projeto, Investimento e Forma de pagamento continuam sendo digitados à mão, nas mesmas quatro primeiras colunas.',
    ]),
]
for titulo, linhas in BLOCOS_NT:
    r = titulo_secao(nt, r, NC_NT, titulo)
    nt.row_dimensions[r].height = 4
    r += 1
    for t in linhas:
        r = par(r, t, h=44 if len(t) > 230 else (30 if len(t) > 115 else 20))
    nt.row_dimensions[r].height = 8
    r += 1
r = par(r, 'O arquivo original foi preservado. Esta planilha é gerada por '
           'gerar-controle-pagamentos.py — edite o script, nunca o .xlsx.',
        f=Font(name=F, size=9, bold=True, color='7A5B17'), h=26, bg=GOLDBG)
print_cfg(nt, f'A1:F{r-1}', retrato=True)

wb.active = 0
SAIDA = '/home/user/valvicorcamentista/painel/planilhas/Valvic_Controle_Pagamentos.xlsx'
wb.save(SAIDA)
print('OK →', SAIDA)
for aba in ANOS:
    m = mapa_anos[aba]
    print(f'  {aba}: linhas {m["ini"]}–{m["fim"]} · {len(m["divisores"])} blocos')
