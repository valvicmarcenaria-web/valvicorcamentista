#!/usr/bin/env python3
"""
Gera Valvic_Relacionamento_Parceiras.xlsx — o registro de relacionamento com as
parceiras decoradoras e arquitetas, e o controle da campanha de cesta de café.

    python3 gerar-relacionamento-parceiras.py

EDITAR SEMPRE ESTE SCRIPT, NUNCA O .xlsx — a próxima geração apaga a mudança feita
direto no arquivo.

Por que a planilha existe: a etapa que mata campanha de relacionamento é o registro.
Mandar duas vezes para a mesma parceira e nunca para outra é pior do que não ter
mandado. Aqui fica quem recebeu o quê, quando, e qual é o próximo contato.

Abas:
  Parceiras       cadastro fixo — a base do ano inteiro
  Campanha Café   o controle da entrega desta campanha, uma linha por parceira
  Cotação Cesta   comparativo dos 3 fornecedores (regra da casa acima de R$ 500)
  Régua do Ano    o planejamento bimestral de contato, parceira a parceira
  Listas          o que alimenta os menus suspensos + como usar
"""
import importlib.util
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill

# O utilitário da casa se chama gerar-planilha.py (com hífen), então entra por importlib.
FERRAMENTA = (Path(__file__).resolve().parents[2]
              / '.claude/skills/alice-assistente-operacional/ferramentas/gerar-planilha.py')
_spec = importlib.util.spec_from_file_location('gerar_planilha', FERRAMENTA)
gp = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(gp)

BORDA, GOLDBG, NAVY2 = gp.BORDA, gp.GOLDBG, gp.NAVY2
cabecalho, largura, menu = gp.cabecalho, gp.largura, gp.menu
nova, salvar, titulo_secao, zebra = gp.nova, gp.salvar, gp.titulo_secao, gp.zebra

SAIDA = Path(__file__).resolve().parent / 'Valvic_Relacionamento_Parceiras.xlsx'

LINHAS = 30          # parceiras cadastráveis
CREME = PatternFill('solid', fgColor='FFFDF6')

SEGMENTO = ['Arquiteta', 'Decoradora', 'Designer de interiores', 'Escritório']
SITUACAO = ['Ativa', 'Morna', 'Fria', 'Nova']
PREFERENCIA = ['Café', 'Café sem lactose', 'Chá', 'Sem glúten', 'Não bebe café', 'A confirmar']
TIPO_CONTATO = ['Cesta/brinde', 'Visita à fábrica', 'Entrega de obra', 'Mensagem',
                'Data comemorativa', 'Reunião', 'Projeto em andamento']
SIM_NAO = ['Sim', 'Não', 'Pendente']
QUEM_ENTREGA = ['Karla', 'Jonathan', 'Deivison', 'Douglas', 'Motoboy', 'Fornecedor']


def nota(ws, linha, coluna, texto, ate_coluna=None):
    """Linha de instrução em itálico cinza, mesclada."""
    if ate_coluna:
        ws.merge_cells(start_row=linha, start_column=coluna,
                       end_row=linha, end_column=ate_coluna)
    c = ws.cell(row=linha, column=coluna, value=texto)
    c.font = Font(name='Calibri', size=9, italic=True, color='6C7785')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[linha].height = 26


def preencher(ws, primeira, ultima, colunas):
    """Marca de creme as colunas que a Karla digita."""
    for r in range(primeira, ultima + 1):
        for c in colunas:
            cel = ws.cell(row=r, column=c)
            cel.fill = CREME
            cel.border = BORDA
            cel.alignment = Alignment(vertical='center', wrap_text=True)


# ---------------------------------------------------------------- aba 1
def aba_parceiras(wb):
    ws = wb.active
    titulo_secao(ws, 'PARCEIRAS — cadastro e relacionamento · Valvic Marcenaria', 1, 15)
    nota(ws, 2, 1, 'Uma linha por parceira. O que manda a régua é a coluna PRÓXIMO CONTATO: '
                   'toda segunda-feira, filtre por ela e veja quem vence na semana.', 15)
    cols = ['Nº', 'Parceira', 'Escritório / estúdio', 'Celular', 'Instagram', 'Segmento',
            'Situação', 'Endereço de entrega', 'Melhor dia e horário', 'Preferência',
            'Último contato', 'O que foi', 'Próximo contato', 'Projetos indicados',
            'Observação']
    cabecalho(ws, cols, linha=3)
    primeira, ultima = 4, 3 + LINHAS
    for i, r in enumerate(range(primeira, ultima + 1), start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 30
    preencher(ws, primeira, ultima, range(2, 16))
    menu(ws, SEGMENTO, f'F{primeira}:F{ultima}')
    menu(ws, SITUACAO, f'G{primeira}:G{ultima}')
    menu(ws, PREFERENCIA, f'J{primeira}:J{ultima}')
    menu(ws, TIPO_CONTATO, f'L{primeira}:L{ultima}')
    largura(ws, [5, 24, 22, 16, 18, 16, 12, 34, 18, 18, 13, 20, 14, 12, 32])
    zebra(ws, primeira, ultima, 15)
    return ws


# ---------------------------------------------------------------- aba 2
def aba_campanha(wb):
    ws = wb.create_sheet('Campanha Café')
    titulo_secao(ws, 'CAMPANHA CESTA DE CAFÉ — setembro/2026 · entrega na sexta', 1, 12)
    nota(ws, 2, 1, 'Só existe campanha depois desta aba preenchida. Marque a entrega no '
                   'MESMO dia — registro feito de memória na segunda-feira já nasce errado.', 12)
    cols = ['Nº', 'Parceira', 'Endereço confirmado', 'Cesta separada', 'Cartão escrito',
            'Data prevista', 'Quem entrega', 'Data da entrega', 'Recebeu em mãos',
            'Respondeu', 'O que ela disse', 'Próxima ação e data']
    cabecalho(ws, cols, linha=3)
    primeira, ultima = 4, 9          # as 6 parceiras
    for i, r in enumerate(range(primeira, ultima + 1), start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 30
    preencher(ws, primeira, ultima, range(2, 13))
    for col in ('C', 'D', 'E', 'I', 'J'):
        menu(ws, SIM_NAO, f'{col}{primeira}:{col}{ultima}')
    menu(ws, QUEM_ENTREGA, f'G{primeira}:G{ultima}')
    largura(ws, [5, 24, 18, 15, 15, 14, 15, 15, 15, 12, 38, 26])
    zebra(ws, primeira, ultima, 12)

    l = ultima + 2
    titulo_secao(ws, 'O QUE SE MEDE EM 30 DIAS (medir em 11/10/2026)', l, 12)
    medidas = [
        ('Cestas entregues', '6 de 6 — se faltar uma, a campanha não fechou'),
        ('Parceiras que responderam', 'meta: pelo menos 4'),
        ('Visitas à fábrica agendadas', 'no contato seguinte, não neste'),
        ('Projetos ou contatos novos vindos das 6', 'medida real da campanha'),
    ]
    cabecalho(ws, ['Indicador', 'Meta / leitura', 'Resultado'], linha=l + 1)
    for i, (ind, met) in enumerate(medidas):
        r = l + 2 + i
        ws.cell(row=r, column=1, value=ind).alignment = Alignment(vertical='center')
        ws.cell(row=r, column=2, value=met).alignment = Alignment(vertical='center',
                                                                  wrap_text=True)
        ws.cell(row=r, column=3).fill = CREME
        ws.cell(row=r, column=3).border = BORDA
        ws.row_dimensions[r].height = 22
    ws.freeze_panes = 'A4'      # cabecalho() congela onde é chamado; vale o de cima
    return ws


# ---------------------------------------------------------------- aba 3
def aba_cotacao(wb):
    ws = wb.create_sheet('Cotação Cesta')
    titulo_secao(ws, 'COTAÇÃO DA CESTA — 3 fornecedores, regra da casa acima de R$ 500', 1, 8)
    nota(ws, 2, 1, 'Cotar 3 e documentar. O total das 6 cestas vai no pedido de autorização '
                   'ao Jonathan — compra nenhuma sai sem o "pode" dele por escrito.', 8)
    cols = ['Fornecedor', 'Contato', 'O que vem na cesta', 'Valor unitário',
            'Total 6 cestas', 'Frete / entrega', 'Prazo', 'Escolhido']
    cabecalho(ws, cols, linha=3)
    primeira, ultima = 4, 6
    preencher(ws, primeira, ultima, range(1, 9))
    for r in range(primeira, ultima + 1):
        ws.row_dimensions[r].height = 34
        ws.cell(row=r, column=5, value=f'=D{r}*6').number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=4).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=6).number_format = 'R$ #,##0.00'
    menu(ws, SIM_NAO, f'H{primeira}:H{ultima}')
    largura(ws, [26, 20, 42, 16, 16, 16, 14, 12])

    l = ultima + 2
    ws.cell(row=l, column=1, value='Faixa de bom gosto: cesta pequena e bem apresentada. '
            'Caro constrange e cria dívida — o gesto é que comunica, não o preço.').font = \
        Font(name='Calibri', size=9, italic=True, color='6C7785')
    ws.merge_cells(start_row=l, start_column=1, end_row=l, end_column=8)
    return ws


# ---------------------------------------------------------------- aba 4
def aba_regua(wb):
    ws = wb.create_sheet('Régua do Ano')
    titulo_secao(ws, 'RÉGUA DO ANO — um contato a cada 2 meses, sem pedir nada', 1, 8)
    nota(ws, 2, 1, 'Escreva em cada bimestre O QUE foi ou será o contato. Bimestre em branco '
                   'na hora de olhar = parceira que esfriou sem ninguém perceber.', 8)
    cols = ['Nº', 'Parceira', 'Jan-Fev', 'Mar-Abr', 'Mai-Jun', 'Jul-Ago', 'Set-Out', 'Nov-Dez']
    cabecalho(ws, cols, linha=3)
    primeira, ultima = 4, 9
    for i, r in enumerate(range(primeira, ultima + 1), start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 30
    preencher(ws, primeira, ultima, range(2, 9))
    largura(ws, [5, 24, 22, 22, 22, 22, 22, 22])
    zebra(ws, primeira, ultima, 8)

    l = ultima + 2
    titulo_secao(ws, 'O QUE ENTRA EM CADA BIMESTRE — esqueleto, não obrigação', l, 8)
    plano = [
        ('Jan-Fev', 'Retomada do ano — mensagem curta, sem pedir nada'),
        ('Mar-Abr', 'Dia da Mulher (08/03) — grande parte das parceiras'),
        ('Mai-Jun', 'Fotos das obras do 1º trimestre que ela indicou'),
        ('Jul-Ago', 'Convite para conhecer a fábrica — o que mais esquenta parceria fria'),
        ('Set-Out', 'Onda de café / brinde de relacionamento'),
        ('Nov-Dez', 'Dia do Arquiteto (15/12) e fim de ano'),
    ]
    cabecalho(ws, ['Bimestre', 'Contato padrão'], linha=l + 1)
    for i, (b, t) in enumerate(plano):
        r = l + 2 + i
        ws.cell(row=r, column=1, value=b)
        ws.cell(row=r, column=2, value=t).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 20
    ws.freeze_panes = 'A4'
    return ws


# ---------------------------------------------------------------- aba 5
def aba_listas(wb):
    ws = wb.create_sheet('Listas')
    titulo_secao(ws, 'LISTAS E COMO USAR', 1, 6)
    blocos = [('Segmento', SEGMENTO), ('Situação', SITUACAO), ('Preferência', PREFERENCIA),
              ('Tipo de contato', TIPO_CONTATO), ('Sim / Não', SIM_NAO),
              ('Quem entrega', QUEM_ENTREGA)]
    cabecalho(ws, [b[0] for b in blocos], linha=3)
    for ci, (_, itens) in enumerate(blocos, start=1):
        for ri, item in enumerate(itens, start=4):
            ws.cell(row=ri, column=ci, value=item)
    largura(ws, [22, 14, 20, 22, 12, 16])

    l = 4 + max(len(b[1]) for b in blocos) + 1
    titulo_secao(ws, 'AS 3 REGRAS DA CAMPANHA', l, 6)
    regras = [
        'Presente e pedido não andam juntos. Nada de indicação, RT ou orçamento na mesma '
        'mensagem do café — presente com pedido junto vira cobrança.',
        'RT, comissão e desconto são assunto do Jonathan. A Karla cuida da relação; '
        'a conversa de dinheiro é dele.',
        'Registrar no mesmo dia. Campanha sem registro repete parceira e esquece outra.',
    ]
    for i, txt in enumerate(regras):
        r = l + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=f'{i + 1}. {txt}')
        c.alignment = Alignment(wrap_text=True, vertical='center')
        c.fill = PatternFill('solid', fgColor=GOLDBG)
        c.font = Font(name='Calibri', size=10, color=NAVY2)
        ws.row_dimensions[r].height = 32
    return ws


def main():
    wb = nova('Parceiras')
    aba_parceiras(wb)
    aba_campanha(wb)
    aba_cotacao(wb)
    aba_regua(wb)
    aba_listas(wb)
    salvar(wb, str(SAIDA))


if __name__ == '__main__':
    main()
